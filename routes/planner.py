from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response, current_app
from extensions import db
from models import *
from utils import *
import utils
from services.mail_service import MailService
from utils import _vehicle_payload

# Explicitly map underscore-prefixed functions from utils (they are not imported by *)
_parse_files = utils._parse_files
_dump_files = utils._dump_files
_subproject_allowed_for_project = utils._subproject_allowed_for_project
_effective_main_project_id_for_subprojects = utils._effective_main_project_id_for_subprojects
_normalize_kanban_status = utils._normalize_kanban_status
_promote_job_kanban_status = utils._promote_job_kanban_status
_set_job_kanban_status = utils._set_job_kanban_status
_cell_has_meaningful_job = utils._cell_has_meaningful_job
_effective_team_name_for_cell = utils._effective_team_name_for_cell
_sync_job_from_cell = utils._sync_job_from_cell
_publish_cell = utils._publish_cell
upsert_jobs_for_range = utils.upsert_jobs_for_range
_sqlite_db_path = utils._sqlite_db_path
_create_sqlite_backup_file = utils._create_sqlite_backup_file
_csrf_verify = utils._csrf_verify
_rate_limit = utils._rate_limit
_user_is_admin_or_planner = utils._user_is_admin_or_planner
_save_feedback_uploads = utils._save_feedback_uploads

from datetime import date, datetime, timedelta
from sqlalchemy import or_, and_, desc, func, case
import json
import io
import os
import math
import re
import time as _time
import colorsys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
from werkzeug.utils import secure_filename
import uuid

planner_bp = Blueprint('planner', __name__)

@planner_bp.route("/")
def index():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    u = get_current_user()
    if u and (u.role or "").strip().lower() == "field":
        return redirect(url_for('planner.portal_home'))
    return redirect(url_for('planner.plan_week'))


@planner_bp.get("/admin/db/backup") 
@login_required 
@planner_or_admin_required 
def admin_db_backup_download(): 
    src = _sqlite_db_path()
    if not src:
        return jsonify({"ok": False, "error": "Sadece SQLite icin destekleniyor."}), 400
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backups_dir = os.path.join(current_app.instance_path, "backups")
    os.makedirs(backups_dir, exist_ok=True)
    dest = os.path.join(backups_dir, f"backup_{stamp}.db")
    if not _create_sqlite_backup_file(dest): 
        return jsonify({"ok": False, "error": "Backup olusturulamadi."}), 500 
    
    # Doğrulama: Yedek dosyasının geçerli olduğunu kontrol et
    try:
        import sqlite3
        verify_conn = sqlite3.connect(dest)
        tables = verify_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        table_count = len(tables)
        verify_conn.close()
        
        if table_count == 0:
            try:
                os.remove(dest)
            except:
                pass
            return jsonify({"ok": False, "error": "Yedek dosyası geçersiz."}), 500
    except Exception as e:
        try:
            os.remove(dest)
        except:
            pass
        return jsonify({"ok": False, "error": f"Yedek doğrulama hatası: {str(e)}"}), 500
    
    return send_file(dest, as_attachment=True, download_name=f"planner_backup_{stamp}.db", mimetype="application/octet-stream")


@planner_bp.get("/admin/db/view")
@login_required
@kivanc_required
def admin_db_view():
    """Veritabanındaki tüm tabloları ve verileri görüntüle (sadece Kıvanc)"""
    src = _sqlite_db_path()
    if not src:
        flash("Sadece SQLite veritabanları görüntülenebilir.", "danger")
        return redirect(url_for("admin_users"))
    
    tables = {}
    try:
        import sqlite3
        conn = sqlite3.connect(src)
        conn.row_factory = sqlite3.Row  # Row factory for dict-like access
        
        # Tüm tabloları al
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        table_names = [row[0] for row in cursor.fetchall()]
        
        for table_name in table_names:
            try:
                # Tablo sütunlarını al
                cursor = conn.execute(f"PRAGMA table_info({table_name})")
                columns = [row[1] for row in cursor.fetchall()]
                
                # Tablo verilerini al (limit 1000 kayıt)
                cursor = conn.execute(f"SELECT * FROM {table_name} LIMIT 1000")
                rows = []
                for row in cursor.fetchall():
                    row_dict = {}
                    for i, col_name in enumerate(columns):
                        row_dict[col_name] = row[i]
                    rows.append(row_dict)
                
                tables[table_name] = {
                    "columns": columns,
                    "rows": rows
                }
            except Exception as e:
                # Tablo okunamazsa atla
                tables[table_name] = {
                    "columns": [],
                    "rows": [],
                    "error": str(e)
                }
        
        conn.close()
    except Exception as e:
        flash(f"Veritabanı okunurken hata oluştu: {str(e)}", "danger")
        return redirect(url_for("admin_users"))
    
    return render_template("admin_db_view.html", tables=tables)


@planner_bp.post("/admin/db/restore")
@login_required
@planner_or_admin_required
def admin_db_restore():
    """Veritabanını yedek dosyadan geri yükle"""
    user = get_current_user()
    if not user or not bool(getattr(user, "is_admin", False)):
        return jsonify({"ok": False, "error": "Sadece admin kullanıcılar veritabanını geri yükleyebilir."}), 403
    
    src = _sqlite_db_path()
    if not src:
        return jsonify({"ok": False, "error": "Sadece SQLite icin destekleniyor."}), 400
    
    # Dosya kontrolü
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "Dosya yüklenmedi."}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"ok": False, "error": "Dosya seçilmedi."}), 400
    
    # Dosya uzantısı kontrolü
    if not file.filename.lower().endswith('.db'):
        return jsonify({"ok": False, "error": "Sadece .db dosyaları yüklenebilir."}), 400
    
    # CSRF token kontrolü
    csrf_token = request.form.get('csrf_token', '')
    if not _csrf_verify(csrf_token):
        return jsonify({"ok": False, "error": "CSRF token hatası."}), 403
    
    try:
        # Mevcut veritabanını yedekle (güvenlik için)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backups_dir = os.path.join(current_app.instance_path, "backups")
        os.makedirs(backups_dir, exist_ok=True)
        pre_restore_backup = os.path.join(backups_dir, f"pre_restore_backup_{stamp}.db")
        
        if not _create_sqlite_backup_file(pre_restore_backup):
            return jsonify({"ok": False, "error": "Mevcut veritabanı yedeklenemedi. Restore iptal edildi."}), 500
        
        # Yüklenen dosyayı geçici olarak kaydet
        import tempfile
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"restore_{stamp}.db")
        file.save(temp_file)
        
        # Dosyanın geçerli bir SQLite dosyası olduğunu kontrol et ve içeriğini doğrula
        try:
            import sqlite3
            test_conn = sqlite3.connect(temp_file)
            try:
                # Temel SQLite kontrolü
                test_conn.execute("SELECT 1")
                
                # Tablo sayısını kontrol et
                tables = test_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
                table_count = len(tables)
                
                if table_count == 0:
                    test_conn.close()
                    os.remove(temp_file)
                    return jsonify({"ok": False, "error": "Yüklenen dosya boş bir veritabanı içeriyor."}), 400
                
                # Dosya boyutunu kontrol et (çok küçükse şüpheli)
                file_size = os.path.getsize(temp_file)
                if file_size < 1024:  # 1KB'dan küçükse
                    test_conn.close()
                    os.remove(temp_file)
                    return jsonify({"ok": False, "error": "Yüklenen dosya çok küçük, geçerli bir veritabanı olmayabilir."}), 400
                    
            finally:
                test_conn.close()
        except Exception as e:
            try:
                os.remove(temp_file)
            except:
                pass
            return jsonify({"ok": False, "error": f"Geçersiz SQLite dosyası: {str(e)}"}), 400
        
        # Mevcut veritabanı bilgilerini kaydet (doğrulama için)
        import sqlite3
        old_conn = sqlite3.connect(src)
        old_tables = old_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        old_table_count = len(old_tables)
        old_conn.close()
        
        # Tüm bağlantıları kapat
        db.session.close()
        db.engine.dispose()
        
        # Mevcut veritabanını yedek dosya ile değiştir (byte-byte kopyalama - tamamen aynı dosya)
        import shutil
        shutil.copy2(temp_file, src)
        
        # Doğrulama: Yeni veritabanının içeriğini kontrol et
        verify_conn = sqlite3.connect(src)
        try:
            new_tables = verify_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
            new_table_count = len(new_tables)
            
            # Tablo sayısı kontrolü (en azından bazı tablolar olmalı)
            if new_table_count == 0:
                # Eğer restore başarısız olduysa, yedeği geri yükle
                if os.path.exists(pre_restore_backup):
                    shutil.copy2(pre_restore_backup, src)
                verify_conn.close()
                try:
                    os.remove(temp_file)
                except:
                    pass
                return jsonify({"ok": False, "error": "Geri yüklenen veritabanı geçersiz görünüyor. Restore iptal edildi ve önceki veritabanı geri yüklendi."}), 500
        finally:
            verify_conn.close()
        
        # Geçici dosyayı sil
        try:
            os.remove(temp_file)
        except Exception:
            pass
        
        # Veritabanı bağlantısını yeniden başlat
        db.create_all()
        
        return jsonify({
            "ok": True,
            "message": f"Veritabanı başarıyla geri yüklendi. ({new_table_count} tablo yüklendi)",
            "pre_restore_backup": os.path.basename(pre_restore_backup),
            "table_count": new_table_count
        })
        
    except Exception as e:
        # Hata durumunda önceki yedeği geri yükle (eğer varsa)
        if 'pre_restore_backup' in locals() and os.path.exists(pre_restore_backup):
            try:
                import shutil
                shutil.copy2(pre_restore_backup, src)
            except Exception:
                pass
        
        return jsonify({"ok": False, "error": f"Restore hatası: {str(e)}"}), 500 
 
 
@planner_bp.post("/admin/publish/week/preview") 
@login_required 
@planner_or_admin_required 
def admin_publish_week_preview(): 
    data = request.get_json(force=True, silent=True) or {} 
    token = str(data.get("csrf_token") or "") 
    if not _csrf_verify(token): 
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400 
 
    ws = parse_date(str(data.get("week_start") or "").strip()) 
    if not ws: 
        return jsonify({"ok": False, "error": "week_start gecersiz"}), 400 
 
    start = week_start(ws) 
    end = start + timedelta(days=6) 
 
    cells = PlanCell.query.filter(PlanCell.work_date >= start, PlanCell.work_date <= end).all() 
    publishable = [c for c in cells if _cell_has_meaningful_job(c)] 
 
    project_ids = {int(c.project_id) for c in publishable} 
    projects_by_id: Dict[int, Project] = {} 
    if project_ids: 
        for p in Project.query.filter(Project.id.in_(project_ids)).all(): 
            projects_by_id[int(p.id)] = p 
 
    subproject_ids = {int(c.subproject_id) for c in publishable if c.subproject_id} 
    subprojects_by_id: Dict[int, SubProject] = {} 
    if subproject_ids: 
        for sp in SubProject.query.filter(SubProject.id.in_(subproject_ids)).all(): 
            subprojects_by_id[int(sp.id)] = sp 
 
    cities = sorted( 
        { 
            (projects_by_id[c.project_id].region or "").strip() 
            for c in publishable 
            if c.project_id in projects_by_id 
            and (projects_by_id[c.project_id].region or "").strip() 
            and projects_by_id[c.project_id].region != "-" 
        } 
    ) 
 
    dist: Dict[Tuple[int, int], int] = {} 
    for c in publishable: 
        pid = int(c.project_id) 
        spid = int(c.subproject_id or 0) 
        dist[(pid, spid)] = dist.get((pid, spid), 0) + 1 
 
    dist_rows = [] 
    for (pid, spid), cnt in sorted(dist.items(), key=lambda x: (-x[1], x[0][0], x[0][1])): 
        p = projects_by_id.get(pid) 
        if not p: 
            continue 
        sp = subprojects_by_id.get(spid) if spid else None 
        dist_rows.append( 
            { 
                "project_id": int(pid), 
                "city": p.region, 
                "project_code": p.project_code, 
                "project_name": p.project_name, 
                "subproject_id": int(spid), 
                "subproject_name": (sp.name if sp else ""), 
                "count": int(cnt), 
            } 
        ) 
 
    max_rows = 25 
    return jsonify( 
        { 
            "ok": True, 
            "week_start": iso(start), 
            "week_end": iso(end), 
            "total_jobs": int(len(publishable)), 
            "city_count": int(len(cities)), 
            "cities": cities, 
            "distribution": dist_rows[:max_rows], 
            "distribution_total": int(len(dist_rows)), 
        } 
    ) 
 
 
@planner_bp.post("/admin/publish/cell/preview") 
@login_required 
@planner_or_admin_required 
def admin_publish_cell_preview(): 
    data = request.get_json(force=True, silent=True) or {} 
    token = str(data.get("csrf_token") or "") 
    if not _csrf_verify(token): 
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400 
 
    try: 
        project_id = int(data.get("project_id", 0) or 0) 
    except Exception: 
        project_id = 0 
    d = parse_date(str(data.get("work_date") or "").strip()) 
    if not project_id or not d: 
        return jsonify({"ok": False, "error": "project_id/work_date gecersiz"}), 400 
 
    cell = PlanCell.query.filter_by(project_id=project_id, work_date=d).first() 
    if not cell or not _cell_has_meaningful_job(cell): 
        return jsonify({"ok": False, "error": "Yayinlanacak is bulunamadi"}), 404 
 
    project = Project.query.get(project_id) 
    if not project: 
        return jsonify({"ok": False, "error": "Proje bulunamadi"}), 404 
 
    people = [ 
        (name or "").strip() 
        for (name,) in ( 
            db.session.query(Person.full_name) 
            .join(CellAssignment, CellAssignment.person_id == Person.id) 
            .filter(CellAssignment.cell_id == cell.id) 
            .order_by(Person.full_name.asc()) 
            .all() 
        ) 
        if (name or "").strip() 
    ] 
 
    sp = cell.subproject if getattr(cell, "subproject_id", None) else None 
    return jsonify( 
        { 
            "ok": True, 
            "work_date": iso(d), 
            "project": { 
                "id": int(project.id), 
                "city": project.region, 
                "project_code": project.project_code, 
                "project_name": project.project_name, 
                "responsible": project.responsible, 
            }, 
            "subproject": ( 
                {"id": int(sp.id), "name": sp.name, "code": sp.code} if sp else {"id": 0, "name": "", "code": ""} 
            ), 
            "cell": { 
                "shift": (cell.shift or ""), 
                "vehicle_info": (cell.vehicle_info or ""), 
                "note": (cell.note or ""), 
                "important_note": (getattr(cell, "important_note", None) or ""), 
                "team_id": int(cell.team_id or 0), 
                "team_name": _effective_team_name_for_cell(cell), 
            }, 
            "people": people, 
        } 
    ) 
 
 
@planner_bp.post("/admin/publish/team_week/preview") 
@login_required 
@planner_or_admin_required 
def admin_publish_team_week_preview(): 
    data = request.get_json(force=True, silent=True) or {} 
    token = str(data.get("csrf_token") or "") 
    if not _csrf_verify(token): 
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400 
 
    ws = parse_date(str(data.get("week_start") or "").strip()) 
    if not ws: 
        return jsonify({"ok": False, "error": "week_start gecersiz"}), 400 
 
    try: 
        team_id = int(data.get("team_id", 0) or 0) 
    except Exception: 
        team_id = 0 
    if team_id <= 0: 
        return jsonify({"ok": False, "error": "team_id gecersiz"}), 400 
 
    start = week_start(ws) 
    end = start + timedelta(days=6) 
 
    team = Team.query.get(team_id) 
 
    cells = ( 
        PlanCell.query 
        .filter( 
            PlanCell.work_date >= start, 
            PlanCell.work_date <= end, 
            PlanCell.team_id == team_id, 
        ) 
        .all() 
    ) 
    publishable = [c for c in cells if _cell_has_meaningful_job(c)] 
 
    counts: Dict[str, int] = {iso(start + timedelta(days=i)): 0 for i in range(7)} 
    for c in publishable: 
        k = iso(c.work_date) 
        if k in counts: 
            counts[k] += 1 
 
    day_counts = [{"date": iso(start + timedelta(days=i)), "count": int(counts[iso(start + timedelta(days=i))])} for i in range(7)] 
    return jsonify( 
        { 
            "ok": True, 
            "team_id": int(team_id), 
            "team_name": (team.name if team else ""), 
            "week_start": iso(start), 
            "week_end": iso(end), 
            "total_jobs": int(len(publishable)), 
            "day_counts": day_counts, 
        } 
    ) 
 
 
@planner_bp.post("/admin/publish/week") 
@login_required 
@planner_or_admin_required 
def admin_publish_week(): 
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400

    ws = parse_date(str(data.get("week_start") or "").strip())
    if not ws:
        return jsonify({"ok": False, "error": "week_start gecersiz"}), 400

    start = week_start(ws)
    end = start + timedelta(days=6)
    publisher = get_current_user()
    now = datetime.now()

    cells = PlanCell.query.filter(PlanCell.work_date >= start, PlanCell.work_date <= end).all()
    published = 0
    for cell in cells:
        if not _cell_has_meaningful_job(cell):
            continue
        _publish_cell(cell, publisher=publisher, now=now)
        published += 1

    db.session.commit()
    return jsonify({"ok": True, "published": int(published), "week_start": iso(start), "week_end": iso(end)})


@planner_bp.post("/admin/publish/cell")
@login_required
@planner_or_admin_required
def admin_publish_cell():
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400

    try:
        project_id = int(data.get("project_id", 0) or 0)
    except Exception:
        project_id = 0
    d = parse_date(str(data.get("work_date") or "").strip())
    if not project_id or not d:
        return jsonify({"ok": False, "error": "project_id/work_date gecersiz"}), 400

    cell = PlanCell.query.filter_by(project_id=project_id, work_date=d).first()
    if not cell or not _cell_has_meaningful_job(cell):
        return jsonify({"ok": False, "error": "Yayinlanacak is bulunamadi"}), 404

    publisher = get_current_user()
    now = datetime.now()
    job = _publish_cell(cell, publisher=publisher, now=now)
    db.session.commit()
    return jsonify({"ok": True, "job_id": int(job.id), "published_at": job.published_at.isoformat() if job.published_at else None})


@planner_bp.post("/admin/publish/team_week")
@login_required
@planner_or_admin_required
def admin_publish_team_week():
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400

    ws = parse_date(str(data.get("week_start") or "").strip())
    if not ws:
        return jsonify({"ok": False, "error": "week_start gecersiz"}), 400

    try:
        team_id = int(data.get("team_id", 0) or 0)
    except Exception:
        team_id = 0
    if team_id <= 0:
        return jsonify({"ok": False, "error": "team_id gecersiz"}), 400

    start = week_start(ws)
    end = start + timedelta(days=6)
    publisher = get_current_user()
    now = datetime.now()

    cells = (
        PlanCell.query
        .filter(
            PlanCell.work_date >= start,
            PlanCell.work_date <= end,
            PlanCell.team_id == team_id,
        )
        .all()
    )
    published = 0
    for cell in cells:
        if not _cell_has_meaningful_job(cell):
            continue
        _publish_cell(cell, publisher=publisher, now=now)
        published += 1

    db.session.commit()
    return jsonify({
        "ok": True,
        "published": int(published),
        "team_id": int(team_id),
        "week_start": iso(start),
        "week_end": iso(end),
    })


# ---------- PLAN ----------
@planner_bp.get("/plan")
@login_required
def plan_week():
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", ""))
    start = ws if ws else week_start(d)
    days = [start + timedelta(days=i) for i in range(7)]

    # Satirdaki isler (sehir bazli) = region != "-" ; proje sablonlari = region == "-"
    # Projeleri sadece bu haftada is eklenmisse goster
    template_projects = Project.query.filter(Project.is_active == True, Project.region == "-").order_by(Project.id.desc()).all()
    
    # Once bu haftada is eklenmis projeleri bul
    cells = PlanCell.query.filter(PlanCell.work_date >= days[0], PlanCell.work_date <= days[-1]).all()
    project_ids_with_work = {c.project_id for c in cells}
    
    # Projeleri goster: sadece bu haftada isi olanlari ekle
    base_q = Project.query.filter(Project.region != "-")
    if project_ids_with_work:
        projects = base_q.filter(Project.id.in_(project_ids_with_work))\
            .order_by(Project.region.asc(), Project.project_code.asc()).all()
        # Dropdown için benzersiz project_code'lara göre grupla (her project_code için en yüksek ID'li projeyi al)
        unique_projects_dict = {}
        for p in projects:
            code = p.project_code or ""
            if code not in unique_projects_dict or p.id > unique_projects_dict[code].id:
                unique_projects_dict[code] = p
        unique_projects_for_dropdown = sorted(unique_projects_dict.values(), key=lambda x: (x.region or "", x.project_code or ""))
    else:
        projects = []  # Bu haftada hic is yoksa projeleri gosterme
        unique_projects_for_dropdown = []
    
    code_colors = {p.project_code: pastel_color(p.project_code) for p in projects}
    cell_by_key: Dict[Tuple[int, str], PlanCell] = {(c.project_id, iso(c.work_date)): c for c in cells}

    # Satır bazında (Proje kolonunda göstermek için) tekil alt proje etiketi
    row_subproject_label: Dict[int, str] = {}
    if cells:
        monday_key = iso(start)
        sub_ids_by_project: Dict[int, set] = {}
        for c in cells:
            sid = int(c.subproject_id or 0)
            if sid <= 0:
                continue
            sub_ids_by_project.setdefault(int(c.project_id), set()).add(sid)

        # Öncelik: bu haftanın Pazartesi (week_start) hücresine set edilen alt proje
        # (Yoksa: tüm hafta boyunca tekil kalan alt proje)
        label_id_by_project: Dict[int, int] = {}
        for pid, sids in sub_ids_by_project.items():
            monday_cell = cell_by_key.get((int(pid), monday_key))
            monday_sid = int((monday_cell.subproject_id if monday_cell else 0) or 0)
            if monday_sid > 0:
                label_id_by_project[int(pid)] = monday_sid
            elif len(sids) == 1:
                label_id_by_project[int(pid)] = next(iter(sids))

        wanted_ids = sorted(set(label_id_by_project.values()))
        sub_by_id: Dict[int, SubProject] = {}
        if wanted_ids:
            for sp in SubProject.query.filter(SubProject.id.in_(wanted_ids)).all():
                sub_by_id[int(sp.id)] = sp

        for pid, sid in label_id_by_project.items():
            sp = sub_by_id.get(int(sid))
            if not sp:
                continue
            code = (sp.code or "").strip()
            row_subproject_label[int(pid)] = f"{sp.name}{f' ({code})' if code else ''}"

    ass_map: Dict[int, list] = {}
    cell_person_ids: Dict[int, list] = {}
    if cells:
        cell_ids = [c.id for c in cells]
        rows = (
            db.session.query(CellAssignment.cell_id, Person.full_name)
            .join(Person, Person.id == CellAssignment.person_id)
            .filter(CellAssignment.cell_id.in_(cell_ids))
            .all()
        )
        for cid, name in rows:
            ass_map.setdefault(cid, []).append(name)

        # Cell person IDs for team signature
        for cid, pid in db.session.query(CellAssignment.cell_id, CellAssignment.person_id).filter(CellAssignment.cell_id.in_(cell_ids)).all():
            cell_person_ids.setdefault(cid, []).append(pid)

    # Overtime data fetching logic:
    # We need to know if a cell has ANY overtime.
    # Map: cell_id -> { has_overtime: bool, total_hours: float, person_count: int, people_names: list }
    cell_overtime_info: Dict[int, dict] = {}
    if cells:
        cell_ids = [c.id for c in cells]
        overtime_rows = (
            db.session.query(TeamOvertime.cell_id, TeamOvertime.person_id, TeamOvertime.duration_hours, Person.full_name)
            .join(Person, Person.id == TeamOvertime.person_id)
            .filter(TeamOvertime.cell_id.in_(cell_ids))
            .filter(TeamOvertime.duration_hours > 0)
            .all()
        )
        for cid, pid, hours, pname in overtime_rows:
             if cid not in cell_overtime_info:
                 cell_overtime_info[cid] = {
                     "has_overtime": True,
                     "total_hours": 0.0,
                     "person_count": 0,
                     "people": set(),
                     "people_list": []
                 }
             info = cell_overtime_info[cid]
             info["total_hours"] += (hours or 0.0)
             if pid not in info["people"]:
                 info["people"].add(pid)
                 info["person_count"] += 1
                 info["people_list"].append(f"{pname} ({hours} sa)")
    
    # Clean up set for template usage (lists are safer)
    for cid in cell_overtime_info:
        cell_overtime_info[cid]["people"] = list(cell_overtime_info[cid]["people"])

    # Calculate grey cells for consecutive same teams
    cell_grey: Dict[Tuple[int, str], bool] = {}
    # Collect per day per team signature projects
    day_signature_projects: Dict[str, Dict[str, list]] = {}
    for cell in cells:
        k = iso(cell.work_date)
        person_ids = sorted(cell_person_ids.get(cell.id, []))
        sig = team_signature(person_ids) if person_ids else None
        if sig:
            if k not in day_signature_projects:
                day_signature_projects[k] = {}
            if sig not in day_signature_projects[k]:
                day_signature_projects[k][sig] = []
            day_signature_projects[k][sig].append(cell.project_id)
    
    # Mark grey if a team signature works in multiple projects on the same day
    for k, sig_dict in day_signature_projects.items():
        for sig, project_ids in sig_dict.items():
            if len(project_ids) > 1:
                for proj_id in project_ids:
                    cell_grey[(proj_id, k)] = True

    people = Person.query.order_by(Person.full_name.asc()).all()
    # Netmon firma ID'sini bul
    netmon_firma = Firma.query.filter_by(name="Netmon").first()
    netmon_firma_id = netmon_firma.id if netmon_firma else None
    
    people_json = []
    for p in people:
        person_data = {
            "id": p.id,
            "full_name": p.full_name,
            "email": p.email or "",
            "phone": p.phone or "",
            "tc_no": p.tc_no or "",
            "firma_id": p.firma_id,
            "firma_name": p.firma.name if p.firma else None,
            "seviye_name": p.seviye.name if p.seviye else None
        }
        people_json.append(person_data)

    status_map = get_person_status_map(days)

    # Aktif durum tiplerini al (özet için görünür olanlar)
    status_types = PersonnelStatusType.query.filter_by(is_active=True).order_by(PersonnelStatusType.display_order.asc()).all()
    
    # Subproject ID -> StatusType eşleştirmesi
    subproject_to_status = {}
    for st in status_types:
        if st.subproject_id:
            subproject_to_status[st.subproject_id] = st
    
    # Status Type JSON listesi (frontend için)
    status_types_json = [
        {
            "id": st.id,
            "name": st.name,
            "code": st.code,
            "color": st.color or "#64748b",
            "icon": st.icon or "📋",
            "visible": getattr(st, 'visible_in_summary', True),
            "subproject_id": st.subproject_id
        }
        for st in status_types
    ]

    # Office tipi (fallback için)
    office_status_type = next((st for st in status_types if st.code == 'office'), None)

    # Her gün için atama ve durum bilgilerini hesapla
    busy_map = {}
    status_assignments_map = {}  # day_iso -> {person_id: status_type_id}
    
    for d in days:
        day_iso = iso(d)
        
        # Tüm cell atamalarını al (subproject_id ve project_code ile birlikte)
        cell_q = db.session.query(
            CellAssignment.person_id,
            PlanCell.subproject_id,
            Project.project_code
        ).join(PlanCell, PlanCell.id == CellAssignment.cell_id)\
         .join(Project, Project.id == PlanCell.project_id)\
         .filter(PlanCell.work_date == d, PlanCell.status != 'cancelled')
        
        busy_ids = set()
        status_assignments = {}  # person_id -> status_type_id
        
        for row in cell_q.all():
            person_id, subproject_id, project_code = row
            busy_ids.add(person_id)
            
            mapped = False
            # 1. Subproject_id ile eşleşen status type varsa kaydet
            if subproject_id and subproject_id in subproject_to_status:
                status_assignments[person_id] = subproject_to_status[subproject_id].id
                mapped = True
            
            # 2. Fallback: Project Code "9026-0001%" ise ve mapped değilse -> Ofis
            if not mapped and project_code and project_code.startswith("9026-0001") and office_status_type:
                status_assignments[person_id] = office_status_type.id
                mapped = True
        
        busy_map[day_iso] = busy_ids
        status_assignments_map[day_iso] = status_assignments

    # Personel durum özeti
    personnel_summary = {}
    for d in days:
        day_iso = iso(d)
        busy_ids = busy_map[day_iso]
        status_assignments = status_assignments_map.get(day_iso, {})
        
        # Seviyeler: Ana kadro, Yardımcı, Alt yüklenici
        summary = {
            "total_available": [],
            "ana_kadro_available": [],
            "yardimci_available": [],
            "alt_yuklenici_available": []
        }
        firm_available = {}
        
        # Dinamik status kategorileri için sayaçlar
        status_counts = {st.id: [] for st in status_types}  # status_type_id -> [isimler]
        
        for p in people:
            st = status_map.get((p.id, day_iso), "available")
            firm_label = p.firma.name if p.firma else "Firma belirtilmemiş"
            
            # PersonDayStatus tablosunda izinli olanlar
            if st == "leave":
                # Varsayılan olarak ilk leave tipine ekle
                for status_type in status_types:
                    if status_type.code == "leave" or status_type.code == "annual_leave":
                        status_counts[status_type.id].append(p.full_name)
                        break
                continue
            
            # Cell ataması ile status type eşleştirmesi
            if p.id in status_assignments:
                status_type_id = status_assignments[p.id]
                if status_type_id in status_counts:
                    status_counts[status_type_id].append(p.full_name)
                continue
            
            # Normal projede çalışıyor (status type olmayan)
            if p.id in busy_ids:
                continue
            
            # Boşta
            summary["total_available"].append(p.full_name)
            firm_available.setdefault(firm_label, []).append(p.full_name)
            
            seviye_name = p.seviye.name if p.seviye else None
            if seviye_name == "Ana kadro":
                summary["ana_kadro_available"].append(p.full_name)
            elif seviye_name == "Yardımcı":
                summary["yardimci_available"].append(p.full_name)
            elif seviye_name == "Alt yüklenici":
                summary["alt_yuklenici_available"].append(p.full_name)
        
        # Sonuç sözlüğü
        day_summary = {
            "total": len(summary["total_available"]),
            "ana_kadro": len(summary["ana_kadro_available"]),
            "yardimci": len(summary["yardimci_available"]),
            "alt_yuklenici": len(summary["alt_yuklenici_available"]),
            "total_names": summary["total_available"],
            "ana_kadro_names": summary["ana_kadro_available"],
            "yardimci_names": summary["yardimci_available"],
            "alt_yuklenici_names": summary["alt_yuklenici_available"],
            "firm_available": firm_available,
            "firm_available_counts": {k: len(v) for k, v in firm_available.items()},
            "status_categories": {}  # Dinamik kategoriler
        }
        
        # Dinamik status kategorilerini ekle
        for st in status_types:
            cat_key = f"status_{st.id}"
            day_summary["status_categories"][st.id] = {
                "id": st.id,
                "name": st.name,
                "code": st.code,
                "color": st.color or "#64748b",
                "icon": st.icon or "📋",
                "visible": getattr(st, 'visible_in_summary', True),
                "count": len(status_counts[st.id]),
                "names": status_counts[st.id]
            }
        
        personnel_summary[day_iso] = day_summary


    # iller dropdown: önce TR_CITIES, ayrıca DB'deki mevcut iller (yanlış yazılmış eski şehirler de kaybolmasın)
    existing_cities = [r[0] for r in db.session.query(Project.region).filter(Project.region != '-', Project.region != None).distinct().all()]
    cities = []
    for c in TR_CITIES + sorted(set(existing_cities)):
        if c and c not in cities:
            cities.append(c)

    # Alt projeleri projelerin altına grupla (görünür başka yerde)
    project_ids = [int(p.id) for p in projects] if projects else []
    project_subprojects = {}
    if project_ids:
        subs = (
            SubProject.query
            .filter(SubProject.project_id.in_(project_ids))
            .order_by(SubProject.project_id.asc(), SubProject.created_at.asc(), SubProject.id.asc())
            .all()
        )
        for sp in subs:
            project_subprojects.setdefault(int(sp.project_id), []).append({
                "name": sp.name,
                "code": sp.code or "",
                "is_active": bool(sp.is_active)
            })

    # Araçları veritabanından çek
    vehicles = Vehicle.query.order_by(Vehicle.plate.asc()).all()
    field_users = User.query.filter(User.is_active == True, User.role == "field").order_by(User.full_name.asc(), User.email.asc()).all()
    
    # HAFTALIK ARAÇ ATAMA KONTROLÜ
    # Bu haftaki PlanCell'lerde hangi araçlar kullanılmış? (plate -> team_id)
    # Böylece yeni hafta için atamalar sıfırdan başlar
    week_end = start + timedelta(days=6)
    weekly_vehicle_usage = db.session.query(
        PlanCell.vehicle_info,
        PlanCell.team_id
    ).filter(
        PlanCell.work_date >= start,
        PlanCell.work_date <= week_end,
        PlanCell.vehicle_info.isnot(None),
        db.func.trim(PlanCell.vehicle_info) != "",
        PlanCell.team_id.isnot(None)
    ).distinct().all()
    
    # plate -> team_id haritası (bu hafta için)
    weekly_plate_team_map = {}
    for vinfo, tid in weekly_vehicle_usage:
        plate = (vinfo or "").strip().split("(")[0].strip()  # "34ABC123 (Ford)" -> "34ABC123"
        if plate and tid:
            weekly_plate_team_map[plate] = tid
    
    vehicles_json = []
    for v in vehicles:
        # Bu araç bu hafta herhangi bir ekibe atanmış mı?
        weekly_team_id = weekly_plate_team_map.get(v.plate)
        payload = _vehicle_payload(v, weekly_team_id)
        if payload:
            vehicles_json.append(payload)

    return render_template(
        "plan.html",
        start=start, days=days,
        template_projects=template_projects,
        projects=projects,
        unique_projects_for_dropdown=unique_projects_for_dropdown,
        cell_by_key=cell_by_key,
        cell_overtime_info=cell_overtime_info,
        row_subproject_label=row_subproject_label,
        ass_map=ass_map,
        cell_person_ids=cell_person_ids,
        cell_grey=cell_grey,
        people_json=people_json,
                template_projects_json=[
            {"id":p.id,"project_code":p.project_code,"project_name":p.project_name,"responsible":p.responsible}
            for p in template_projects
        ],
        prev_week=iso(start - timedelta(days=7)),
        next_week=iso(start + timedelta(days=7)),
        selected_week=iso(start),
        week_start_iso=iso(start),
        code_colors=code_colors,
        tr_days=TR_DAYS,
        today_iso=iso(date.today()),
        status_map=status_map,
        personnel_summary=personnel_summary,
        status_types_json=status_types_json,
        cities=cities,
        tr_cities=TR_CITIES,  # All 81 Turkish provinces for route starting point
        vehicles=vehicles,
        vehicles_json=vehicles_json,
        project_subprojects=project_subprojects,
        field_users=field_users
    )



def hsl_to_rgb_hex(hsl_str: str) -> str:
    """HSL string'i (hsl(hue 70% 88%)) RGB hex'e çevir"""
    import re
    # Try to match hsl(hue sat% light%) format
    match = re.match(r'hsl\((\d+)\s+(\d+)%\s+(\d+)%\)', hsl_str)
    if not match:
        # Try hex format directly
        if hsl_str.startswith('#'):
            hex_val = hsl_str[1:]
            if len(hex_val) == 6:
                return hex_val.upper()
        return "FEF3C7"  # Varsayılan sarı
    h = int(match.group(1)) / 360.0
    s = int(match.group(2)) / 100.0
    l = int(match.group(3)) / 100.0
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"


@planner_bp.get("/plan/export/excel")
@login_required
def plan_export_excel():
    """Haftalık planı Excel olarak indir"""
    try:
        d = parse_date(request.args.get("week_start", "")) or date.today()
        start = week_start(d)
        days = [start + timedelta(days=i) for i in range(7)]
    except Exception:
        flash("Hafta tarihi geçersiz.", "danger")
        return redirect(url_for("planner.plan_week"))

    # Projeleri ve hücreleri al
    cells = PlanCell.query.filter(PlanCell.work_date >= days[0], PlanCell.work_date <= days[-1]).all()
    project_ids_with_work = {c.project_id for c in cells}
    
    base_q = Project.query.filter(Project.region != "-")
    if project_ids_with_work:
        projects = base_q.filter(Project.id.in_(project_ids_with_work))\
            .order_by(Project.region.asc(), Project.project_code.asc()).all()
    else:
        projects = []
    
    cell_by_key: Dict[Tuple[int, str], PlanCell] = {(c.project_id, iso(c.work_date)): c for c in cells}
    
    # Proje kodlarına göre renkleri hesapla
    code_colors = {p.project_code: pastel_color(p.project_code) for p in projects}
    
    # Satır bazında (Proje kolonunda göstermek için) tekil alt proje etiketi
    row_subproject_label: Dict[int, str] = {}
    if cells:
        monday_key = iso(start)
        sub_ids_by_project: Dict[int, set] = {}
        for c in cells:
            sid = int(c.subproject_id or 0)
            if sid <= 0:
                continue
            sub_ids_by_project.setdefault(int(c.project_id), set()).add(sid)

        # Öncelik: bu haftanın Pazartesi (week_start) hücresine set edilen alt proje
        # (Yoksa: tüm hafta boyunca tekil kalan alt proje)
        label_id_by_project: Dict[int, int] = {}
        for pid, sids in sub_ids_by_project.items():
            monday_cell = cell_by_key.get((int(pid), monday_key))
            monday_sid = int((monday_cell.subproject_id if monday_cell else 0) or 0)
            if monday_sid > 0:
                label_id_by_project[int(pid)] = monday_sid
            elif len(sids) == 1:
                label_id_by_project[int(pid)] = next(iter(sids))

        wanted_ids = sorted(set(label_id_by_project.values()))
        sub_by_id: Dict[int, SubProject] = {}
        if wanted_ids:
            for sp in SubProject.query.filter(SubProject.id.in_(wanted_ids)).all():
                sub_by_id[int(sp.id)] = sp

        for pid, sid in label_id_by_project.items():
            sp = sub_by_id.get(int(sid))
            if not sp:
                continue
            code = (sp.code or "").strip()
            row_subproject_label[int(pid)] = f"{sp.name}{f' ({code})' if code else ''}"
    
    # Personel atamalarını al
    ass_map: Dict[int, list] = {}
    if cells:
        cell_ids = [c.id for c in cells]
        rows = (
            db.session.query(CellAssignment.cell_id, Person.full_name)
            .join(Person, Person.id == CellAssignment.person_id)
            .filter(CellAssignment.cell_id.in_(cell_ids))
            .all()
        )
        for cid, name in rows:
            ass_map.setdefault(cid, []).append(name)
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = f"Hafta {start.strftime('%d.%m.%Y')}"
    
    # Başlık satırı (gün adı locale'dan bağımsız)
    _gun_adlari = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    header_row = ["İL", "PROJE", "SORUMLU"]
    for d in days:
        gun_adi = _gun_adlari[d.weekday()]  # 0=Monday
        header_row.append(d.strftime("%d.%m.%Y") + " " + gun_adi)
    ws.append(header_row)
    
    # Başlık stil - koyu gri arka plan
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Renk tanımlamaları
    il_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # Yeşil açık
    resp_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")  # Sarı açık
    cell_filled_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Dolu hücre - açık yeşil
    cell_personnel_fill = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")  # Personel var - koyu yeşil
    
    # Veri satırları
    row_num = 2
    for p in projects:
        # PROJE kolonu: proje kodu + proje adı, karşı firma sorumlusu (varsa), alt proje (varsa)
        proj_parts = [f"{p.project_code} {p.project_name}"]
        if p.karsi_firma_sorumlusu:
            proj_parts.append(p.karsi_firma_sorumlusu)
        sp_label = row_subproject_label.get(p.id)
        if sp_label:
            proj_parts.append(sp_label)
        proj_text = "\\n".join(proj_parts)
        
        row_data = [p.region or "", proj_text, p.responsible or ""]
        
        for d in days:
            k = iso(d)
            cell = cell_by_key.get((p.id, k))
            cell_text = ""
            
            if cell:
                parts = []
                # Alt proje bilgisi (varsa)
                if cell.subproject_id:
                    subproject = SubProject.query.get(cell.subproject_id)
                    if subproject:
                        code = (subproject.code or "").strip()
                        subproject_label = f"{subproject.name}{f' ({code})' if code else ''}"
                        parts.append(subproject_label)
                
                # Personel
                if ass_map.get(cell.id):
                    parts.append(", ".join(ass_map[cell.id]))
                
                # Çalışma Saati (Varsa)
                hours = calculate_hours_from_shift(cell.shift)
                if hours > 0:
                    parts.append(f"Saat: {hours}")
                
                # Çalışma Detayı
                if cell.job_mail_body:
                    parts.append(f"Detay: {cell.job_mail_body}")
                
                # Not (varsa)
                if cell.note:
                    parts.append(f"Not: {cell.note}")
                
                cell_text = " | ".join(parts) if parts else "-"
            else:
                cell_text = "-"
            
            row_data.append(cell_text)
        
        ws.append(row_data)
        
        # İL kolonu - yeşil
        ws[f'A{row_num}'].fill = il_fill
        ws[f'A{row_num}'].font = Font(bold=True, color="166534")
        
        # PROJE kolonu - proje koduna göre renk
        proj_color = code_colors.get(p.project_code, '#fef3c7')
        rgb_hex = hsl_to_rgb_hex(proj_color)
        proj_fill = PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")
        ws[f'B{row_num}'].fill = proj_fill
        
        # PROJE SORUMLUSU kolonu - sarı
        ws[f'C{row_num}'].fill = resp_fill
        ws[f'C{row_num}'].font = Font(color="92400E")
        
        # Gün hücreleri - dolu ise renkli
        col_idx = 4
        for d in days:
            k = iso(d)
            cell = cell_by_key.get((p.id, k))
            col_letter = get_column_letter(col_idx)
            
            if cell and ass_map.get(cell.id):
                # Personel varsa koyu yeşil
                ws[f'{col_letter}{row_num}'].fill = cell_personnel_fill
            elif cell:
                # Hücre dolu ama personel yok
                ws[f'{col_letter}{row_num}'].fill = cell_filled_fill
            else:
                # Boş hücre - beyaz
                ws[f'{col_letter}{row_num}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            col_idx += 1
        
        row_num += 1
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 15  # İL
    ws.column_dimensions['B'].width = 30  # PROJE
    ws.column_dimensions['C'].width = 20  # SORUMLU
    for i in range(4, 11):  # Günler
        ws.column_dimensions[get_column_letter(i)].width = 25
    
    # Satır yüksekliklerini ayarla
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 60
    
    # Excel dosyasını response olarak döndür
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"haftalik_plan_{start.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ---------- PROJECTS ----------
@planner_bp.route("/api/next-project-code", methods=["GET"])
@login_required
@observer_required
def api_next_project_code():
    """En son eklenen projenin kodunu bulup bir sonraki kodunu döndürür (otomatik öneri)"""
    # Sadece ana projeleri al (region == '-'), en yeni eklenen (id'ye göre) proje
    last_project = Project.query.filter(Project.region == "-")\
        .order_by(Project.id.desc()).first()
    
    if not last_project:
        return jsonify({"next_code": "P-001"})
    
    current_code = last_project.project_code or ""
    
    # Kod formatını analiz et ve sonraki kodu hesapla
    import re
    
    # Format: "P-XXX" veya "XXXX-XXXX" gibi formatları destekle
    # Sayısal kısmı bul
    numeric_match = re.search(r'(\d+)$', current_code)
    
    if numeric_match:
        numeric_part = numeric_match.group(1)
        next_numeric = int(numeric_part) + 1
        
        # Aynı formatta yeni kod oluştur
        prefix = current_code[:len(current_code) - len(numeric_part)]
        next_code = f"{prefix}{next_numeric:0{len(numeric_part)}d}"
    else:
        # Format tanınmazsa basit bir ekleme yap
        next_code = f"{current_code}-1"
    
    return jsonify({"next_code": next_code})

@planner_bp.route("/projects", methods=["GET", "POST"])
@login_required
@observer_required
def projects_page():
    if request.method == "POST":
        # Artık proje eklemede il/bölge kullanıcıdan alınmıyor (legacy kolon: region)
        region = "-"
        project_code = request.form.get("project_code", "").strip()
        project_name = request.form.get("project_name", "").strip()
        responsible = request.form.get("responsible", "").strip()
        karsi_firma_sorumlusu = request.form.get("karsi_firma_sorumlusu", "").strip()
        is_active = (request.form.get("is_active") == "1")

        if not all([project_code, project_name, responsible]):
            # flash("Zorunlu alanlar boş.", "danger")
            # return redirect(url_for('planner.projects_page'))
            return jsonify({"ok": False, "error": "Zorunlu alanlar boş."}), 400

        # Validasyon: Aynı proje koduyla başka ana proje (region '-') olamaz
        existing_project = Project.query.filter(
            Project.project_code == project_code,
            Project.region == "-"
        ).first()
        if existing_project:
            return jsonify({"ok": False, "error": f"Hata: '{project_code}' kodlu bir proje zaten mevcut! Aynı kodla ikinci proje açılamaz."}), 400

        notification_enabled = (request.form.get("notification_enabled") == "1")

        # Proje Başlangıç Dosyası İşlemleri
        initiation_file = request.files.get("initiation_file")
        no_initiation_file = (request.form.get("no_initiation_file") == "1")
        no_file_reason = request.form.get("no_file_reason", "").strip()

        initiation_file_path = None
        initiation_file_type = None
        initiation_file_name = None

        if not initiation_file and not no_initiation_file:
             return jsonify({"ok": False, "error": "Proje başlangıç dosyası veya nedeni zorunludur."}), 400
             
        if no_initiation_file and not no_file_reason:
             return jsonify({"ok": False, "error": "Dosya eklenmediyse nedeni belirtilmelidir."}), 400

        if initiation_file and initiation_file.filename:
            filename = secure_filename(initiation_file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            
            # Upload folder check
            if not os.path.exists(current_app.config["UPLOAD_FOLDER"]):
                os.makedirs(current_app.config["UPLOAD_FOLDER"])
            
            upload_path = os.path.join(current_app.config["UPLOAD_FOLDER"], unique_filename)
            initiation_file.save(upload_path)
            
            initiation_file_path = unique_filename
            initiation_file_type = initiation_file.content_type
            initiation_file_name = filename

        new_project = Project(region=region, project_code=project_code, project_name=project_name,
                               responsible=responsible, karsi_firma_sorumlusu=karsi_firma_sorumlusu if karsi_firma_sorumlusu else None, 
                               is_active=is_active, notification_enabled=notification_enabled,
                               initiation_file_path=initiation_file_path,
                               initiation_file_type=initiation_file_type,
                               initiation_file_name=initiation_file_name,
                               no_initiation_file=no_initiation_file,
                               no_file_reason=no_file_reason)
        db.session.add(new_project)
        db.session.commit()
        
        # Yeni Proje Eklendiğinde Mail Bildirimi (Otomatik)
        if notification_enabled:
            try:
                # Admin ve Planner yetkili aktif kullanıcıları bul
                admin_users = User.query.filter(User.role.in_(["admin", "planner"])).filter(User.is_active == True).all()
                recipients = [u.email for u in admin_users if u.email and "@" in u.email]
                
                if recipients:
                    subject = f"Yeni Proje Eklendi: {project_code} - {project_name}"
                    current_user = get_current_user()
                    
                    import html
                    from services.mail_service import MailService
                    
                    # Mail içeriği oluştur
                    body_content = f"""
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 12px;">
                        <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                            <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Proje Kodu</div>
                            <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(project_code)}</div>
                        </div>
                        <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                            <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Proje Adı</div>
                            <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(project_name)}</div>
                        </div>
                        <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                            <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Sorumlu</div>
                            <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(responsible)}</div>
                        </div>
                    </div>
                    """
                    
                    for recipient in recipients:
                        try:
                            html_body = render_template(
                                "email_base.html",
                                title=subject,
                                heading="Yeni Proje Eklendi",
                                intro="Sisteme yeni bir proje eklendi.",
                                body_html=body_content,
                                footer="Bu bildirim Görev Takip Sistemi tarafından otomatik olarak gönderilmiştir.",
                            )
                            
                            MailService.send(
                                mail_type="project",
                                recipients=[recipient],
                                subject=subject,
                                html=html_body,
                                user_id=getattr(current_user, "id", None),
                                project_id=new_project.id,
                                meta={"type": "project_added", "project_code": project_code, "project_name": project_name},
                            )
                        except Exception as ex:
                            current_app.logger.error(f"Proje ekleme maili gonderilemedi ({recipient}): {ex}")
            except Exception as e:
                current_app.logger.exception(f"Proje ekleme mail hatasi: {str(e)}")

        return jsonify({"ok": True, "message": "Proje eklendi."})

    # Projeler menüsü sadece "proje şablonlarını" gösterir (region == '-')
    # Plan sayfasından eklenen "iş satırları" (region != '-') burada görünmez.
    base_query = Project.query.filter(Project.region == "-")
    
    # Pasif projeleri göster filtresi
    show_inactive = request.args.get("show_inactive", "0") == "1"
    if not show_inactive:
        # Varsayılan: yalnızca aktif projeler
        base_query = base_query.filter(Project.is_active == True)
    # show_inactive True ise: Hem aktif hem pasif projeleri göster (filtre yok)
    
    # Arama filtresi
    search_query = request.args.get("search", "").strip()
    if search_query:
        search_filter = f"%{search_query}%"
        # ProjectComment ile join yap (outer join ki yorumu olmayanlar da gelsin)
        base_query = base_query.outerjoin(ProjectComment, ProjectComment.project_id == Project.id)
        
        base_query = base_query.filter(
            db.or_(
                Project.project_code.ilike(search_filter),
                Project.project_name.ilike(search_filter),
                Project.responsible.ilike(search_filter),
                Project.karsi_firma_sorumlusu.ilike(search_filter),
                ProjectComment.comment.ilike(search_filter),      # Yorum içeriğinde ara
                ProjectComment.file_name.ilike(search_filter)     # Dosya adında ara
            )
        ).distinct() # Aynı proje birden fazla yorumla eşleşirse tekrar etmesin
    
    # Sıralama
    sort_column = request.args.get("sort", "project_code")
    sort_order = request.args.get("order", "desc")  # Varsayılan: büyükten küçüğe
    
    if sort_column == "project_code":
        order_by = Project.project_code.asc() if sort_order == "asc" else Project.project_code.desc()
    elif sort_column == "project_name":
        order_by = Project.project_name.asc() if sort_order == "asc" else Project.project_name.desc()
    elif sort_column == "responsible":
        order_by = Project.responsible.asc() if sort_order == "asc" else Project.responsible.desc()
    elif sort_column == "karsi_firma_sorumlusu":
        order_by = Project.karsi_firma_sorumlusu.asc() if sort_order == "asc" else Project.karsi_firma_sorumlusu.desc()
    elif sort_column == "is_active":
        order_by = Project.is_active.desc() if sort_order == "asc" else Project.is_active.asc()
    else:
        order_by = Project.project_code.desc()  # Varsayılan: büyükten küçüğe
    
    # Proje yorumlarını da getir (arama ekranında göstermek için)
    base_query = base_query.options(db.joinedload(Project.comments))
    projects = base_query.order_by(order_by).all()
    users = User.query.filter(User.is_active == True).order_by(User.full_name.asc(), User.username.asc()).all()
    
    # Proje sayılarını hesapla
    open_main_projects_count = Project.query.filter(Project.region == "-", Project.is_active == True).count()
    open_closed_projects_count = Project.query.filter(Project.region != "-", Project.is_active == True).count()
    
    if request.args.get("partial"):
        return render_template("partials/projects_table_body.html", projects=projects)

    return render_template("projects.html", projects=projects, users=users, search_query=search_query, sort_column=sort_column, sort_order=sort_order, show_inactive=show_inactive, open_main_projects_count=open_main_projects_count, open_closed_projects_count=open_closed_projects_count)


# ---------- PERSONEL DURUM AYARLARI ----------
@planner_bp.get("/personnel-settings")
@login_required
@planner_or_admin_required
def personnel_settings_page():
    """Personel durum tipleri ayarları sayfası"""
    # Durum tiplerini al
    status_types = PersonnelStatusType.query.order_by(PersonnelStatusType.display_order.asc()).all()
    
    # 9026-0001 alt projelerini al
    office_project = Project.query.filter(Project.project_code.like("9026-0001%")).first()
    office_subprojects = []
    if office_project:
        subs = SubProject.query.filter_by(project_id=office_project.id, is_active=True)\
            .order_by(SubProject.code.asc()).all()
        office_subprojects = [{"id": sp.id, "code": sp.code or "", "name": sp.name} for sp in subs]
    
    return render_template(
        "personnel_settings.html",
        status_types=status_types,
        office_subprojects=office_subprojects
    )


def calculate_hours_from_shift(shift_value: str) -> float:
    """
    Çalışma saatlerinden toplam saat hesapla.
    Görseldeki mantığa göre:
    - "08:30 - 18:00" -> 8.5 saat (1 saat mola çıkarılıyor)
    - "08:30 - 18:00 YOL" -> 8.5 saat (1 saat mola çıkarılıyor)
    - "00:00 - 06:00" -> 8.5 saat (gece vardiyası)
    - "08:30 - 12:30" -> 4 saat (yarım gün, mola yok)
    - "13:30 - 18:00" -> 4.5 saat (yarım gün, mola yok)
    """
    if not shift_value or not shift_value.strip():
        return 0.0
    
    shift = shift_value.strip()
    has_yol = " YOL" in shift
    
    # Saat aralığını parse et (örn: "08:30 - 18:00")
    if " - " in shift:
        parts = shift.split(" - ", 1)
        if len(parts) == 2:
            start_str = parts[0].strip()
            end_str = parts[1].strip()
            
            # "YOL" kelimesini kaldır
            end_str = end_str.replace(" YOL", "").strip()
            
            try:
                # Saat:dakika formatını parse et
                start_parts = start_str.split(":")
                end_parts = end_str.split(":")
                
                if len(start_parts) == 2 and len(end_parts) == 2:
                    start_hour = int(start_parts[0])
                    start_min = int(start_parts[1])
                    end_hour = int(end_parts[0])
                    end_min = int(end_parts[1])
                    
                    # Toplam dakikayı hesapla
                    start_total_min = start_hour * 60 + start_min
                    end_total_min = end_hour * 60 + end_min
                    
                    # Eğer bitiş saati başlangıçtan küçükse, ertesi güne geçmiş demektir (örn: 00:00 - 06:00)
                    if end_total_min < start_total_min:
                        end_total_min += 24 * 60
                    
                    diff_min = end_total_min - start_total_min
                    hours = diff_min / 60.0
                    
                    # Tam gün çalışma (08:30 - 18:00) için 1 saat mola çıkar
                    # Görseldeki mantığa göre: 08:30-18:00 = 8.5 saat
                    if start_total_min == 510 and end_total_min == 1080:  # 08:30 - 18:00
                        hours = 8.5
                    # Gece vardiyası (00:00 - 06:00) için 8.5 saat
                    elif start_total_min == 0 and end_total_min == 360:  # 00:00 - 06:00
                        hours = 8.5
                    # Yarım gün çalışmalar için mola çıkarılmaz
                    # Diğer durumlar için hesaplanan saat kullanılır
                    
                    return round(hours, 1)
            except (ValueError, IndexError):
                pass
    
    return 0.0


@planner_bp.get("/projects/export/excel")
@login_required
@observer_required
def projects_export_excel():
    """Projeleri Excel olarak indir"""
    from io import BytesIO
    from flask import send_file
    
    project_type = request.args.get("type", "main")  # "main" veya "sub"
    show_inactive = request.args.get("show_inactive", "0") == "1"
    search_query = request.args.get("search", "").strip()
    
    # Ana projeler için sorgu
    if project_type == "main":
        base_query = Project.query.filter(Project.region == "-")
        
        # Pasif projeler filtresi
        if not show_inactive:
            base_query = base_query.filter(Project.is_active == True)
        
        # Arama filtresi
        if search_query:
            search_filter = f"%{search_query}%"
            base_query = base_query.filter(
                db.or_(
                    Project.project_code.ilike(search_filter),
                    Project.project_name.ilike(search_filter),
                    Project.responsible.ilike(search_filter),
                    Project.karsi_firma_sorumlusu.ilike(search_filter)
                )
            )
        
        projects = base_query.order_by(Project.project_code.asc()).all()
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Ana Projeler"
        
        # Başlık satırı
        headers = ["Proje Kodu", "Proje Adı", "Sorumlu", "Karşı Firma Sorumlusu", "Durum"]
        ws.append(headers)
        
        # Başlık stil
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Veri satırları
        for p in projects:
            ws.append([
                p.project_code or "",
                p.project_name or "",
                p.responsible or "",
                p.karsi_firma_sorumlusu or "",
                "Aktif" if p.is_active else "Pasif"
            ])
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        
        filename = f"ana_projeler_{date.today().strftime('%Y%m%d')}.xlsx"
    
    else:  # Alt projeler
        base_query = SubProject.query.join(Project)
        
        # Pasif projeler filtresi
        if not show_inactive:
            base_query = base_query.filter(Project.is_active == True)
        
        # Arama filtresi
        if search_query:
            search_filter = f"%{search_query}%"
            base_query = base_query.filter(
                db.or_(
                    Project.project_code.ilike(search_filter),
                    Project.project_name.ilike(search_filter),
                    SubProject.name.ilike(search_filter)
                )
            )
        
        subprojects = base_query.order_by(Project.project_code.asc(), SubProject.name.asc()).all()
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Alt Projeler"
        
        # Başlık satırı
        headers = ["Proje Kodu", "Proje Adı", "Alt Proje Adı", "Durum"]
        ws.append(headers)
        
        # Başlık stil
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Veri satırları
        for sp in subprojects:
            p = sp.project
            ws.append([
                p.project_code if p else "",
                p.project_name if p else "",
                sp.name or "",
                "Aktif" if (p and p.is_active) else "Pasif"
            ])
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        
        filename = f"alt_projeler_{date.today().strftime('%Y%m%d')}.xlsx"
    
    # Excel dosyasını oluştur
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@planner_bp.post("/projects/<int:project_id>/delete")
@login_required
@observer_required
def project_delete(project_id: int):
    p = Project.query.get_or_404(project_id)

    # 1. İlişkili Job kayıtlarını bul (Project ile bağlantısı olup cascade olmayanlar)
    # Job modeli Project'e bağlı ama cascade tanımlı değil.
    # Ayrıca JobReport modeli Job'a bağlı ve cascade tanımlı değil.
    jobs = Job.query.filter_by(project_id=project_id).all()
    for job in jobs:
        # JobReport kayıtlarını sil
        JobReport.query.filter_by(job_id=job.id).delete()
        # Job kaydını sil (Assignments, Feedback, History cascade ile silinir)
        db.session.delete(job)
    
    # 2. Projeyi sil (SubProject, PlanCell vb. cascade ile silinir)
    db.session.delete(p)
    db.session.commit()
    flash("Proje silindi.", "success")
    return redirect(url_for('planner.projects_page'))







@planner_bp.post("/api/project_comment")
@login_required
@observer_required
def api_project_comment():
    try:
        data = request.get_json(force=True, silent=True) or {}
        token = str(data.get("csrf_token") or "")
        if not _csrf_verify(token):
            return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400
        
        project_id = int(data.get("project_id", 0) or 0)
        subproject_id = int(data.get("subproject_id", 0) or 0) if data.get("subproject_id") else None
        comment = str(data.get("comment") or "").strip()
        
        if not comment:
            return jsonify({"ok": False, "error": "Yorum boş olamaz."}), 400
        
        if not project_id:
            return jsonify({"ok": False, "error": "project_id eksik."}), 400
        
        # Proje veya alt proje var mı kontrol et
        if subproject_id:
            sp = SubProject.query.get(subproject_id)
            if not sp or sp.project_id != project_id:
                return jsonify({"ok": False, "error": "Alt proje bulunamadi."}), 404
        else:
            p = Project.query.get(project_id)
            if not p:
                return jsonify({"ok": False, "error": "Proje bulunamadi."}), 404
        
        user = get_current_user()
        if not user:
            return jsonify({"ok": False, "error": "Kullanici bulunamadi. Lutfen tekrar giris yapin."}), 401
        
        # Dosya İşlemleri (Base64)
        file_path = None
        file_type = None
        file_name = None
        
        file_data = data.get("file_data") # data:image/png;base64,....
        file_name_input = data.get("file_name")
        
        if file_data and file_name_input:
            import base64
            import uuid
            
            try:
                # data URL header'ını temizle
                if "base64," in file_data:
                    header, encoded = file_data.split("base64,", 1)
                else:
                    encoded = file_data
                
                decoded_data = base64.b64decode(encoded)
                
                # Dosya uzantısını ve tipini belirle
                ext = file_name_input.split('.')[-1].lower() if '.' in file_name_input else "bin"
                is_image = ext in ['png', 'jpg', 'jpeg', 'gif', 'webp']
                file_type = "image" if is_image else "document"
                
                # Benzersiz isim oluştur
                unique_filename = f"comment_{uuid.uuid4().hex}.{ext}"
                upload_folder = current_app.config["UPLOAD_FOLDER"]
                full_path = os.path.join(upload_folder, unique_filename)
                
                with open(full_path, "wb") as f:
                    f.write(decoded_data)
                
                file_path = unique_filename
                file_name = file_name_input
            except Exception as e:
                print(f"Dosya kaydetme hatası: {e}")
                # Hata olsa bile yorumu kaydet, ama dosyasız
        
        # Alt proje için de ana proje ID'sini sakla
        new_comment = ProjectComment(
            project_id=project_id,
            subproject_id=subproject_id,
            comment=comment,
            created_by_user_id=user.id,
            file_path=file_path,
            file_type=file_type,
            file_name=file_name
        )
        # (Pass ve eski yorum satırı silindi)
        
        db.session.add(new_comment)
        db.session.commit()
        
        return jsonify({"ok": True, "comment_id": int(new_comment.id)})
    except ValueError as e:
        return jsonify({"ok": False, "error": f"Gecersiz veri: {str(e)}"}), 400
    except Exception as e:
        db.session.rollback()
        import traceback
        error_msg = str(e)
        print(f"Project comment error: {error_msg}")
        print(traceback.format_exc())
        return jsonify({"ok": False, "error": f"Yorum eklenirken bir hata olustu: {error_msg}"}), 500


@planner_bp.get("/api/project_comments")
@login_required
@observer_required
def api_project_comments():
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0) if request.args.get("subproject_id") else None
    
    if not project_id:
        return jsonify({"ok": False, "error": "project_id eksik."}), 400
    
    query = ProjectComment.query
    if subproject_id:
        query = query.filter(ProjectComment.subproject_id == subproject_id)
    else:
        query = query.filter(ProjectComment.project_id == project_id, ProjectComment.subproject_id.is_(None))
    
    comments = query.order_by(ProjectComment.created_at.asc()).all()
    
    result = []
    for c in comments:
        user = User.query.get(c.created_by_user_id)
        result.append({
            "id": int(c.id),
            "comment": c.comment,
            "created_at": iso(c.created_at) if c.created_at else None,
            "created_by_name": (user.full_name or user.username) if user else "Bilinmeyen",
            "file_path": c.file_path,
            "file_type": c.file_type,
            "file_name": c.file_name
        })
    
    # Proje Başlangıç Bilgisi (Sadece ana proje detaylarında gösterilir)
    initiation = None
    if not subproject_id and project_id:
        project = Project.query.get(project_id)
        if project:
            initiation = {
                "file_path": project.initiation_file_path,
                "file_name": project.initiation_file_name,
                "file_type": project.initiation_file_type,
                "no_file": project.no_initiation_file,
                "reason": project.no_file_reason
            }

    return jsonify({"ok": True, "comments": result, "initiation": initiation})


@planner_bp.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
@observer_required
def project_edit(project_id: int):
    try:
        p = Project.query.get_or_404(project_id)
        if request.method == "POST":
            # Pasif durum kontrolü (İstek 2)
            new_is_active = (request.form.get("is_active") == "1")
            
            # Eğer proje Pasif yapılıyorsa ve şu an Aktif ise, alt projeleri kontrol et
            if not new_is_active and p.is_active:
                active_subprojects_count = SubProject.query.filter_by(project_id=p.id, is_active=True).count()
                if active_subprojects_count > 0:
                    flash(f"Alt projeleri aktif olan proje pasif yapılamaz! Lütfen önce {active_subprojects_count} adet aktif alt projeyi pasif duruma getirin.", "warning")
                    return redirect(url_for("planner.project_edit", project_id=project_id))

            new_code = request.form.get("project_code", "").strip()
            p.project_name = request.form.get("project_name", "").strip()
            p.responsible = request.form.get("responsible", "").strip()
            p.karsi_firma_sorumlusu = request.form.get("karsi_firma_sorumlusu", "").strip() or None
            p.is_active = new_is_active
            if not all([new_code, p.project_name, p.responsible]):
                flash("Zorunlu alanlar boş.", "danger")
                return redirect(url_for("planner.project_edit", project_id=project_id))
            # Ana projelerde (region '-') proje kodu benzersiz olmalı
            existing = Project.query.filter(
                Project.project_code == new_code,
                Project.region == "-",
                Project.id != project_id
            ).first()
            if existing:
                flash(f"'{new_code}' kodlu bir proje zaten mevcut. Aynı kodla ikinci proje açılamaz.", "danger")
                return redirect(url_for("planner.project_edit", project_id=project_id))
            p.project_code = new_code
            db.session.commit()
            flash("Proje güncellendi.", "success")
            return redirect(url_for('planner.projects_page'))
        
        # User query with soft delete check
        try:
            # Check if is_deleted column exists
            from sqlalchemy import inspect
            inspector = inspect(User)
            has_is_deleted = any(col.name == 'is_deleted' for col in inspector.columns)
            
            if has_is_deleted:
                users = User.query.filter(
                    User.is_active == True,
                    (User.is_deleted == False) | (User.is_deleted.is_(None))
                ).order_by(User.full_name.asc(), User.username.asc()).all()
            else:
                users = User.query.filter(User.is_active == True).order_by(User.full_name.asc(), User.username.asc()).all()
        except Exception as e:
            # Fallback to simple query if there's any issue
            users = User.query.filter(User.is_active == True).order_by(User.full_name.asc(), User.username.asc()).all()
        
        # Alt projeleri koduna göre büyükten küçüğe, en son eklenen en üste sırala
        subprojects_raw = SubProject.query.filter(SubProject.project_id == p.id).all()
        # Python'da sırala: kodun sonundaki sayısal kısmına göre büyükten küçüğe, sonra created_at DESC
        import re
        def get_code_numeric_suffix(code_str):
            if not code_str:
                return -1  # NULL kodlar en alta
            # Kodun sonundaki sayısal kısmı bul (örn: "9025-0002-05" -> 5, "05_v" -> 5)
            parts = code_str.split('-')
            if len(parts) > 0:
                last_part = parts[-1]
                # Son sayısal kısmı bul
                numbers = re.findall(r'\d+', last_part)
                if numbers:
                    return int(numbers[-1])  # Son sayısal kısmı
            return 0
        
        subprojects = sorted(
            subprojects_raw,
            key=lambda sp: (
                -get_code_numeric_suffix(sp.code),  # Kod suffix'ine göre büyükten küçüğe (05 > 04 > 03)
                -(sp.created_at.timestamp() if sp.created_at else 0)  # En son eklenen en üste
            ),
            reverse=False
        )

        subproject_usage = {}
        try:
            sp_ids = [int(sp.id) for sp in subprojects if sp and sp.id]
            if sp_ids:
                cell_counts = dict(
                    db.session.query(PlanCell.subproject_id, db.func.count(PlanCell.id))
                    .filter(PlanCell.subproject_id.in_(sp_ids))
                    .group_by(PlanCell.subproject_id)
                    .all()
                )
                job_counts = dict(
                    db.session.query(Job.subproject_id, db.func.count(Job.id))
                    .filter(Job.subproject_id.in_(sp_ids))
                    .group_by(Job.subproject_id)
                    .all()
                )
                for sid in sp_ids:
                    subproject_usage[sid] = (int(cell_counts.get(sid, 0) or 0) + int(job_counts.get(sid, 0) or 0)) > 0
        except Exception:
            subproject_usage = {}

        return render_template("project_edit.html", p=p, users=users, subprojects=subprojects, subproject_usage=subproject_usage)
    except Exception as e:
        import traceback
        print(f"Project edit error: {str(e)}")
        print(traceback.format_exc())
        flash(f"Proje düzenleme sayfası yüklenirken bir hata oluştu: {str(e)}", "danger")
        return redirect(url_for('planner.projects_page'))


@planner_bp.post("/projects/<int:project_id>/subprojects/add")
@login_required
@observer_required
def subproject_add(project_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    p = Project.query.get_or_404(project_id)
    name = (request.form.get("name") or "").strip()
    send_notify = (request.form.get("send_notify") == "1")
    code_input = (request.form.get("code") or "").strip()
    code = code_input or None
    if not name:
        flash("Alt proje adi zorunlu.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    dup_name = (
        SubProject.query
        .filter(SubProject.project_id == p.id, db.func.lower(SubProject.name) == name.lower())
        .first()
    )
    if dup_name:
        flash("Bu proje icin alt proje adi zaten var.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))
    
    # Global Check: Aynı proje koduna sahip başka bir alt proje (global olarak) olmamalı (istek 1 ve 4)
    # Ancak "proje ve alt proje açarken kod tek bir tane açılmalı" diyor kullanımı.
    # Proje kodu zaten Project tablosunda. Alt proje kodu SubProject tablosunda.
    # Kullanıcı muhtemelen SubProject kodu benzersiz olsun istiyor.
    # Kontrol edelim:
    if code:
        # Sadece bu projede değil, TÜM projelerde bu alt proje kodu var mı?
        # "1-Proje ve alt proje açarken koddam tek bir tane açılmalı" -> Global unique.
        existing_global = SubProject.query.filter(db.func.lower(SubProject.code) == code.lower()).first()
        if existing_global:
             flash(f"Hata: '{code}' kodlu bir alt proje zaten mevcut (Başka bir projede olabilir)!", "danger")
             return redirect(url_for("planner.project_edit", project_id=project_id))

    if not code:
        code = _next_subproject_code(p.id, p.project_code)

    if code:
        dup = SubProject.query.filter(SubProject.project_id == p.id, db.func.lower(SubProject.code) == code.lower()).first()
        if dup:
            flash("Bu proje icin alt proje kodu zaten var.", "danger")
            return redirect(url_for("planner.project_edit", project_id=project_id))

    sp = SubProject(project_id=p.id, name=name, code=code, is_active=True)
    db.session.add(sp)
    db.session.commit()
    if send_notify:
        current_user = get_current_user()
        mail_cfg = dict(load_mail_settings())
        notify_to = (mail_cfg.get("notify_to") or "").strip()
        notify_cc = (mail_cfg.get("notify_cc") or "").strip()
        recipient = notify_to or None
        if not recipient:
            responsible_user = User.query.filter(
                or_(User.full_name == p.responsible, User.username == p.responsible)
            ).first()
            if responsible_user and responsible_user.email:
                recipient = responsible_user.email
        if not recipient:
            # Fallback: Adminleri bul
            admin_users = User.query.filter(User.role == 'admin', User.is_active == True).all()
            for admin in admin_users:
                if admin.email and "@" in admin.email:
                    recipient = admin.email
                    break
        
        if not recipient:
            flash("Alt proje bildirimi için geçerli bir alıcı (Sorumlu veya Admin) bulunamadı; mail gönderilmedi.", "warning")
        else:
            actor = (current_user.full_name if current_user and current_user.full_name else
                     (current_user.username if current_user and current_user.username else "Bilinmiyor"))
            now = datetime.now()
            subject = f"Alt Proje Eklendi: {p.project_code}/{p.project_name} -> {sp.name}"
            
            context = {
                "heading": "Yeni Alt Proje Eklendi",
                "intro": f"Projeye yeni bir alt proje eklendi: <strong>{sp.name}</strong>",
                "subproject_name": sp.name,
                "project_name": p.name,
                "created_at": now.strftime("%d.%m.%Y %H:%M"),
                "actor": actor,
                "footer": "Bu bildirim Görev Takip Sistemi tarafından otomatik olarak gönderilmiştir.",
                "table_headers": ["Bilgi", "Değer"],
                "table_rows": [
                    ["Ana Proje", f"{p.name} ({p.project_code})"],
                    ["Alt Proje", sp.name],
                    ["Ekleyen", actor]
                ]
            }

            import html
            # Fallback için body içeriği
            body_content = f"""
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 12px;">
                <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                    <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Ana Proje</div>
                    <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(p.project_name)} ({html.escape(p.project_code)})</div>
                </div>
                <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                    <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Alt Proje</div>
                    <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(sp.name)}</div>
                </div>
                <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px;">
                    <div style="font-size: 11px; color: #64748b; text-transform: uppercase; margin-bottom: 4px; font-weight: 600;">Ekleyen</div>
                    <div style="font-size: 14px; color: #1e293b; font-weight: 600;">{html.escape(actor)}</div>
                </div>
            </div>
            """

            try:
                html_body = render_template(
                    "email_base.html",
                    title=subject,
                    heading="Yeni Alt Proje Eklendi",
                    intro=f"Projeye yeni bir alt proje eklendi: <strong>{html.escape(sp.name)}</strong>",
                    body_html=body_content,
                    footer=context.get("footer", "Bu bildirim Görev Takip Sistemi tarafından otomatik olarak gönderilmiştir."),
                )
                
                ok = MailService.send(
                    mail_type="subproject",
                    recipients=[recipient],
                    subject=subject,
                    html=html_body,
                    cc=notify_cc or None,
                    user_id=getattr(current_user, "id", None),
                    project_id=p.id,
                    meta={"type": "subproject", "project_id": p.id, "subproject_id": sp.id},
                )
            except Exception as ex:
                current_app.logger.error(f"Alt proje bildirimi gonderilemedi ({recipient}): {ex}")
                ok = False

            if ok:
                flash("Alt proje bildirimi gonderildi.", "success")
            else:
                flash("Alt proje bildirimi gonderilemedi. Detay icin loglari kontrol edin.", "danger")
    return redirect(url_for("planner.project_edit", project_id=project_id) + "#subprojects-table")


def _next_subproject_code(project_id: int, project_code: str) -> str:
    base = (project_code or "").strip()
    if not base:
        base = f"PRJ{project_id}"
    prefix = f"{base}-"
    lower_prefix = prefix.lower()
    suffixes = []
    for (existing_code,) in (
        db.session.query(SubProject.code)
        .filter(SubProject.project_id == project_id)
        .all()
    ):
        if not existing_code:
            continue
        normalized = existing_code.strip()
        if normalized.lower().startswith(lower_prefix):
            tail = normalized[len(prefix):]
            if tail.isdigit():
                suffixes.append(int(tail))
    next_idx = max(suffixes) + 1 if suffixes else 1  # İlk alt proje 01 ile başlar
    return f"{base}-{next_idx:02d}"


@planner_bp.post("/projects/<int:project_id>/subprojects/edit")
@login_required
@observer_required
def subproject_edit(project_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    subproject_id = int(request.form.get("subproject_id", 0) or 0)
    sp = SubProject.query.get_or_404(subproject_id)
    if int(sp.project_id or 0) != int(project_id):
        flash("Alt proje / proje eslesmedi.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Alt proje adi zorunlu.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    dup = (
        SubProject.query
        .filter(SubProject.project_id == project_id, db.func.lower(SubProject.name) == name.lower(), SubProject.id != sp.id)
        .first()
    )
    if dup:
        flash("Bu proje icin alt proje adi zaten var.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    sp.name = name
    db.session.commit()
    flash("Alt proje guncellendi.", "success")
    return redirect(url_for("planner.project_edit", project_id=project_id) + "#subprojects-table")


@planner_bp.post("/projects/<int:project_id>/subprojects/<int:subproject_id>/toggle")
@login_required
@observer_required
def subproject_toggle(project_id: int, subproject_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))
    sp = SubProject.query.get_or_404(subproject_id)
    if int(sp.project_id or 0) != int(project_id):
        flash("Alt proje / proje eslesmedi.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    is_active = (request.form.get("is_active") or "").strip() == "1"
    sp.is_active = bool(is_active)
    db.session.commit()
    flash("Alt proje durumu guncellendi.", "success")
    return redirect(url_for("planner.project_edit", project_id=project_id) + "#subprojects-table")


@planner_bp.post("/projects/<int:project_id>/subprojects/<int:subproject_id>/delete")
@login_required
@observer_required
def subproject_delete(project_id: int, subproject_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))
    sp = SubProject.query.get_or_404(subproject_id)
    if int(sp.project_id or 0) != int(project_id):
        flash("Alt proje / proje eslesmedi.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id))

    used_cells = PlanCell.query.filter(PlanCell.subproject_id == sp.id).count()
    used_jobs = Job.query.filter(Job.subproject_id == sp.id).count()
    if (used_cells or 0) > 0 or (used_jobs or 0) > 0:
        flash("Bu alt proje daha once islerde kullanilmis. Silinemez, pasif yapin.", "danger")
        return redirect(url_for("planner.project_edit", project_id=project_id) + "#subprojects-table")

    db.session.delete(sp)
    db.session.commit()
    flash("Alt proje silindi.", "success")
    return redirect(url_for("planner.project_edit", project_id=project_id) + "#subprojects-table")



# ---------- PEOPLE ----------
@planner_bp.route("/people", methods=["GET", "POST"])
@login_required
@observer_required
def people_page():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            flash("Personel adı zorunlu.", "danger")
            return redirect(url_for('planner.people_page'))

        firma_id = request.form.get("firma_id", "").strip()
        seviye_id = request.form.get("seviye_id", "").strip()
        durum = request.form.get("durum", "Aktif").strip()
        
        db.session.add(Person(
            full_name=full_name,
            tc_no=request.form.get("tc_no", "").strip(),
            role=request.form.get("role", "").strip(),
            email=request.form.get("email", "").strip(),
            phone=request.form.get("phone", "").strip(),
            firma_id=int(firma_id) if firma_id else None,
            seviye_id=int(seviye_id) if seviye_id else None,
            durum=durum,
        ))
        db.session.commit()
        flash("Personel eklendi.", "success")
        return redirect(url_for('planner.people_page'))

    people = Person.query.order_by(Person.full_name.asc()).all()
    firmalar = Firma.query.order_by(Firma.name.asc()).all()  # Tüm firmalar (aktif/pasif)
    seviyeler = Seviye.query.order_by(Seviye.name.asc()).all()  # Tüm seviyeler (aktif/pasif)
    # Check if current user is kivanc
    current_user = get_current_user()
    is_kivanc = current_user and (current_user.username == 'kivanc' or current_user.email == 'kivancozcan@netmon.com.tr')
    return render_template("people.html", people=people, firmalar=firmalar, seviyeler=seviyeler, is_kivanc=is_kivanc)


@planner_bp.get("/people/export.xlsx")
@login_required
@observer_required
def people_excel():
    people = Person.query.order_by(Person.full_name.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Personel"
    
    # Header
    headers = ["Ad Soyad", "TC", "Telefon", "Email", "Görev", "Firma", "Seviye", "Durum"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for row_idx, p in enumerate(people, 2):
        ws.cell(row=row_idx, column=1, value=p.full_name or "")
        ws.cell(row=row_idx, column=2, value=p.tc_no or "")
        ws.cell(row=row_idx, column=3, value=p.phone or "")
        ws.cell(row=row_idx, column=4, value=p.email or "")
        ws.cell(row=row_idx, column=5, value=p.role or "")
        ws.cell(row=row_idx, column=6, value=p.firma.name if p.firma else "")
        ws.cell(row=row_idx, column=7, value=p.seviye.name if p.seviye else "")
        ws.cell(row=row_idx, column=8, value=p.durum or "")
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"personel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    as_attachment=True, download_name=filename)


@planner_bp.route("/people/<int:person_id>/edit", methods=["GET", "POST"])
@kivanc_required
def person_edit(person_id: int):
    p = Person.query.get_or_404(person_id)
    
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            flash("Personel adı zorunlu.", "danger")
            return redirect(url_for("person_edit", person_id=person_id))
        
        firma_id = request.form.get("firma_id", "").strip()
        seviye_id = request.form.get("seviye_id", "").strip()
        durum = request.form.get("durum", "Aktif").strip()
        
        p.full_name = full_name
        p.tc_no = request.form.get("tc_no", "").strip() or None
        p.role = request.form.get("role", "").strip() or None
        p.email = request.form.get("email", "").strip() or None
        p.phone = request.form.get("phone", "").strip() or None
        p.firma_id = int(firma_id) if firma_id else None
        p.seviye_id = int(seviye_id) if seviye_id else None
        p.durum = durum
        
        db.session.commit()
        flash("Personel güncellendi.", "success")
        return redirect(url_for('planner.people_page'))
    
    firmalar = Firma.query.order_by(Firma.name.asc()).all()
    seviyeler = Seviye.query.order_by(Seviye.name.asc()).all()
    return render_template("person_edit.html", p=p, firmalar=firmalar, seviyeler=seviyeler)


@planner_bp.post("/people/<int:person_id>/delete")
@login_required
@kivanc_required
def person_delete(person_id: int):
    p = Person.query.get_or_404(person_id)
    
    # İlgili kayıtları temizle (Foreign Key hatasını önlemek için)
    # 1. JobAssignment
    JobAssignment.query.filter_by(person_id=person_id).delete()
    
    # 2. CellAssignment
    CellAssignment.query.filter_by(person_id=person_id).delete()
    
    # 3. PersonDayStatus
    PersonDayStatus.query.filter_by(person_id=person_id).delete()
    
    # 4. TeamOvertime (person_id null yapılabilir)
    TeamOvertime.query.filter_by(person_id=person_id).update({"person_id": None})
    
    db.session.delete(p)
    db.session.commit()
    flash("Personel ve ilişkili atamalar silindi.", "success")
    return redirect(url_for('planner.people_page'))


# ---------- TANIMLAR: FIRMA ----------
@planner_bp.route("/tanimlar/firma", methods=["GET", "POST"])
@login_required
@observer_required
def firma_page():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Firma adı zorunlu.", "danger")
            return redirect(url_for("firma_page"))
        
        is_active = (request.form.get("is_active") == "1")
        gender = request.form.get("gender", "").strip() or None
        email = request.form.get("email", "").strip() or None
        mail_recipient_name = request.form.get("mail_recipient_name", "").strip() or None
        
        db.session.add(Firma(name=name, is_active=is_active, gender=gender, email=email, mail_recipient_name=mail_recipient_name))
        db.session.commit()
        flash("Firma eklendi.", "success")
        return redirect(url_for("firma_page"))
    
    firmalar = Firma.query.order_by(Firma.name.asc()).all()
    return render_template("firma.html", firmalar=firmalar)


@planner_bp.route("/tanimlar/firma/<int:firma_id>/edit", methods=["GET", "POST"])
@login_required
@observer_required
def firma_edit(firma_id: int):
    f = Firma.query.get_or_404(firma_id)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Firma adı zorunlu.", "danger")
            return redirect(url_for("firma_edit", firma_id=firma_id))
        
        f.name = name
        f.is_active = (request.form.get("is_active") == "1")
        f.gender = request.form.get("gender", "").strip() or None
        f.email = request.form.get("email", "").strip() or None
        f.mail_recipient_name = request.form.get("mail_recipient_name", "").strip() or None
        db.session.commit()
        flash("Firma güncellendi.", "success")
        # Eğer people sayfasından geliyorsa oraya dön
        if request.referrer and 'people' in request.referrer:
            return redirect(url_for('planner.people_page') + "?tab=firma")
        return redirect(url_for("firma_page"))
    
    return render_template("firma_edit.html", f=f)


@planner_bp.post("/tanimlar/firma/<int:firma_id>/delete")
@login_required
@observer_required
def firma_delete(firma_id: int):
    f = Firma.query.get_or_404(firma_id)
    # Check if firma is used by any person
    if Person.query.filter_by(firma_id=firma_id).count() > 0:
        flash("Bu firma personeller tarafından kullanılıyor, silinemez.", "danger")
        if request.referrer and 'people' in request.referrer:
            return redirect(url_for('planner.people_page') + "?tab=firma")
        return redirect(url_for("firma_page"))
    db.session.delete(f)
    db.session.commit()
    flash("Firma silindi.", "success")
    # Eğer people sayfasından geliyorsa oraya dön
    if request.referrer and 'people' in request.referrer:
        return redirect(url_for('planner.people_page') + "?tab=firma")
    return redirect(url_for("firma_page"))


# ---------- TANIMLAR: SEVIYE ----------
@planner_bp.route("/tanimlar/seviye", methods=["GET", "POST"])
@login_required
@observer_required
def seviye_page():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Seviye adı zorunlu.", "danger")
            return redirect(url_for("people_page", tab="seviye"))
        
        is_active = (request.form.get("is_active") == "1")
        db.session.add(Seviye(name=name, is_active=is_active))
        db.session.commit()
        flash("Seviye eklendi.", "success")
        return redirect(url_for('planner.people_page') + "?tab=seviye")
    
    seviyeler = Seviye.query.order_by(Seviye.name.asc()).all()
    return render_template("seviye.html", seviyeler=seviyeler)


@planner_bp.post("/tanimlar/seviye/<int:seviye_id>/delete")
@login_required
@observer_required
def seviye_delete(seviye_id: int):
    s = Seviye.query.get_or_404(seviye_id)
    # Check if seviye is used by any person
    if Person.query.filter_by(seviye_id=seviye_id).count() > 0:
        flash("Bu seviye personeller tarafından kullanılıyor, silinemez.", "danger")
        if request.referrer and 'people' in request.referrer:
            return redirect(url_for('planner.people_page') + "?tab=seviye")
        return redirect(url_for("seviye_page"))
    db.session.delete(s)
    db.session.commit()
    flash("Seviye silindi.", "success")
    # Eğer people sayfasından geliyorsa oraya dön
    if request.referrer and 'people' in request.referrer:
        return redirect(url_for('planner.people_page') + "?tab=seviye")
    return redirect(url_for("seviye_page"))


# ---------- VEHICLES (ARAÇLAR) ----------
@planner_bp.route("/tools", methods=["GET", "POST"])
@login_required
@observer_required
def tools_page():
    if request.method == "POST":
        plate = request.form.get("plate", "").strip().upper()
        brand = request.form.get("brand", "").strip()
        model = request.form.get("model", "").strip()
        year_str = request.form.get("year", "").strip()
        vehicle_type = request.form.get("vehicle_type", "").strip()
        status = request.form.get("status", "available").strip()
        capacity_str = request.form.get("capacity", "").strip()
        vodafone_approval = str(request.form.get("vodafone_approval", "0") or "0").strip() in {"1", "true", "on"}
        notes = request.form.get("notes", "").strip()
        
        if not plate or not brand:
            flash("Plaka ve marka zorunlu.", "danger")
            return redirect(url_for('planner.tools_page'))
        
        # Check if plate already exists
        existing = Vehicle.query.filter_by(plate=plate).first()
        if existing:
            flash("Bu plaka zaten kayıtlı.", "danger")
            return redirect(url_for('planner.tools_page'))
        
        year = None
        if year_str:
            try:
                year = int(year_str)
            except ValueError:
                pass
        capacity = None
        if capacity_str:
            try:
                capacity = int(capacity_str)
            except ValueError:
                capacity = None

        db.session.add(Vehicle(
            plate=plate,
            brand=brand,
            model=model if model else None,
            year=year,
            vehicle_type=vehicle_type if vehicle_type else None,
            status=status if status else "available",
            capacity=capacity,
            vodafone_approval=vodafone_approval,
            notes=notes if notes else None,
            created_at=datetime.now()
        ))
        db.session.commit()
        flash("Araç eklendi.", "success")
        return redirect(url_for('planner.tools_page'))
    
    # Hafta hesaplama
    from datetime import timedelta
    week_start_param = request.args.get("week_start", "")
    if week_start_param:
        try:
            selected_week = date.fromisoformat(week_start_param)
            # Pazartesi'ye yuvarla
            current_week_start = selected_week - timedelta(days=selected_week.weekday())
        except:
            current_week_start = date.today() - timedelta(days=date.today().weekday())
    else:
        current_week_start = date.today() - timedelta(days=date.today().weekday())
    
    current_week_end = current_week_start + timedelta(days=6)
    prev_week_start = current_week_start - timedelta(days=7)
    prev_week_end = prev_week_start + timedelta(days=6)
    next_week_start = current_week_start + timedelta(days=7)
    
    vehicles = Vehicle.query.order_by(Vehicle.plate.asc()).all()
    
    # Araç-Ekip eşleştirmesi için ekipleri al (statik atama)
    teams_with_vehicles = Team.query.filter(Team.vehicle_id != None).all()
    vehicle_team_map = {t.vehicle_id: t for t in teams_with_vehicles}
    
    # Bu hafta için haftalık atamalar
    from models import VehicleAssignment
    current_assignments = VehicleAssignment.query.filter(
        VehicleAssignment.week_start == current_week_start,
        VehicleAssignment.is_active == True
    ).all()
    assignment_map = {a.vehicle_id: a for a in current_assignments}
    
    # Önceki hafta atamaları (tooltip için)
    prev_assignments = VehicleAssignment.query.filter(
        VehicleAssignment.week_start == prev_week_start
    ).all()
    prev_assignment_map = {a.vehicle_id: a for a in prev_assignments}
    
    # Her araç için ekip ve atama bilgisini ekle
    vehicles_with_teams = []
    for v in vehicles:
        assignment = assignment_map.get(v.id)
        prev_assignment = prev_assignment_map.get(v.id)
        vehicles_with_teams.append({
            'vehicle': v,
            'team': vehicle_team_map.get(v.id),
            'assignment': assignment,
            'assigned_person': assignment.person if assignment else None,
            'secondary_person': assignment.secondary_person if assignment else None,
            'prev_assignment': prev_assignment,
            'prev_person': prev_assignment.person if prev_assignment else None
        })
    
    # Tüm personelleri al (atama için dropdown)
    people = Person.query.filter(Person.durum == "Aktif").order_by(Person.full_name).all()
    
    return render_template("tools.html", 
                           vehicles=vehicles, 
                           vehicles_with_teams=vehicles_with_teams,
                           people=people,
                           current_week_start=current_week_start,
                           current_week_end=current_week_end,
                           prev_week_start=prev_week_start,
                           next_week_start=next_week_start)


@planner_bp.post("/tools/<int:vehicle_id>/delete")
@login_required
@observer_required
def vehicle_delete(vehicle_id: int):
    v = Vehicle.query.get_or_404(vehicle_id)
    db.session.delete(v)
    db.session.commit()
    flash("Araç silindi.", "success")
    return redirect(url_for('planner.tools_page'))


@planner_bp.route("/tools/<int:vehicle_id>/edit", methods=["GET", "POST"])
@login_required
@observer_required
def vehicle_edit(vehicle_id: int):
    v = Vehicle.query.get_or_404(vehicle_id)
    if request.method == "POST":
        plate = request.form.get("plate", "").strip().upper()
        brand = request.form.get("brand", "").strip()
        model = request.form.get("model", "").strip()
        year_str = request.form.get("year", "").strip()
        vehicle_type = request.form.get("vehicle_type", "").strip()
        status = request.form.get("status", "available").strip()
        capacity_str = request.form.get("capacity", "").strip()
        vodafone_approval = str(request.form.get("vodafone_approval", "0") or "0").strip() in {"1", "true", "on"}
        notes = request.form.get("notes", "").strip()
        
        if not plate or not brand:
            flash("Plaka ve marka zorunlu.", "danger")
            return redirect(url_for("vehicle_edit", vehicle_id=vehicle_id))
        
        # Check if plate already exists for another vehicle
        existing = Vehicle.query.filter(Vehicle.plate == plate, Vehicle.id != vehicle_id).first()
        if existing:
            flash("Bu plaka başka bir araçta kullanılıyor.", "danger")
            return redirect(url_for("vehicle_edit", vehicle_id=vehicle_id))
        
        v.plate = plate
        v.brand = brand
        v.model = model if model else None
        v.vehicle_type = vehicle_type if vehicle_type else None
        v.status = status if status else "available"
        v.capacity = None
        if capacity_str:
            try:
                v.capacity = int(capacity_str)
            except ValueError:
                v.capacity = None
        v.vodafone_approval = vodafone_approval
        v.notes = notes if notes else None
        
        if year_str:
            try:
                v.year = int(year_str)
            except ValueError:
                v.year = None
        else:
            v.year = None
        
        db.session.commit()
        flash("Araç güncellendi.", "success")
        return redirect(url_for('planner.tools_page'))
    
    return render_template("vehicle_edit.html", v=v)


# ---------- API: CELL GET/SET/CLEAR ----------
@planner_bp.get("/api/cell")
def api_cell_get():
    try:
        try:
            log.debug("api_cell_get request.args: %s", request.args)
        except Exception:
            pass
        project_id = int(request.args.get("project_id", 0))
        date_str = request.args.get("date", "")
        if not project_id or not date_str:
            return jsonify({"ok": False, "error": "Parametreler eksik (project_id, date)"}), 400
            
        d = parse_date(date_str) or date.today()
        ws = parse_date(request.args.get("week_start", ""))
        cell = PlanCell.query.filter_by(project_id=project_id, work_date=d).first()
        if not cell:
            return jsonify({"exists": False, "cell": None, "assigned": []})

        assigned = [a.person_id for a in cell.assignments]
        team = Team.query.get(cell.team_id) if cell.team_id else None
        team_vehicle = _vehicle_payload(team.vehicle if team else None, cell.team_id) if team else None
    
        # Alt proje bilgisini al
        subproject_label = ""
        subproject_id = getattr(cell, "subproject_id", None)
        if subproject_id:
            subproject = SubProject.query.get(subproject_id)
            if subproject:
                code = (subproject.code or "").strip()
                subproject_label = f"{subproject.name}{f' ({code})' if code else ''}"

        # Get project info
        project = Project.query.get(cell.project_id)
        project_name = project.project_name if project else ""

        # Get person overtimes (hours/shifts)
        person_overtimes = {}
        t_ots = TeamOvertime.query.filter_by(cell_id=cell.id).all()
        for ot in t_ots:
            if ot.description:
                person_overtimes[int(ot.person_id)] = ot.description.strip()
            # fallback if description is empty but duration > 0? 
            # Ideally description holds the hours string "2" or "2.5"

        # Haftalık dinamik ekip adı hesapla
        from utils import get_weekly_team_display_name
        team_display_name = ""
        if cell.team_id:
            # Hafta başlangıcını hesapla
            week_start_date = ws if ws else week_start(d)
            team_display_name = get_weekly_team_display_name(cell.team_id, week_start_date)
        
        # Fallback: Eğer dinamik ad boşsa, mevcut team_name veya veritabanındaki adı kullan
        if not team_display_name:
            team_display_name = cell.team_name or (team.name if team else "")

        return jsonify({
            "exists": True,
            "team_vehicle": team_vehicle,
            "cell": {
                "id": cell.id,
                "project_id": cell.project_id,
                "project_name": project_name,
                "work_date": iso(cell.work_date),
                "subproject_id": subproject_id,
                "subproject_label": subproject_label,
                "shift": cell.shift or "",
                "vehicle_info": cell.vehicle_info or "",
                "note": cell.note or "",
                "isdp_info": getattr(cell, "isdp_info", "") or "",
                "po_info": getattr(cell, "po_info", "") or "",
                "important_note": getattr(cell, "important_note", "") or "",
                "team_id": cell.team_id or None,
                "team_name": team_display_name,  # Haftalık dinamik ekip adı
                "assigned_user_id": getattr(cell, "assigned_user_id", None),
                "job_mail_body": getattr(cell, "job_mail_body", "") or "",
                "lld_hhd_files": _parse_files(getattr(cell, "lld_hhd_files", None)) or ([cell.lld_hhd_path] if getattr(cell, "lld_hhd_path", None) else []),
                "tutanak_files": _parse_files(getattr(cell, "tutanak_files", None)) or ([cell.tutanak_path] if getattr(cell, "tutanak_path", None) else []),
            },
            "assigned": assigned,
            "person_overtimes": person_overtimes
        })
    except Exception as e:
        import traceback

        err_msg = traceback.format_exc()
        print(f"ERROR in api_cell_get: {str(e)}")
        with open("c:\\Users\\USER\\Desktop\\Saha\\error_log.txt", "w") as f:
            f.write(err_msg)
        return jsonify({"error": str(e), "exists": False, "cell": None, "assigned": []}), 500


@planner_bp.get("/api/subprojects")
@login_required
def api_subprojects():
    project_id = int(request.args.get("project_id", 0) or 0)
    if not project_id:
        return jsonify({"ok": False, "error": "project_id eksik"}), 400

    include_inactive = str(request.args.get("include_inactive", "0") or "0").strip() in {"1", "true", "yes"}
    owner_project_id = _effective_main_project_id_for_subprojects(project_id)
    if not owner_project_id:
        return jsonify({"ok": True, "project_id": 0, "subprojects": []})

    q = SubProject.query.filter(SubProject.project_id == owner_project_id)
    if not include_inactive:
        q = q.filter(SubProject.is_active == True)
    rows_raw = q.all()

    # Kodun sayısal suffix'ine göre DESC, sonra created_at DESC
    import re
    def get_code_numeric_suffix(code_str):
        if not code_str:
            return -1
        parts = code_str.split("-")
        if parts:
            last_part = parts[-1]
            nums = re.findall(r"\d+", last_part)
            if nums:
                return int(nums[-1])
        return 0

    rows = sorted(
        rows_raw,
        key=lambda sp: (
            -get_code_numeric_suffix(sp.code),
            -(sp.created_at.timestamp() if getattr(sp, "created_at", None) else 0),
        ),
        reverse=False
    )

    return jsonify({
        "ok": True,
        "project_id": owner_project_id,
        "subprojects": [
            {"id": int(sp.id), "name": sp.name, "code": (sp.code or ""), "is_active": bool(sp.is_active)}
            for sp in rows
        ],
    })


@planner_bp.get("/api/projects/<int:project_id>/subprojects")
@login_required
def api_project_subprojects(project_id: int):
    include_inactive = str(request.args.get("include_inactive", "0") or "0").strip() in {"1", "true", "yes"}
    owner_project_id = _effective_main_project_id_for_subprojects(int(project_id or 0))
    if not owner_project_id:
        return jsonify({"ok": True, "project_id": 0, "requested_project_id": project_id, "subprojects": []})

    # Debug: Proje bilgilerini kontrol et
    requested_project = Project.query.get(project_id) if project_id else None
    owner_project = Project.query.get(owner_project_id) if owner_project_id else None

    q = SubProject.query.filter(SubProject.project_id == owner_project_id)
    if not include_inactive:
        q = q.filter(SubProject.is_active == True)
    rows_raw = q.all()
    
    # Güvenlik: Alt proje kodlarının proje koduna uyumunu kontrol et
    verified_rows = []
    owner_project_code = owner_project.project_code if owner_project else None
    
    for sp in rows_raw:
        # Alt projenin project_id'si owner_project_id ile eşleşmeli
        if sp.project_id != owner_project_id:
            import logging
            logging.warning(f"Subproject {sp.id} ({sp.code}) belongs to project {sp.project_id} but was requested for project {project_id} (owner: {owner_project_id})")
            continue
        
        # Alt proje kodunun proje koduna uyumunu kontrol et (eğer kod varsa)
        if sp.code and owner_project_code:
            # Alt proje kodu proje kodundan başlamalı (örn: "9026-0011-01" -> "9026-0011")
            sp_code_prefix = sp.code.split('-')[0] + '-' + sp.code.split('-')[1] if '-' in sp.code else sp.code
            if not sp_code_prefix.startswith(owner_project_code):
                import logging
                logging.warning(f"Subproject {sp.id} code '{sp.code}' doesn't match project code '{owner_project_code}' for project {project_id}")
                continue
        
        verified_rows.append(sp)
    
    rows_raw = verified_rows
    
    # Python'da sırala: kodun sonundaki sayısal kısmına göre büyükten küçüğe, sonra created_at DESC
    import re
    def get_code_numeric_suffix(code_str):
        if not code_str:
            return -1  # NULL kodlar en alta
        # Kodun sonundaki sayısal kısmı bul (örn: "9025-0002-05" -> 5, "05_v" -> 5)
        parts = code_str.split('-')
        if len(parts) > 0:
            last_part = parts[-1]
            # Son sayısal kısmı bul
            numbers = re.findall(r'\d+', last_part)
            if numbers:
                return int(numbers[-1])  # Son sayısal kısmı
        return 0
    
    rows = sorted(
        rows_raw,
        key=lambda sp: (
            -get_code_numeric_suffix(sp.code),  # Kod suffix'ine göre büyükten küçüğe (05 > 04 > 03)
            -(sp.created_at.timestamp() if sp.created_at else 0)  # En son eklenen en üste
        ),
        reverse=False
    )

    return jsonify({
        "ok": True,
        "project_id": owner_project_id,
        "requested_project_id": project_id,  # Debug için gönderilen ID'yi de döndür
        "subprojects": [
            {"id": int(sp.id), "name": sp.name, "code": (sp.code or ""), "is_active": bool(sp.is_active), "created_at": iso(sp.created_at) if sp.created_at else None}
            for sp in rows
        ],
    })


@planner_bp.get("/api/assignments_week")
@login_required
def api_assignments_week():
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)

    assignments = {}
    cells = PlanCell.query.filter(PlanCell.work_date >= start, PlanCell.work_date <= end).all()
    for cell in cells:
        d = iso(cell.work_date)
        if d not in assignments:
            assignments[d] = set()
        # assignments ilişkisi CellAssignment objeleri döner; person_id'leri ekleyelim
        for ass in cell.assignments:
            if ass and ass.person_id:
                assignments[d].add(int(ass.person_id))

    # dict values set'ten list'e çevir
    for d in assignments:
        assignments[d] = list(assignments[d])

    return jsonify({"ok": True, "assignments": assignments})

# ---------- NOTIFICATIONS ----------
SLA_DAYS = 1
_last_sla_check = None


def _notify_user(*, user_id: int, event: str, title: str, body: str = "", link_url: str = None,
                 job_id: int = None, mail_log_id: int = None, meta: dict = None):
    try:
        import json
        n = Notification(
            user_id=int(user_id),
            event=(event or "").strip() or "event",
            title=(title or "").strip() or "Bildirim",
            body=(body or "").strip() or None,
            link_url=(link_url or None),
            job_id=(int(job_id) if job_id else None),
            mail_log_id=(int(mail_log_id) if mail_log_id else None),
            meta_json=(json.dumps(meta or {}, ensure_ascii=False) if meta is not None else None),
        )
        db.session.add(n)
        return n
    except Exception:
        return None


def _notify_admins(*, event: str, title: str, body: str = "", link_url: str = None,
                   job_id: int = None, mail_log_id: int = None, meta: dict = None):
    try:
        admins = User.query.filter(
            User.is_active == True,
            or_(User.is_admin == True, User.role.in_(["admin", "planner"]))
        ).all()
    except Exception:
        admins = []
    for u in admins:
        _notify_user(user_id=u.id, event=event, title=title, body=body, link_url=link_url, job_id=job_id, mail_log_id=mail_log_id, meta=meta)


def _notify_users_by_emails(emails, **kwargs):
    emails = [e.strip().lower() for e in (emails or []) if e and isinstance(e, str) and "@" in e]
    if not emails:
        return []
    users = User.query.filter(db.func.lower(User.email).in_(emails)).all()
    out = []
    for u in users:
        n = _notify_user(user_id=u.id, **kwargs)
        if n:
            out.append(n)
    return out


def _check_sla_notifications():
    global _last_sla_check
    today = date.today()
    if _last_sla_check == today:
        return
    _last_sla_check = today

    try:
        cutoff = today - timedelta(days=SLA_DAYS)
        overdue = Job.query.filter(Job.status != "completed", Job.work_date <= cutoff).all()
        if not overdue:
            return

        ids = [j.id for j in overdue if j and j.id]
        existing = set()
        if ids:
            for (jid,) in db.session.query(Notification.job_id).filter(Notification.event == "sla", Notification.job_id.in_(ids)).distinct().all():
                if jid:
                    existing.add(int(jid))

        for j in overdue:
            if not j or not j.id or int(j.id) in existing:
                continue
            project = Project.query.get(j.project_id)
            title = "SLA aşıldı"
            body = "{} {} | {} | Ekip: {}".format(
                (project.project_code if project else ""),
                (project.region if project else ""),
                iso(j.work_date),
                (j.team_name or "-"),
            )
            _notify_admins(
                event="sla",
                title=title,
                body=body,
                link_url=url_for("assignment_page", job_id=j.id),
                job_id=j.id,
                meta={"project_id": j.project_id, "work_date": iso(j.work_date)},
            )

        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass



# ---------- API: MAIL TEMPLATES (client) ----------
@planner_bp.get("/api/mail_templates")
@login_required
def api_mail_templates():
    return jsonify({"ok": False, "error": "Mail şablonu özelliği kaldırıldı."}), 404


@planner_bp.get("/api/job_mail_last")
@login_required
def api_job_mail_last():
    project_id = int(request.args.get("project_id", 0) or 0)
    d = parse_date(request.args.get("date", "")) or date.today()
    if not project_id:
        return jsonify({"ok": False, "error": "project_id eksik"}), 400

    try:
        import json as _json
        rows = MailLog.query.filter_by(kind="send", ok=True).order_by(MailLog.created_at.desc()).limit(200).all()
        for r in rows:
            try:
                meta = _json.loads(r.meta_json or "{}")
            except Exception:
                meta = {}
            if meta.get("type") != "job":
                continue
            if int(meta.get("project_id") or 0) != project_id:
                continue
            if str(meta.get("work_date") or "") != iso(d):
                continue
            return jsonify({
                "ok": True,
                "found": True,
                "sent_at": r.created_at.isoformat() if r.created_at else None,
                "to": r.to_addr,
                "subject": r.subject
            })
    except Exception:
        pass

    return jsonify({"ok": True, "found": False})


@planner_bp.post("/api/cell")
@login_required
@observer_required
def api_cell_set():
    try:
        data = request.get_json(force=True, silent=True) or {}
        try:
            log.debug("api_cell_set data: %s", data)
        except Exception:
            pass
        
        project_id = int(data.get("project_id", 0))
        work_date_str = data.get("work_date", "")
        
        if not project_id or not work_date_str:
            return jsonify({"ok": False, "error": "Parametreler eksik (project_id, work_date)"}), 400
            
        d = parse_date(work_date_str) or date.today()
        subproject_id = int(data.get("subproject_id", 0) or 0)
        if subproject_id and not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id):
            return jsonify({"ok": False, "error": "Alt proje secimi gecersiz."}), 400

        shift = (data.get("shift") or "").strip()
        vehicle_info = (data.get("vehicle_info") or "").strip()
        note = (data.get("note") or "").strip()
        isdp_info = (data.get("isdp_info") or "").strip()
        po_info = (data.get("po_info") or "").strip()
        important_note = (data.get("important_note") or "").strip()
        team_name = (data.get("team_name") or "").strip()
        job_mail_body = (data.get("job_mail_body") or "").strip()
        assigned_user_id = int(data.get("assigned_user_id", 0) or 0)
        remove_lld_names = _parse_files(data.get("remove_lld_list"))
        remove_tutanak_names = _parse_files(data.get("remove_tutanak_list"))
        person_ids = [int(pid) for pid in data.get("person_ids", []) if pid]
        vehicle_dirty_raw = data.get("vehicle_dirty")
        if isinstance(vehicle_dirty_raw, str):
            vehicle_dirty = vehicle_dirty_raw.strip().lower() in {"1", "true", "yes"}
        else:
            vehicle_dirty = bool(vehicle_dirty_raw)
        vehicle_id_specified = "vehicle_id" in data
        requested_vehicle_id = None
        if vehicle_id_specified:
            raw_vehicle_id = data.get("vehicle_id")
            try:
                requested_vehicle_id = int(raw_vehicle_id or 0)
            except Exception:
                requested_vehicle_id = None
            if requested_vehicle_id is not None and requested_vehicle_id <= 0:
                requested_vehicle_id = None

        team_vehicle_requested = False
        team_vehicle_target_id = None
        if "team_vehicle_id" in data:
            team_vehicle_requested = True
            try:
                team_vehicle_target_id = int(data.get("team_vehicle_id") or 0)
            except Exception:
                team_vehicle_target_id = 0
        elif vehicle_id_specified:
            team_vehicle_requested = True
            team_vehicle_target_id = requested_vehicle_id or 0

        cell = ensure_cell(project_id, d)
        if subproject_id:
            cell.subproject_id = subproject_id
        else:
            cell.subproject_id = None
        cell.shift = shift if shift else None
        cell.vehicle_info = vehicle_info if vehicle_info else None
        cell.note = note if note else None
        cell.isdp_info = isdp_info if isdp_info else None
        cell.po_info = po_info if po_info else None
        cell.important_note = important_note if important_note else None
        cell.job_mail_body = job_mail_body if job_mail_body else None
        cell.assigned_user_id = assigned_user_id if assigned_user_id else None

        # assignment + ekip (team_id)
        added_ids = set_assignments_and_team(
            cell,
            person_ids,
            preferred_vehicle_info=vehicle_info if (vehicle_dirty or vehicle_info) else None,
        )

        # Per-person shift/overtime saving
        person_overtimes = data.get("person_overtimes") or {}
        
        # Get current user for created_by
        current_u_id = get_current_user().id if get_current_user() else 1
        
        # Existing records in a map
        existing_ots = {
            ot.person_id: ot 
            for ot in TeamOvertime.query.filter_by(cell_id=cell.id).all()
        }
        
        processed_pids = set()
        
        if isinstance(person_overtimes, dict):
            for pid_str, shift_val in person_overtimes.items():
                try:
                    pid = int(pid_str)
                    if pid not in person_ids:
                        continue # Skip if person is no longer assigned
                    
                    shift_str = (shift_val or "").strip()
                    # User might send "0" or empty string to clear
                    if not shift_str:
                         continue
                        
                    processed_pids.add(pid)
                    # Try to parse float hours from string "2", "2.5", etc.
                    try:
                        hours = float(shift_str)
                    except:
                        hours = 0.0
                    
                    ot = existing_ots.get(pid)
                    if not ot:
                        ot = TeamOvertime(
                            cell_id=cell.id, 
                            person_id=pid, 
                            work_date=d,
                            created_by_user_id=current_u_id
                        )
                        db.session.add(ot)
                    
                    ot.duration_hours = hours
                    ot.description = shift_str 
                    ot.updated_at = datetime.now()
                    
                except Exception as ex:
                    print(f"Error saving person overtime for {pid_str}: {ex}")
            
        # Cleanup: Remove overtime records for persons currently assigned BUT who have no overtime value in the payload (cleared)
        # Javascript sends ALL values. If value is empty, it might be missing from dict or empty string.
        # If it was missing from dict, it's not in processed_pids.
        for pid in person_ids:
            if pid not in processed_pids and pid in existing_ots:
                db.session.delete(existing_ots[pid])
        
        # Cleanup: Remove overtime records for persons NOT in the assignment list anymore
        for pid, ot in existing_ots.items():
            if pid not in person_ids:
                db.session.delete(ot)




        team = Team.query.get(cell.team_id) if cell.team_id else None
        if team and team_vehicle_requested:
            vehicle_for_team = None
            if team_vehicle_target_id and team_vehicle_target_id > 0:
                vehicle_for_team = Vehicle.query.get(team_vehicle_target_id)
                if not vehicle_for_team:
                    db.session.rollback()
                    return jsonify({"ok": False, "error": "Araç bulunamadı"}), 404
                conflict_team = Team.query.filter(
                    Team.vehicle_id == team_vehicle_target_id,
                    Team.id != team.id
                ).first()
                if conflict_team:
                    # Aracı yeni ekibe verebilmek için önce eski ekibin atamasını kaldır
                    conflict_team.vehicle_id = None
                    db.session.query(PlanCell).filter(
                        PlanCell.team_id == conflict_team.id
                    ).update({PlanCell.vehicle_info: None}, synchronize_session=False)
            team.vehicle_id = vehicle_for_team.id if vehicle_for_team else None
            plate = vehicle_for_team.plate if vehicle_for_team else None
            db.session.flush()
            db.session.query(PlanCell).filter(PlanCell.team_id == team.id).update(
                {PlanCell.vehicle_info: plate}, synchronize_session=False
            )
            cell.vehicle_info = plate

        # ekip adı (rapor için)

        # remove attachments if requested
        # remove selected attachments
        if remove_lld_names:
            current = _parse_files(cell.lld_hhd_files)
            remaining = [f for f in current if f not in remove_lld_names]
            for fname in current:
                if fname in remove_lld_names:
                    delete_upload(fname)
            cell.lld_hhd_files = _dump_files(remaining) if remaining else None
            if cell.lld_hhd_path and cell.lld_hhd_path in remove_lld_names:
                delete_upload(cell.lld_hhd_path)
                cell.lld_hhd_path = None
        if remove_tutanak_names:
            current = _parse_files(cell.tutanak_files)
            remaining = [f for f in current if f not in remove_tutanak_names]
            for fname in current:
                if fname in remove_tutanak_names:
                    delete_upload(fname)
            cell.tutanak_files = _dump_files(remaining) if remaining else None
            if cell.tutanak_path and cell.tutanak_path in remove_tutanak_names:
                delete_upload(cell.tutanak_path)
                cell.tutanak_path = None

        # updated_at'i güncelle
        cell.updated_at = datetime.now()

        db.session.commit()
        team_vehicle_payload = None
        if cell.team_id:
            updated_team = Team.query.get(cell.team_id)
            if updated_team:
                assigned_vehicle = None
                if getattr(updated_team, "vehicle_id", None):
                    assigned_vehicle = Vehicle.query.get(updated_team.vehicle_id)
                team_vehicle_payload = _vehicle_payload(assigned_vehicle, updated_team.id)

        # Alt proje bilgisini al
        subproject_label = ""
        if cell.subproject_id:
            subproject = SubProject.query.get(cell.subproject_id)
            if subproject:
                code = (subproject.code or "").strip()
                subproject_label = f"{subproject.name}{f' ({code})' if code else ''}"
        
        curr_u = get_current_user()
        evt_data = {
            "cell_id": cell.id,
            "project_id": cell.project_id,
            "work_date": iso(cell.work_date),
            "shift": cell.shift,
            "note": cell.note,
            "vehicle_info": cell.vehicle_info,
            "team_id": cell.team_id,
            "team_name": cell.team_name,
            "subproject_id": cell.subproject_id,
            "subproject_label": subproject_label,
            "person_ids": [int(pid) for pid in person_ids],
            "hasAttachment": bool(cell.lld_hhd_files or cell.tutanak_files or cell.lld_hhd_path or cell.tutanak_path),
            "team_vehicle": team_vehicle_payload,
            "assigned_user_id": cell.assigned_user_id,
            "updated_at": cell.updated_at.timestamp() if cell.updated_at else datetime.now().timestamp(),
            "updated_by": (curr_u.full_name or curr_u.email) if curr_u else "Sistem",
            "updated_by_id": curr_u.id if curr_u else 0
        }
        socketio.emit('cell_updated', evt_data, namespace='/')

        # Hücreye atanan kişi Job'a hemen yansısın (Benim İşlerim'de görünsün)
        try:
            job = Job.query.filter_by(cell_id=cell.id).first()
            if job:
                _sync_job_from_cell(job, cell)
                db.session.add(job)
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

        # Ensure Job exists for this cell (assignment panel)
        try:
            upsert_jobs_for_range(d, d)
        except Exception:
            pass

        # Notifications: new assignment to newly added people (matching User.email)
        try:
            added = list(added_ids or [])
            if added:
                job = Job.query.filter_by(cell_id=cell.id).first()
                project = Project.query.get(project_id)
                title = 'Yeni atama'
                body = '{} {} | {} | Ekip: {}'.format((project.project_code if project else ''), (project.region if project else ''), iso(d), (cell.team_name or ''))
                link_url = url_for('planner.assignment_page', job_id=job.id) if job else None
                people_rows = Person.query.filter(Person.id.in_(added)).all()
                emails = [p.email for p in people_rows if p and p.email]
                created = _notify_users_by_emails(emails, event='new_assignment', title=title, body=body, link_url=link_url, job_id=(job.id if job else None), meta={'project_id': project_id, 'work_date': iso(d)})
                if not created:
                    _notify_admins(event='new_assignment', title=title, body=body, link_url=link_url, job_id=(job.id if job else None), meta={'project_id': project_id, 'work_date': iso(d)})
                db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
        
        return jsonify({"ok": True, "cell_id": cell.id, "team_id": cell.team_id, "team_name": cell.team_name or "", "team_vehicle": team_vehicle_payload, "subproject_id": cell.subproject_id, "subproject_label": subproject_label})
    except Exception as e:
        db.session.rollback()
        import traceback
        err_msg = traceback.format_exc()
        print(f"ERROR in api_cell_set: {str(e)}")
        with open("c:\\Users\\USER\\Desktop\\Saha\\error_log_post.txt", "w") as f:
            f.write(err_msg)
        return jsonify({"ok": False, "error": f"Kaydetme hatası: {str(e)}"}), 500


@planner_bp.post("/api/save_overtime_only")
@login_required
@observer_required
def api_save_overtime_only():
    try:
        data = request.get_json(force=True, silent=True) or {}
        project_id = int(data.get("project_id", 0))
        work_date_str = data.get("work_date", "")
        if not project_id or not work_date_str:
            return jsonify({"ok": False, "error": "Parametreler eksik"}), 400
            
        d = parse_date(work_date_str) or date.today()
        cell = ensure_cell(project_id, d)
        
        # Update Assignments (this is necessary to link overtime to people)
        # passing preferred_vehicle_info=None ensures we don't accidentally update vehicle or cause conflict
        person_ids = [int(pid) for pid in data.get("person_ids", []) if pid]
        set_assignments_and_team(cell, person_ids, preferred_vehicle_info=None)
        
        # Update Overtime
        person_overtimes = data.get("person_overtimes") or {}
        current_u_id = get_current_user().id if get_current_user() else 1
        
        existing_ots = {ot.person_id: ot for ot in TeamOvertime.query.filter_by(cell_id=cell.id).all()}
        processed_pids = set()
        
        if isinstance(person_overtimes, dict):
            for pid_str, shift_val in person_overtimes.items():
                try:
                    pid = int(pid_str)
                    if pid not in person_ids: continue
                    shift_str = (shift_val or "").strip()
                    if not shift_str: continue
                    
                    processed_pids.add(pid)
                    hours = 0.0
                    try: hours = float(shift_str)
                    except: pass
                    
                    ot = existing_ots.get(pid)
                    if not ot:
                        ot = TeamOvertime(cell_id=cell.id, person_id=pid, work_date=d, created_by_user_id=current_u_id)
                        db.session.add(ot)
                    ot.duration_hours = hours
                    ot.description = shift_str
                    ot.updated_at = datetime.now()
                except: pass
        
        # Cleanup removed overtimes
        for pid in person_ids:
            if pid not in processed_pids and pid in existing_ots:
                db.session.delete(existing_ots[pid])
        # Cleanup overtimes for people no longer in assignment
        for pid, ot in existing_ots.items():
            if pid not in person_ids:
                db.session.delete(ot)
                
        cell.updated_at = datetime.now()
        curr_u = get_current_user()
        cell.updated_by_id = curr_u.id if curr_u else 0
        db.session.commit()

        # Emit socket event
        try:
            evt_data = {
                "cell_id": cell.id,
                "project_id": cell.project_id,
                "work_date": iso(cell.work_date),
                "person_ids": person_ids,
                "updated_at": cell.updated_at.timestamp(),
                "updated_by": curr_u.full_name if curr_u else "Sistem"
            }
            socketio.emit('cell_updated', evt_data, namespace='/')
        except: pass

        return jsonify({"ok": True})

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@planner_bp.post("/api/cell/copy_to_friday")
@login_required
@planner_or_admin_required
def api_cell_copy_to_friday():
    try:
        data = request.get_json(force=True, silent=True) or {}
        project_id = int(data.get("project_id", 0) or 0)
        work_date = parse_date(data.get("work_date", "")) or None
        if project_id <= 0 or not work_date:
            return jsonify({"ok": False, "error": "Parametre eksik."}), 400

        source_cell = PlanCell.query.filter_by(project_id=project_id, work_date=work_date).first()
        if not source_cell:
            return jsonify({"ok": False, "error": "Kaynak hücre bulunamadı."}), 404

        week_begin = week_start(work_date)
        friday = week_begin + timedelta(days=4)
        if work_date > friday:
            return jsonify({"ok": False, "error": "İş bu haftanın cumasından sonra."}), 400

        person_ids = [a.person_id for a in source_cell.assignments]
        next_day = work_date + timedelta(days=1)
        copied_dates = []
        while next_day <= friday:
            dest = ensure_cell(project_id, next_day)
            dest.shift = source_cell.shift if source_cell.shift else None
            dest.vehicle_info = source_cell.vehicle_info if source_cell.vehicle_info else None
            dest.note = source_cell.note if source_cell.note else None
            dest.isdp_info = source_cell.isdp_info if source_cell.isdp_info else None
            dest.po_info = source_cell.po_info if source_cell.po_info else None
            dest.important_note = source_cell.important_note if source_cell.important_note else None
            dest.team_name = source_cell.team_name if source_cell.team_name else None
            dest.job_mail_body = source_cell.job_mail_body if source_cell.job_mail_body else None
            dest.assigned_user_id = source_cell.assigned_user_id if source_cell.assigned_user_id else None
            dest.subproject_id = source_cell.subproject_id if source_cell.subproject_id else None
            dest.updated_at = datetime.now()
            set_assignments_and_team(dest, person_ids, preferred_vehicle_info=(source_cell.vehicle_info or None))
            copied_dates.append(iso(next_day))
            next_day += timedelta(days=1)

        db.session.commit()
        try:
            socketio.emit("update_table", namespace="/")
        except Exception:
            pass
        return jsonify({"ok": True, "copied_dates": copied_dates})
    except Exception as exc:
        db.session.rollback()
        log.exception("api_cell_copy_to_friday error: %s", str(exc))
        return jsonify({"ok": False, "error": "İş kopyalanamadı."}), 500


@planner_bp.post("/api/cell/upload_attachments")
@login_required
@observer_required
def api_cell_upload_attachments():
    try:
        project_id = int(request.form.get("project_id", 0))
        d = parse_date(request.form.get("work_date", "")) or date.today()
        if not project_id:
            return jsonify({"ok": False, "error": "Geçersiz proje"}), 400

        cell = ensure_cell(project_id, d)
        prefix = f"{project_id}-{iso(d)}"
        updated = {}

        lld_files = request.files.getlist("lld_hhd")
        if lld_files:
            cur = _parse_files(cell.lld_hhd_files)
            for fs in lld_files:
                if fs and fs.filename:
                    fname = save_uploaded_file(fs, f"lldhhd-{prefix}")
                    cur.append(fname)
            if cur:
                cell.lld_hhd_files = _dump_files(cur)
                updated["lld_hhd_files"] = cur

        tutanak_files = request.files.getlist("tutanak")
        if tutanak_files:
            cur = _parse_files(cell.tutanak_files)
            for fs in tutanak_files:
                if fs and fs.filename:
                    fname = save_uploaded_file(fs, f"tutanak-{prefix}")
                    cur.append(fname)
            if cur:
                cell.tutanak_files = _dump_files(cur)
                updated["tutanak_files"] = cur

        cell.updated_at = datetime.now()
        db.session.commit()
        socketio.emit('update_table', namespace='/')
        return jsonify({"ok": True, **updated})
    except ValueError as ve:
        return jsonify({"ok": False, "error": str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Yükleme hatası: {str(e)}"}), 500
@planner_bp.post("/api/cell/clear")
@login_required
@observer_required
def api_cell_clear():
    data = request.get_json(force=True, silent=True) or {}
    project_id = int(data.get("project_id", 0))
    d = parse_date(data.get("work_date", "")) or date.today()

    cell = PlanCell.query.filter_by(project_id=project_id, work_date=d).first()
    if not cell:
        return jsonify({"ok": True, "cleared": False})

    apply_snapshot(cell, {"exists": False})
    cell.updated_at = datetime.now()
    db.session.commit()
    socketio.emit('update_table', namespace='/')
    return jsonify({"ok": True, "cleared": True})


# ---------- API: PLAN SYNC (Eş zamanlı güncelleme) ----------
@planner_bp.get("/api/plan_sync")
@login_required
def api_plan_sync():
    """Hafta için son güncelleme zamanını döndürür"""
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", ""))
    start = ws if ws else week_start(d)
    end = start + timedelta(days=6)
    
    # Bu hafta için en son güncellenen cell'in updated_at'ini bul
    last_update = db.session.query(db.func.max(PlanCell.updated_at)).filter(
        PlanCell.work_date >= start,
        PlanCell.work_date <= end
    ).scalar()
    
    # En son güncelleme zamanını hesapla
    max_timestamp = None
    if last_update:
        max_timestamp = last_update.timestamp()
    
    return jsonify({
        "ok": True,
        "last_update": max_timestamp,
        "last_update_str": last_update.isoformat() if last_update else None
    })



# ---------- API: VEHICLES WEEK (Araç sütunu toplu) ----------
@planner_bp.get("/api/vehicles_week")
@login_required
def api_vehicles_week():
    # Belirli hafta için (Pzt..Paz) proje satırında ilk plaka bilgisi döner.
    # Amaç: front-end'in her hücre için /api/cell çağırmasını engelleyip tek sorguda haftalık araçları getirmek.
    # Params: week_start (YYYY-MM-DD, tercih Pazartesi), project_id (opsiyonel).
    # Response: { ok: true, week_start, vehicles: { '<project_id>': '34ABC123', ... } }
    ws = parse_date(request.args.get("week_start", "")) or parse_date(request.args.get("date", "")) or date.today()
    start = week_start(ws)
    end = start + timedelta(days=6)
    project_id = int(request.args.get("project_id", 0) or 0)

    # first non-empty vehicle_info per project in that week
    base = db.session.query(
        PlanCell.project_id.label("project_id"),
        db.func.min(PlanCell.work_date).label("min_date"),
    ).filter(
        PlanCell.work_date >= start,
        PlanCell.work_date <= end,
        PlanCell.vehicle_info.isnot(None),
        db.func.trim(PlanCell.vehicle_info) != "",
    )
    if project_id:
        base = base.filter(PlanCell.project_id == project_id)

    subq = base.group_by(PlanCell.project_id).subquery()

    rows = db.session.query(PlanCell.project_id, PlanCell.vehicle_info).join(
        subq,
        db.and_(
            PlanCell.project_id == subq.c.project_id,
            PlanCell.work_date == subq.c.min_date,
        ),
    ).all()

    def _plate_only(s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        # Split by '(' to clean brand info (e.g. "34ABC123 (Fiat)") but keep spaces in "06 BJ 8300"
        return s.split('(')[0].strip()

    vehicles = {}
    for pid, vinfo in rows:
        plate = _plate_only(vinfo or "")
        if plate:
            vehicles[str(pid)] = plate

    return jsonify({"ok": True, "week_start": iso(start), "vehicles": vehicles})

# ---------- API: CELL TEAM REPORT (Modal tablo) ----------
@planner_bp.get("/api/cell_team_report")
def api_cell_team_report():
    project_id = int(request.args.get("project_id", 0))
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", ""))
    cell = PlanCell.query.filter_by(project_id=project_id, work_date=d).first()
    if not cell:
        return jsonify({"ok": False, "error": "Bu hücrede kayıt yok."}), 404

    project = Project.query.get(project_id)
    persons = (
        db.session.query(Person.full_name, Person.phone, Person.tc_no)
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .filter(CellAssignment.cell_id == cell.id)
        .order_by(Person.full_name.asc())
        .all()
    )

    rows = [{"full_name": n, "phone": ph or "", "tc": tc or ""} for (n, ph, tc) in persons]

    date_range = iso(d)
    team = Team.query.get(cell.team_id) if cell.team_id else None
    if ws:
        we = ws + timedelta(days=6)
        work_days = (
            db.session.query(PlanCell.work_date)
            .join(CellAssignment, CellAssignment.cell_id == PlanCell.id)
            .filter(PlanCell.project_id == project_id)
            .filter(PlanCell.work_date >= ws, PlanCell.work_date <= we)
            .distinct()
            .order_by(PlanCell.work_date.asc())
            .all()
        )
        if work_days:
            dates = [row[0] for row in work_days]
            date_range = format_date_range(dates[0], dates[-1])
    else:
        date_range = format_date_range(d, d)

    return jsonify({
        "ok": True,
        "date": iso(d),
        "city": project.region if project else "",
        "project_code": project.project_code if project else "",
        "project_name": project.project_name if project else "",
        "shift": cell.shift or "",
        "vehicle": cell.vehicle_info or "",
        "team_id": int(team.id) if team else None,
        "team_vehicle": _vehicle_payload(team.vehicle if team else None),
        "note": cell.note or "",
        "people": rows
    })


# ---------- API: JOB MAIL PREVIEW/SEND ----------
@planner_bp.post("/api/preview_job_email")
@login_required
def api_preview_job_email():
    data = request.get_json(force=True, silent=True) or {}
    project_id = int(data.get("project_id", 0) or 0)
    d = parse_date(data.get("work_date", "")) or date.today()
    if not project_id:
        return jsonify({"ok": False, "error": "project_id eksik"}), 400

    short_summary = (data.get("short_summary") or "").strip()
    job_details = (data.get("job_details") or "").strip()
    checklist = data.get("checklist") or []
    links = data.get("links") or []

    ctx = _job_mail_context_for_cell(project_id, d, short_summary=short_summary, job_details=job_details, checklist=checklist, links=links)
    try:
        subject, body_html = _render_inline_mail_templates(DEFAULT_JOB_MAIL_SUBJECT_TEMPLATE, DEFAULT_JOB_MAIL_BODY_TEMPLATE, ctx)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Mail icerigi hazirlanamadi: {str(e)}"}), 400
    subject_override = (data.get("subject_override") or "").strip()
    if subject_override:
        subject = subject_override

    email_html = render_template(
        "email_base.html",
        title=subject,
        heading=subject,
        intro="Bu mail bir onizlemedir. Gonderim oncesi kontrol icindir.",
        body_html=body_html,
        table_headers=None,
        table_rows=None,
        footer=f"{ctx.get('site_code','')} | {ctx.get('date','')}",
    )

    # Default TO: selected people emails (if any)
    to_addr = (data.get("to_addr") or "").strip()
    if not to_addr:
        to_addr = ", ".join([p.get("email") for p in (ctx.get("people") or []) if p.get("email")])

    create_mail_log(kind="preview", ok=True, to_addr=to_addr or "-", subject=subject, meta={
            "type": "job",
            "project_id": project_id,
            "work_date": iso(d),
            "checklist": checklist,
            "links": links,
        })
    return jsonify({"ok": True, "subject": subject, "to": to_addr, "html": email_html})


@planner_bp.post("/api/send_job_email")
@login_required
@observer_required
def api_send_job_email():
    # multipart form
    if not _csrf_verify(request.form.get("csrf_token", "")):
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz"}), 403

    project_id = int(request.form.get("project_id", 0) or 0)
    d = parse_date(request.form.get("work_date", "")) or date.today()
    if not project_id:
        return jsonify({"ok": False, "error": "project_id eksik"}), 400

    to_addr = (request.form.get("to_addr") or "").strip()
    cc_addrs = (request.form.get("cc_addrs") or "").strip()
    bcc_addrs = (request.form.get("bcc_addrs") or "").strip()
    if not to_addr:
        return jsonify({"ok": False, "error": "Alici (To) zorunlu"}), 400

    short_summary = (request.form.get("short_summary") or "").strip()
    job_details = (request.form.get("job_details") or "").strip()

    import json as _json
    try:
        checklist = _json.loads(request.form.get("checklist_json") or "[]")
    except Exception:
        checklist = []
    try:
        links = _json.loads(request.form.get("links_json") or "[]")
    except Exception:
        links = []
    try:
        include_files = _json.loads(request.form.get("include_files_json") or "[]")
    except Exception:
        include_files = []

    ctx = _job_mail_context_for_cell(project_id, d, short_summary=short_summary, job_details=job_details, checklist=checklist, links=links)
    try:
        subject, body_html = _render_inline_mail_templates(DEFAULT_JOB_MAIL_SUBJECT_TEMPLATE, DEFAULT_JOB_MAIL_BODY_TEMPLATE, ctx)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Mail icerigi hazirlanamadi: {str(e)}"}), 400
    subject_override = (request.form.get("subject_override") or "").strip()
    if subject_override:
        subject = subject_override

    email_html = render_template(
        "email_base.html",
        title=subject,
        heading=subject,
        intro="Is atamasi mailidir.",
        body_html=body_html,
        table_headers=None,
        table_rows=None,
        footer=f"{ctx.get('site_code','')} | {ctx.get('date','')}",
    )

    # attachments
    allowed_ext = {"pdf", "doc", "docx", "xls", "xlsx", "png", "jpg", "jpeg"}
    total_size = 0
    attachments_payload = []

    def _add_blob(filename: str, data_blob: bytes, content_type: str):
        nonlocal total_size
        if not filename:
            return
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext and ext not in allowed_ext:
            raise ValueError(f"Desteklenmeyen ek: {filename}")
        total_size += len(data_blob or b"")
        attachments_payload.append({"filename": filename, "data": data_blob, "content_type": content_type})

    # existing uploaded files by name
    for fname in (include_files or []):
        if not fname or "/" in fname or "\\" in fname or ".." in fname:
            continue
        full_path = os.path.join(current_app.config["UPLOAD_FOLDER"], fname)
        if not os.path.exists(full_path):
            continue
        with open(full_path, "rb") as f:
            data_blob = f.read()
        ctype = mimetypes.guess_type(full_path)[0] or "application/octet-stream"
        _add_blob(os.path.basename(full_path), data_blob, ctype)

    # new attachments
    for fs in request.files.getlist("attachments"):
        if not fs or not fs.filename:
            continue
        fname = secure_filename(fs.filename)
        data_blob = fs.read() or b""
        ctype = fs.mimetype or (mimetypes.guess_type(fname)[0] or "application/octet-stream")
        _add_blob(fname, data_blob, ctype)

    max_total = 15 * 1024 * 1024
    if total_size > max_total:
        return jsonify({"ok": False, "error": f"Ek boyutu cok buyuk (max {max_total//1024//1024}MB)."}), 400

    meta = {
        "type": "job",
        "project_id": project_id,
        "work_date": iso(d),
        "cc": cc_addrs,
        "bcc": bcc_addrs,
        "checklist": checklist,
        "links": links,
        "include_files": include_files,
        "attachments_count": len(attachments_payload),
    }
    try:
        ok = MailService.send(
            mail_type="job",
            recipients=[to_addr],
            subject=subject,
            html=email_html,
            attachments=attachments_payload,
            cc=cc_addrs,
            bcc=bcc_addrs,
            user_id=getattr(current_user, "id", None),
            project_id=project_id,
            job_id=None,
            meta=meta,
        )
    except Exception as e:
        log.exception("Job mail gonderimi basarisiz: to=%s err=%s", to_addr, str(e))
        return jsonify({"ok": False, "error": "Mail gonderimi basarisiz."}), 500

    if ok:
        return jsonify({"ok": True, "subject": subject})
    return jsonify({"ok": False, "error": "Mail gonderimi basarisiz."}), 500


# ---------- API: TEAM MEMBERS (Ekip kiYileri) ----------
@planner_bp.get("/api/team_members")
@login_required
def api_team_members():
    """List team members by team name or ID (for signature or weekly records)."""
    team_id = int(request.args.get("team_id", 0) or 0)
    ws = parse_date(request.args.get("week_start", ""))
    start = ws if ws else None
    end = (ws + timedelta(days=6)) if ws else None

    if team_id and not team_name:
        t = Team.query.get(team_id)
        if t:
            team_name = t.name or ""

    people_rows = []
    source = None

    # 1) Team tablosundaki imzadan oku
    if team_id:
        t = Team.query.get(team_id)
        if t and t.signature:
            try:
                ids = [int(x) for x in t.signature.split(",") if x.strip()]
            except Exception:
                ids = []
            if ids:
                people_rows = Person.query.filter(Person.id.in_(ids)).order_by(Person.full_name.asc()).all()
                source = "team_signature"

    # 2) İmza bulunamazsa, o isimle planlanan kayŽñtlardan kiYileri topla
    if (not people_rows) and team_name:
        q = (
            db.session.query(
                Person.id, Person.full_name, Person.phone, Person.tc_no,
                Firma.name.label("firma_name"),
                Seviye.name.label("seviye_name")
            )
            .join(CellAssignment, CellAssignment.person_id == Person.id)
            .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
            .outerjoin(Team, Team.id == PlanCell.team_id)
            .outerjoin(Firma, Firma.id == Person.firma_id)
            .outerjoin(Seviye, Seviye.id == Person.seviye_id)
            .filter(or_(Team.name == team_name, PlanCell.team_name == team_name))
        )
        if start and end:
            q = q.filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        people_rows = q.distinct().order_by(Person.full_name.asc()).all()
        source = "week_cells"

    if not people_rows:
        return jsonify({"ok": False, "error": "Ekip bulunamadŽñ"}), 404

    def _row_to_dict(row):
        if isinstance(row, Person):
            return {
                "id": row.id,
                "full_name": row.full_name,
                "phone": row.phone or "",
                "tc": row.tc_no or "",
                "firma": row.firma.name if row.firma else "",
                "seviye": row.seviye.name if row.seviye else ""
            }
        pid, fullname, phone, tc, fname, sname = row
        return {
            "id": pid,
            "full_name": fullname,
            "phone": phone or "",
            "tc": tc or "",
            "firma": fname or "",
            "seviye": sname or ""
        }

    resolved_name = team_name
    if not resolved_name and team_id:
        t = Team.query.get(team_id)
        resolved_name = t.name if t else ""

    return jsonify({
        "ok": True,
        "team_name": resolved_name,
        "source": source or "",
        "people": [_row_to_dict(r) for r in people_rows]
    })


# ---------- API: Team Tooltip (Ekip Tooltip için basit endpoint) ----------
@planner_bp.get("/api/team/<int:team_id>/tooltip")
@login_required
def api_team_tooltip(team_id: int):
    """
    Ekip tooltip için basit endpoint.
    Ekip üyelerinin isimlerini ve haftalık dinamik ekip adını döndürür.
    
    Query params:
        - week_start: Hafta başlangıç tarihi (YYYY-MM-DD) - dinamik ekip adı için
    """
    from utils import get_team_members_names, get_weekly_team_display_name
    
    if not team_id or team_id <= 0:
        return jsonify({"ok": False, "error": "Geçersiz team_id"}), 400
    
    # Ekip üyelerinin isimlerini al
    member_names = get_team_members_names(team_id)
    
    if not member_names:
        return jsonify({"ok": False, "error": "Ekip bulunamadı veya üyesi yok"}), 404
    
    # Haftalık dinamik ekip adını hesapla
    ws = parse_date(request.args.get("week_start", ""))
    if ws:
        display_name = get_weekly_team_display_name(team_id, ws)
    else:
        # Hafta belirtilmemişse veritabanındaki adı kullan
        team = Team.query.get(team_id)
        display_name = team.name if team else f"Ekip #{team_id}"
    
    return jsonify({
        "ok": True,
        "team_id": team_id,
        "display_name": display_name,
        "members": member_names,
        "member_count": len(member_names)
    })


# ---------- API: Weekly Team Names (Haftalık Ekip Numaraları) ----------
@planner_bp.get("/api/weekly_team_names")
@login_required
def api_weekly_team_names():
    """
    Verilen hafta için tüm ekiplerin dinamik numaralarını döndürür.
    Frontend'de toplu gösterim için kullanılır.
    
    Query params:
        - week_start: Hafta başlangıç tarihi (YYYY-MM-DD)
    
    Döndürür: {team_id: "Ekip 1", team_id2: "Ekip 2", ...}
    """
    from utils import get_weekly_team_map, get_team_members_names
    
    ws = parse_date(request.args.get("week_start", ""))
    if not ws:
        return jsonify({"ok": False, "error": "week_start gerekli"}), 400
    
    team_map = get_weekly_team_map(ws)
    
    # Team ID -> display name ve üyeler
    result = {}
    for team_id, display_number in team_map.items():
        display_name = f"Ekip {display_number}"
        members = get_team_members_names(team_id)
        result[str(team_id)] = {
            "display_name": display_name,
            "members": members,
            "member_count": len(members)
        }
    
    return jsonify({
        "ok": True,
        "week_start": iso(ws) if ws else None,
        "teams": result
    })


@planner_bp.get("/api/person_assigned")
@login_required
def api_person_assigned():
    """Belirli bir günde personelin başka bir projede çalışıp çalışmadığını kontrol et"""
    d = parse_date(request.args.get("date", "")) or date.today()
    current_project_id = int(request.args.get("current_project_id", 0) or 0)
    
    # O gün başka bir projede çalışan personelleri bul
    assigned_person_ids = (
        db.session.query(CellAssignment.person_id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .filter(PlanCell.work_date == d)
        .filter(PlanCell.project_id != current_project_id)  # Mevcut proje hariç
        .distinct()
        .all()
    )
    
    assigned_ids = [row[0] for row in assigned_person_ids]
    
    # Personel bilgilerini de al
    assigned_people = []
    if assigned_ids:
        people = Person.query.filter(Person.id.in_(assigned_ids)).all()
        for p in people:
            # Hangi projede çalıştığını bul
            cells = (
                db.session.query(PlanCell.project_id, Project.project_code)
                .join(Project, Project.id == PlanCell.project_id)
                .join(CellAssignment, CellAssignment.cell_id == PlanCell.id)
                .filter(PlanCell.work_date == d)
                .filter(PlanCell.project_id != current_project_id)
                .filter(CellAssignment.person_id == p.id)
                .first()
            )
            project_code = cells[1] if cells else "Bilinmeyen"
            assigned_people.append({
                "person_id": p.id,
                "full_name": p.full_name,
                "project_code": project_code
            })
    
    # Mevcut hücredeki mesai bilgilerini al
    cell_id = int(request.args.get("cell_id", 0) or 0)
    person_overtimes = {}
    if cell_id:
        from models import TeamOvertime
        overtimes = TeamOvertime.query.filter_by(cell_id=cell_id).all()
        for ot in overtimes:
            if ot.person_id:
                if ot.person_id not in person_overtimes:
                    person_overtimes[ot.person_id] = []
                person_overtimes[ot.person_id].append({
                    "hours": ot.duration_hours,
                    "description": ot.description or ""
                })
    
    return jsonify({
        "ok": True,
        "assigned_person_ids": assigned_ids,
        "assigned_people": assigned_people,
        "person_overtimes": person_overtimes
    })


# ---------- API: Team Conflict Check (Ekip çakışması kontrolü) ----------
@planner_bp.get("/api/team_conflict")
@login_required
def api_team_conflict():
    """Seçilen personellerden oluşan ekibin aynı gün başka bir projede çalışıp çalışmadığını kontrol et"""
    d = parse_date(request.args.get("date", "")) or date.today()
    person_ids_str = request.args.get("person_ids", "")
    current_project_id = int(request.args.get("current_project_id", 0) or 0)
    
    if not person_ids_str:
        return jsonify({"ok": True, "has_conflict": False})
    
    try:
        person_ids = [int(pid.strip()) for pid in person_ids_str.split(",") if pid.strip()]
    except:
        return jsonify({"ok": False, "error": "Geçersiz person_ids"}), 400
    
    if not person_ids:
        return jsonify({"ok": True, "has_conflict": False})
    
    # Tüm personellerin aynı gün atandığı hücreleri bul
    cells_with_all_people = (
        db.session.query(PlanCell.id, PlanCell.project_id)
        .join(CellAssignment, CellAssignment.cell_id == PlanCell.id)
        .filter(PlanCell.work_date == d)
        .filter(PlanCell.project_id != current_project_id)  # Mevcut proje hariç
        .filter(CellAssignment.person_id.in_(person_ids))
        .group_by(PlanCell.id)
        .having(db.func.count(db.func.distinct(CellAssignment.person_id)) == len(person_ids))
        .all()
    )
    
    # Bu hücrelerde sadece bu personeller mi var kontrol et
    conflicting_projects = []
    for cell_id, project_id in cells_with_all_people:
        cell_person_ids = [
            a.person_id for a in CellAssignment.query.filter_by(cell_id=cell_id).all()
        ]
        if sorted(cell_person_ids) == sorted(person_ids):
            conflicting_cell = PlanCell.query.get(cell_id)
            if conflicting_cell:
                conflicting_project = Project.query.get(conflicting_cell.project_id)
                if conflicting_project:
                    conflicting_projects.append({
                        "project_id": conflicting_cell.project_id,
                        "project_code": conflicting_project.project_code,
                        "project_name": conflicting_project.project_name
                    })
    
    has_conflict = len(conflicting_projects) > 0
    return jsonify({
        "ok": True,
        "has_conflict": has_conflict,
        "conflicting_projects": conflicting_projects
    })


# ---------- API: Availability (plan ekranında) ----------
@planner_bp.get("/api/availability")
def api_availability():
    d = parse_date(request.args.get("date", "")) or date.today()
    shift = (request.args.get("shift") or "").strip()

    # izinli/üretimde olanları çek
    st_rows = PersonDayStatus.query.filter_by(work_date=d).all()
    status_by_person = {r.person_id: r.status for r in st_rows}

    # çalışanlar
    q = db.session.query(CellAssignment.person_id).join(PlanCell, PlanCell.id == CellAssignment.cell_id)\
        .filter(PlanCell.work_date == d)
    if shift:
        # Hem yeni hem eski formatı kontrol et
        normalized_shift = normalize_shift(shift)
        q = q.filter(db.or_(PlanCell.shift == shift, PlanCell.shift == normalized_shift))
    busy_ids = {r[0] for r in q.distinct().all()}

    people = Person.query.order_by(Person.full_name.asc()).all()

    # “boşta” = çalışmıyor + status available (default)
    available = []
    busy = []
    leave = []
    production = []

    for p in people:
        st = status_by_person.get(p.id, "available")
        if st == "leave":
            leave.append({"id": p.id, "name": p.full_name})
            continue
        if st == "production":
            production.append({"id": p.id, "name": p.full_name})
            continue

        if p.id in busy_ids:
            busy.append({"id": p.id, "name": p.full_name})
        else:
            available.append({"id": p.id, "name": p.full_name})

    return jsonify({
        "date": iso(d),
        "shift": shift,
        "available": available,
        "busy": busy,
        "leave": leave,
        "production": production
    })


# ---------- API: Person day status set ----------
@planner_bp.get("/api/person_status_day")
def api_person_status_day():
    d = parse_date(request.args.get("date", "")) or date.today()
    rows = PersonDayStatus.query.filter_by(work_date=d).all()
    status_by_person = {str(r.person_id): {"status": r.status, "note": r.note or ""} for r in rows}
    return jsonify({"ok": True, "date": iso(d), "status_by_person": status_by_person})

@planner_bp.get("/api/person_status_week")
def api_person_status_week():
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)
    rows = PersonDayStatus.query.filter(PersonDayStatus.work_date >= start,
                                        PersonDayStatus.work_date <= end).all()
    out = {}
    for r in rows:
        pid = str(r.person_id)
        out.setdefault(pid, {})[iso(r.work_date)] = {"status": r.status, "note": r.note or ""}
    return jsonify({"ok": True, "week_start": iso(ws), "status_by_person": out})

@planner_bp.get("/api/person_week")
def api_person_week():
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    person_id = int(request.args.get("person_id", 0) or 0)
    if not person_id:
        return jsonify({"ok": False, "error": "person_id eksik"}), 400

    start = ws
    end = ws + timedelta(days=6)

    rows = (
        db.session.query(PlanCell.work_date, PlanCell.shift, PlanCell.vehicle_info, PlanCell.note,
                         Project.project_code, Project.region, Project.project_name)
        .join(Project, Project.id == PlanCell.project_id)
        .join(CellAssignment, CellAssignment.cell_id == PlanCell.id)
        .filter(CellAssignment.person_id == person_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .order_by(PlanCell.work_date.asc(),
                  db.case((PlanCell.shift == "Gündüz", 1),
                          (PlanCell.shift == "Gündüz Yol", 2),
                          (PlanCell.shift == "Gece", 3),
                          else_=9).asc(),
                  Project.project_code.asc())
        .all()
    )

    items = []
    for wd, shift, vehicle, note, pcode, city, pname in rows:
        items.append({
            "date": iso(wd),
            "shift": shift or "",
            "vehicle": vehicle or "",
            "note": note or "",
            "project_code": pcode or "",
            "city": (city or "").strip(),
            "project_name": pname or "",
        })

    p = Person.query.get(person_id)
    return jsonify({
        "ok": True,
        "week_start": iso(ws),
        "person": {
            "id": person_id,
            "full_name": p.full_name if p else f"#{person_id}",
            "team": getattr(p, "team", "") or ""
        },
        "items": items
    })

@planner_bp.post("/api/person_status")
@login_required
@observer_required
def api_person_status():
    data = request.get_json(force=True, silent=True) or {}
    person_id = int(data.get("person_id", 0))
    d = parse_date(data.get("work_date", "")) or date.today()
    status = (data.get("status") or "available").strip()
    note = (data.get("note") or "").strip()
    isdp_info = (data.get("isdp_info") or "").strip()
    po_info = (data.get("po_info") or "").strip()

    if status not in ("available", "leave", "office", "production"):
        return jsonify({"ok": False, "error": "status geçersiz"}), 400

    row = PersonDayStatus.query.filter_by(person_id=person_id, work_date=d).first()
    if not row:
        row = PersonDayStatus(person_id=person_id, work_date=d, status=status, note=note or None)
        db.session.add(row)
    else:
        row.status = status
        row.note = note or None

    # available seçildiyse satırı silelim (db şişmesin)
    if status == "available":
        db.session.delete(row)

    db.session.commit()
    return jsonify({"ok": True})


# ---------- PERSONNEL STATUS TYPES & IDLE PERSONNEL API ----------

@planner_bp.get("/api/personnel/idle")
@login_required
def api_personnel_idle():
    """
    Belirli bir gün için boşta olan personel listesini döndür.
    Query params: work_date (YYYY-MM-DD)
    """
    work_date_str = request.args.get("work_date", "")
    d = parse_date(work_date_str) or date.today()
    
    # Tüm aktif personelleri al
    people = Person.query.filter(Person.durum == "Aktif").order_by(Person.full_name.asc()).all()
    
    # O gün çalışan personelleri bul (normal projeler)
    busy_q = db.session.query(CellAssignment.person_id).join(PlanCell, PlanCell.id == CellAssignment.cell_id)\
        .filter(PlanCell.work_date == d, PlanCell.status != 'cancelled')
    busy_ids = {r[0] for r in busy_q.distinct().all()}
    
    # 9026-0001 projesine atananlar (izin/raporlu/ofis - bunlar da boşta değil)
    office_q = db.session.query(CellAssignment.person_id)\
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)\
        .join(Project, Project.id == PlanCell.project_id)\
        .filter(PlanCell.work_date == d, Project.project_code.like("9026-0001%"))
    office_leave_ids = {r[0] for r in office_q.distinct().all()}
    
    # O gün izinli/raporlu olan personelleri bul (PersonDayStatus tablosu)
    status_q = PersonDayStatus.query.filter(
        PersonDayStatus.work_date == d,
        PersonDayStatus.status != "available"
    ).all()
    status_map = {s.person_id: s for s in status_q}
    
    idle_list = []
    for p in people:
        # Normal projede çalışıyor
        if p.id in busy_ids and p.id not in office_leave_ids:
            continue
        
        # 9026-0001 projesine atanmış (izin/raporlu/ofis)
        if p.id in office_leave_ids:
            continue
        
        # PersonDayStatus'de izinli/raporlu
        status_rec = status_map.get(p.id)
        if status_rec and status_rec.status != "available":
            continue
        
        # Boşta
        idle_list.append({
            "id": p.id,
            "full_name": p.full_name,
            "role": p.role or "",
            "firma_name": p.firma.name if p.firma else "",
            "seviye_name": p.seviye.name if p.seviye else "",
            "email": p.email or "",
            "phone": p.phone or ""
        })
    
    return jsonify({
        "ok": True,
        "work_date": iso(d),
        "idle_count": len(idle_list),
        "idle_personnel": idle_list
    })


@planner_bp.get("/api/personnel/status-types")
@login_required
def api_personnel_status_types():
    """Tüm durum tiplerini döndür"""
    types = PersonnelStatusType.query.filter_by(is_active=True).order_by(PersonnelStatusType.display_order.asc()).all()
    
    result = []
    for t in types:
        subproject_info = None
        if t.subproject_id and t.subproject:
            subproject_info = {
                "id": t.subproject.id,
                "name": t.subproject.name,
                "code": t.subproject.code or ""
            }
        result.append({
            "id": t.id,
            "name": t.name,
            "code": t.code,
            "color": t.color or "#be123c",
            "icon": t.icon or "📋",
            "display_order": t.display_order,
            "subproject": subproject_info
        })
    
    return jsonify({"ok": True, "status_types": result})


@planner_bp.get("/api/personnel/status-types/all")
@login_required
@planner_or_admin_required
def api_personnel_status_types_all():
    """Tüm durum tiplerini (aktif/pasif) döndür - Ayarlar sayfası için"""
    types = PersonnelStatusType.query.order_by(PersonnelStatusType.display_order.asc()).all()
    
    result = []
    for t in types:
        subproject_info = None
        if t.subproject_id and t.subproject:
            subproject_info = {
                "id": t.subproject.id,
                "name": t.subproject.name,
                "code": t.subproject.code or ""
            }
        result.append({
            "id": t.id,
            "name": t.name,
            "code": t.code,
            "color": t.color or "#be123c",
            "icon": t.icon or "📋",
            "display_order": t.display_order,
            "is_active": t.is_active,
            "subproject_id": t.subproject_id,
            "subproject": subproject_info
        })
    
    return jsonify({"ok": True, "status_types": result})


@planner_bp.post("/api/personnel/status-types")
@login_required
@planner_or_admin_required
def api_personnel_status_type_create():
    """Yeni durum tipi oluştur"""
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF doğrulaması başarısız"}), 400
    
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().lower().replace(" ", "_")
    color = (data.get("color") or "#be123c").strip()
    icon = (data.get("icon") or "📋").strip()
    display_order = int(data.get("display_order", 0) or 0)
    subproject_id = int(data.get("subproject_id", 0) or 0) or None
    
    if not name or not code:
        return jsonify({"ok": False, "error": "name ve code zorunlu"}), 400
    
    # Kod benzersiz mi?
    existing = PersonnelStatusType.query.filter_by(code=code).first()
    if existing:
        return jsonify({"ok": False, "error": f"'{code}' kodu zaten kullanımda"}), 400
    
    new_type = PersonnelStatusType(
        name=name,
        code=code,
        color=color,
        icon=icon,
        display_order=display_order,
        subproject_id=subproject_id,
        is_active=True
    )
    db.session.add(new_type)
    db.session.commit()
    
    return jsonify({"ok": True, "id": new_type.id, "message": f"'{name}' durum tipi oluşturuldu"})


@planner_bp.put("/api/personnel/status-types/<int:type_id>")
@login_required
@planner_or_admin_required
def api_personnel_status_type_update(type_id):
    """Durum tipini güncelle"""
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF doğrulaması başarısız"}), 400
    
    st = PersonnelStatusType.query.get(type_id)
    if not st:
        return jsonify({"ok": False, "error": "Durum tipi bulunamadı"}), 404
    
    name = data.get("name")
    color = data.get("color")
    icon = data.get("icon")
    display_order = data.get("display_order")
    subproject_id = data.get("subproject_id")
    is_active = data.get("is_active")
    visible_in_summary = data.get("visible_in_summary")
    
    if name is not None:
        st.name = str(name).strip()
    if color is not None:
        st.color = str(color).strip()
    if icon is not None:
        st.icon = str(icon).strip()
    if display_order is not None:
        st.display_order = int(display_order or 0)
    if subproject_id is not None:
        st.subproject_id = int(subproject_id or 0) or None
    if is_active is not None:
        st.is_active = bool(is_active)
    if visible_in_summary is not None:
        st.visible_in_summary = bool(visible_in_summary)
    
    db.session.commit()
    return jsonify({"ok": True, "message": f"'{st.name}' güncellendi"})


@planner_bp.delete("/api/personnel/status-types/<int:type_id>")
@login_required
@planner_or_admin_required
def api_personnel_status_type_delete(type_id):
    """Durum tipini sil"""
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF doğrulaması başarısız"}), 400
    
    st = PersonnelStatusType.query.get(type_id)
    if not st:
        return jsonify({"ok": False, "error": "Durum tipi bulunamadı"}), 404
    
    # Kullanan kayıtlar var mı?
    used_count = PersonDayStatus.query.filter_by(status_type_id=type_id).count()
    if used_count > 0:
        return jsonify({"ok": False, "error": f"Bu durum tipi {used_count} kayıtta kullanılıyor, silinemez"}), 400
    
    name = st.name
    db.session.delete(st)
    db.session.commit()
    return jsonify({"ok": True, "message": f"'{name}' silindi"})


@planner_bp.post("/api/personnel/bulk-status")
@login_required
@planner_or_admin_required
def api_personnel_bulk_status():
    """
    Birden fazla personele aynı gün için durum ata.
    Bu endpoint aynı zamanda ilgili alt projeye de atama yapar.
    """
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF doğrulaması başarısız"}), 400
    
    person_ids = data.get("person_ids", [])
    work_date_str = data.get("work_date", "")
    status_type_id = int(data.get("status_type_id", 0) or 0)
    note = (data.get("note") or "").strip()
    
    if not person_ids:
        return jsonify({"ok": False, "error": "En az bir personel seçin"}), 400
    
    d = parse_date(work_date_str)
    if not d:
        return jsonify({"ok": False, "error": "Geçersiz tarih"}), 400
    
    if not status_type_id:
        return jsonify({"ok": False, "error": "Durum tipi seçin"}), 400
    
    status_type = PersonnelStatusType.query.get(status_type_id)
    if not status_type:
        return jsonify({"ok": False, "error": "Durum tipi bulunamadı"}), 404
    
    # Durum tipine göre status string
    status_str = "leave"  # Varsayılan
    if status_type.code in ("leave", "sick", "excuse", "unpaid_leave"):
        status_str = "leave"
    elif status_type.code in ("office", "home_office"):
        status_str = "office"
    else:
        status_str = "leave"
    
    updated_count = 0
    conflict_persons = []
    
    for pid in person_ids:
        pid = int(pid)
        person = Person.query.get(pid)
        if not person:
            continue
        
        # Aynı gün için mevcut bir çalışma kaydı var mı?
        existing_work = db.session.query(CellAssignment).join(PlanCell)\
            .filter(CellAssignment.person_id == pid, PlanCell.work_date == d, PlanCell.status != 'cancelled').first()
        
        if existing_work:
            conflict_persons.append(person.full_name)
            continue
        
        # PersonDayStatus kaydet
        row = PersonDayStatus.query.filter_by(person_id=pid, work_date=d).first()
        if not row:
            row = PersonDayStatus(person_id=pid, work_date=d)
            db.session.add(row)
        
        row.status = status_str
        row.status_type_id = status_type_id
        row.note = note or None
        updated_count += 1
    
    db.session.commit()
    
    result = {
        "ok": True,
        "updated_count": updated_count,
        "message": f"{updated_count} personel için '{status_type.name}' durumu atandı"
    }
    
    if conflict_persons:
        result["conflicts"] = conflict_persons
        result["conflict_message"] = f"{len(conflict_persons)} personelin bu tarihte mevcut işi var: {', '.join(conflict_persons[:5])}"
    
    return jsonify(result)


@planner_bp.get("/api/personnel/day-summary")
@login_required
def api_personnel_day_summary():
    """
    Belirli bir gün için personel özeti döndür.
    Toplam boşta, izinli, raporlu, ofis sayıları ve detayları.
    """
    work_date_str = request.args.get("work_date", "")
    d = parse_date(work_date_str) or date.today()
    
    # Tüm aktif personeller
    people = Person.query.filter(Person.durum == "Aktif").all()
    
    # O gün çalışanlar
    busy_q = db.session.query(CellAssignment.person_id).join(PlanCell, PlanCell.id == CellAssignment.cell_id)\
        .filter(PlanCell.work_date == d, PlanCell.status != 'cancelled')
    busy_ids = {r[0] for r in busy_q.distinct().all()}
    
    # PersonDayStatus kayıtları
    status_recs = PersonDayStatus.query.filter(PersonDayStatus.work_date == d).all()
    status_by_person = {s.person_id: s for s in status_recs}
    
    # Durum tipleri
    status_types = PersonnelStatusType.query.filter_by(is_active=True).order_by(PersonnelStatusType.display_order.asc()).all()
    
    # Sayaçlar
    idle_list = []
    by_status_type = {st.id: [] for st in status_types}
    by_seviye = {"Ana kadro": [], "Yardımcı": [], "Alt yüklenici": []}
    
    for p in people:
        seviye_name = p.seviye.name if p.seviye else None
        firma_name = p.firma.name if p.firma else ""
        
        if p.id in busy_ids:
            continue  # Çalışıyor, sayma
        
        status_rec = status_by_person.get(p.id)
        
        if status_rec and status_rec.status != "available":
            # Durum tipi varsa
            if status_rec.status_type_id and status_rec.status_type_id in by_status_type:
                by_status_type[status_rec.status_type_id].append({
                    "id": p.id,
                    "full_name": p.full_name,
                    "firma": firma_name,
                    "seviye": seviye_name
                })
        else:
            # Boşta
            idle_list.append({
                "id": p.id,
                "full_name": p.full_name,
                "firma": firma_name,
                "seviye": seviye_name
            })
            if seviye_name in by_seviye:
                by_seviye[seviye_name].append(p.full_name)
    
    # Status type sayıları
    status_counts = {}
    for st in status_types:
        status_counts[st.code] = {
            "id": st.id,
            "name": st.name,
            "icon": st.icon,
            "color": st.color,
            "count": len(by_status_type.get(st.id, [])),
            "persons": by_status_type.get(st.id, [])
        }
    
    return jsonify({
        "ok": True,
        "work_date": iso(d),
        "idle": {
            "count": len(idle_list),
            "persons": idle_list,
            "by_seviye": {
                "ana_kadro": len(by_seviye.get("Ana kadro", [])),
                "yardimci": len(by_seviye.get("Yardımcı", [])),
                "alt_yuklenici": len(by_seviye.get("Alt yüklenici", []))
            }
        },
        "by_status": status_counts
    })


@planner_bp.get("/api/subprojects/office")
@login_required
def api_office_subprojects():
    """
    9026-0001 GENEL OFİS GİDERLERİ projesinin alt projelerini döndür.
    Durum tipleriyle eşleştirmek için kullanılır.
    """
    # 9026-0001 proje kodlu projeyi bul
    office_project = Project.query.filter(
        Project.project_code.like("9026-0001%")
    ).first()
    
    if not office_project:
        # Proje yoksa, oluştur
        office_project = Project(
            region="-",
            project_code="9026-0001",
            project_name="GENEL OFİS GİDERLERİ",
            responsible="Sistem",
            is_active=True
        )
        db.session.add(office_project)
        db.session.commit()
    
    # Alt projeleri al
    subprojects = SubProject.query.filter_by(project_id=office_project.id, is_active=True)\
        .order_by(SubProject.code.asc()).all()
    
    # Varsayılan alt projeler yoksa oluştur
    if not subprojects:
        default_subs = [
            {"code": "9026-0001-01", "name": "SENELİK İZİN"},
            {"code": "9026-0001-02", "name": "MAZERET İZNİ"},
            {"code": "9026-0001-03", "name": "ÜCRETSİZ İZİN"},
            {"code": "9026-0001-07", "name": "HOME OFİS"},
            {"code": "9026-0001-08", "name": "OFİS"},
            {"code": "9026-0001-09", "name": "RAPORLU"},
        ]
        for sub_data in default_subs:
            sub = SubProject(
                project_id=office_project.id,
                code=sub_data["code"],
                name=sub_data["name"],
                is_active=True
            )
            db.session.add(sub)
        db.session.commit()
        subprojects = SubProject.query.filter_by(project_id=office_project.id, is_active=True)\
            .order_by(SubProject.code.asc()).all()
    
    result = []
    for sp in subprojects:
        result.append({
            "id": sp.id,
            "code": sp.code or "",
            "name": sp.name,
            "full_label": f"{sp.code} {sp.name}" if sp.code else sp.name
        })
    
    return jsonify({
        "ok": True,
        "project_id": office_project.id,
        "project_code": office_project.project_code,
        "project_name": office_project.project_name,
        "subprojects": result
    })


# ---------- COPY ----------
@planner_bp.post("/api/copy_monday_to_week")
@login_required
@observer_required
def api_copy_monday_to_week():
    data = request.get_json(force=True, silent=True) or {}
    project_id = int(data.get("project_id", 0))
    ws = parse_date(data.get("week_start", ""))

    if not project_id or not ws:
        return jsonify({"ok": False, "error": "project_id / week_start eksik"}), 400

    src = PlanCell.query.filter_by(project_id=project_id, work_date=ws).first()
    snap = snapshot_cell(src)

    for i in range(1, 7):
        dst = ensure_cell(project_id, ws + timedelta(days=i))
        apply_snapshot(dst, snap)

    db.session.commit()
    socketio.emit('update_table', namespace='/')
    return jsonify({"ok": True})


@planner_bp.post("/api/copy_week_to_next")
@login_required
@observer_required
def api_copy_week_to_next():
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", ""))
    if not ws:
        return jsonify({"ok": False, "error": "week_start eksik"}), 400

    src_start = ws
    src_end = ws + timedelta(days=6)
    dst_start = ws + timedelta(days=7)

    src_cells = PlanCell.query.filter(PlanCell.work_date >= src_start, PlanCell.work_date <= src_end).all()
    src_map: Dict[Tuple[int, date], PlanCell] = {(c.project_id, c.work_date): c for c in src_cells}

    for p in Project.query.all():
        for i in range(7):
            dst = ensure_cell(p.id, dst_start + timedelta(days=i))
            apply_snapshot(dst, snapshot_cell(src_map.get((p.id, src_start + timedelta(days=i)))))

    db.session.commit()
    socketio.emit('update_table', namespace='/')
    return jsonify({"ok": True})


@planner_bp.post("/api/copy_week_from_previous")
@login_required
@observer_required
def api_copy_week_from_previous():
    """Önceki haftayı mevcut haftaya kopyala - Tüm projeler ve hücreler"""
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", ""))
    if not ws:
        return jsonify({"ok": False, "error": "week_start eksik"}), 400

    # Önceki hafta (kaynak)
    src_start = ws - timedelta(days=7)
    src_end = src_start + timedelta(days=6)
    # Mevcut hafta (hedef)
    dst_start = ws
    dst_end = ws + timedelta(days=6)

    # Önceki haftanın mevcut hafta olup olmadığını kontrol et
    today = date.today()
    today_week_start = week_start(today)
    src_is_current_week = (src_start == today_week_start)

    # Önceki haftada hangi projeler var (hem dolu hem boş hücreler için)
    src_cells = PlanCell.query.filter(PlanCell.work_date >= src_start, PlanCell.work_date <= src_end).all()
    src_project_ids_with_work = {c.project_id for c in src_cells}
    
    # Önceki haftada görünen tüm projeleri al
    # plan_week() mantığına göre: mevcut hafta ise tüm aktif projeler, değilse sadece iş eklenmiş projeler
    if src_is_current_week:
        # Önceki hafta mevcut hafta ise: TÜM aktif projeleri kopyala
        src_projects = Project.query.filter(
            Project.is_active == True,
            Project.region != '-'
        ).order_by(Project.region.asc(), Project.project_code.asc()).all()
    else:
        # Önceki hafta geçmiş/gelecek hafta ise: Sadece o haftada iş eklenmiş projeleri kopyala
        if not src_project_ids_with_work:
            return jsonify({"ok": False, "error": "Önceki haftada kopyalanacak veri bulunamadı."}), 400
        
        src_projects = Project.query.filter(
            Project.is_active == True,
            Project.region != '-',
            Project.id.in_(src_project_ids_with_work)
        ).order_by(Project.region.asc(), Project.project_code.asc()).all()
    
    if not src_projects:
        return jsonify({"ok": False, "error": "Önceki haftada kopyalanacak proje bulunamadı."}), 400
    
    src_map: Dict[Tuple[int, date], PlanCell] = {(c.project_id, c.work_date): c for c in src_cells}
    copied_count = 0
    projects_copied = 0

    # Önceki haftada görünen her proje için
    for project in src_projects:
        project_id = project.id
        projects_copied += 1
        
        # 7 günün tamamını kopyala (dolu olsun boş olsun)
        for i in range(7):
            src_date = src_start + timedelta(days=i)
            dst_date = dst_start + timedelta(days=i)
            src_cell = src_map.get((project_id, src_date))
            
            # Hücre oluştur (boş bile olsa)
            dst = ensure_cell(project_id, dst_date)
            
            if src_cell:
                # Dolu hücre varsa içeriğini kopyala
                apply_snapshot(dst, snapshot_cell(src_cell))
                copied_count += 1
            # Boş hücre de oluşturuldu, böylece proje görünecek

    db.session.commit()
    socketio.emit('update_table', namespace='/')
    return jsonify({
        "ok": True, 
        "copied_count": copied_count,
        "projects_copied": projects_copied,
        "message": f"{projects_copied} proje ve {copied_count} hücre kopyalandı."
    })


@planner_bp.post("/api/project_create_from_plan")
@login_required
@observer_required
def api_project_create_from_plan(): 
    data = request.get_json(force=True, silent=True) or {} 
    region = (data.get("city") or "").strip()  # city stored in legacy region column 
    project_id = int(data.get("template_project_id") or 0) 
    subproject_id = int(data.get("subproject_id") or 0) 
    week_start_str = data.get("week_start", "") 

    if not region:
        return jsonify({"ok": False, "error": "Şehir seçin."}), 400

    if not project_id:
        return jsonify({"ok": False, "error": "Proje seçin."}), 400

    t = Project.query.get(project_id) 
    if not t or t.region != "-": 
        return jsonify({"ok": False, "error": "Proje şablonu bulunamadı."}), 404 
 
    if subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id)): 
        return jsonify({"ok": False, "error": "Alt proje secimi gecersiz."}), 400 

    p = Project(region=region, project_code=t.project_code, project_name=t.project_name,
                responsible=t.responsible, karsi_firma_sorumlusu=t.karsi_firma_sorumlusu, is_active=True)
    db.session.add(p)
    db.session.flush()  # ID'yi almak için
    
    ws = parse_date(week_start_str) if week_start_str else None 
    if ws: 
        # Haftanın ilk günü için boş bir hücre oluştur (projenin o haftada görünmesi için) 
        cell = ensure_cell(p.id, ws) 
        cell.subproject_id = (subproject_id if subproject_id else None) 
 
    db.session.commit() 
    socketio.emit('update_table') 
    
    return jsonify({"ok": True, "project_id": p.id, "existed": False})


@planner_bp.post("/api/plan_row_update")
@login_required
@observer_required
def api_plan_row_update(): 
    """Plan satırını güncelle: şehir ve proje şablonu değiştir.""" 
    data = request.get_json(force=True, silent=True) or {} 
    row_project_id = int(data.get("project_id") or 0) 
    region = (data.get("city") or "").strip() 
    template_project_id = int(data.get("template_project_id") or 0) 
    subproject_id = int(data.get("subproject_id") or 0) 
    week_start_str = str(data.get("week_start") or "").strip() 

    if not row_project_id or not region or not template_project_id:
        return jsonify({"ok": False, "error": "project_id/city/template_project_id eksik"}), 400

    row_p = Project.query.get(row_project_id)
    if not row_p or row_p.region == "-":
        return jsonify({"ok": False, "error": "Satır bulunamadı."}), 404

    t = Project.query.get(template_project_id) 
    if not t or t.region != "-": 
        return jsonify({"ok": False, "error": "Proje şablonu bulunamadı."}), 404 
 
    if subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=template_project_id)): 
        return jsonify({"ok": False, "error": "Alt proje secimi gecersiz."}), 400 

    row_p.region = region
    row_p.project_code = t.project_code
    row_p.project_name = t.project_name 
    row_p.responsible = t.responsible 
    row_p.karsi_firma_sorumlusu = t.karsi_firma_sorumlusu 
    row_p.is_active = True 
 
    ws = parse_date(week_start_str) if week_start_str else None 
    if ws: 
        cell = ensure_cell(row_p.id, ws) 
        cell.subproject_id = (subproject_id if subproject_id else None) 
    db.session.commit() 
    socketio.emit('update_table', namespace='/') 
    return jsonify({"ok": True}) 


@planner_bp.post("/api/plan_row_delete")
@login_required
@observer_required
def api_plan_row_delete():
    """Plan satırını sil: satırın proje kaydı (region != '-') ve tüm hücreleri/atamaları silinir."""
    data = request.get_json(force=True, silent=True) or {}
    row_project_id = int(data.get("project_id") or 0)
    if not row_project_id:
        return jsonify({"ok": False, "error": "project_id eksik"}), 400

    p = Project.query.get(row_project_id)
    if not p or p.region == "-":
        return jsonify({"ok": False, "error": "Satır bulunamadı."}), 404

    cells = PlanCell.query.filter_by(project_id=row_project_id).all()
    cell_ids = [c.id for c in cells]
    if cell_ids:
        # Bağlı tüm kayıtları sil - Sıralama önemli olabilir (bağımlılıklar)
        # JobAssignment tablosu Job tablosuna bağlı, Job tablosu PlanCell'e bağlı
        # Bu yüzden Job'ları silmeden önce JobAssignment'ları silmek gerekebilir ama cascade varsa gerek yok.
        # Yine de SQLAlchemy ile silerken cascade'e güvenmek yerine manuel silmek daha güvenli
        
        # 1. PlanCell'e bağlı Job'ları bul ve sil
        jobs = Job.query.filter(Job.cell_id.in_(cell_ids)).all()
        job_ids = [j.id for j in jobs]
        if job_ids:
            # Job'a bağlı yan tablolar
            JobAssignment.query.filter(JobAssignment.job_id.in_(job_ids)).delete(synchronize_session=False)
            JobFeedback.query.filter(JobFeedback.job_id.in_(job_ids)).delete(synchronize_session=False)
            JobReport.query.filter(JobReport.job_id.in_(job_ids)).delete(synchronize_session=False)
            JobStatusHistory.query.filter(JobStatusHistory.job_id.in_(job_ids)).delete(synchronize_session=False)
            Job.query.filter(Job.id.in_(job_ids)).delete(synchronize_session=False)

        # 2. Diğer PlanCell bağımlılıklarını sil
        CellAssignment.query.filter(CellAssignment.cell_id.in_(cell_ids)).delete(synchronize_session=False)
        CellLock.query.filter(CellLock.cell_id.in_(cell_ids)).delete(synchronize_session=False)
        CellCancellation.query.filter(CellCancellation.cell_id.in_(cell_ids)).delete(synchronize_session=False)
        CellVersion.query.filter(CellVersion.cell_id.in_(cell_ids)).delete(synchronize_session=False)
        TeamOvertime.query.filter(TeamOvertime.cell_id.in_(cell_ids)).delete(synchronize_session=False)
        
        # 3. Son olarak PlanCell'leri sil
        PlanCell.query.filter(PlanCell.id.in_(cell_ids)).delete(synchronize_session=False)

    db.session.delete(p)
    db.session.commit()
    socketio.emit('update_table')
    return jsonify({"ok": True})


@planner_bp.post("/api/plan_row_copy")
@login_required
@observer_required
def api_plan_row_copy():
    """Plan satırını kopyala: seçilen satırın tüm hücrelerini, atamalarını ve bilgilerini yeni bir satıra kopyala."""
    data = request.get_json(force=True, silent=True) or {}
    source_project_id = int(data.get("source_project_id") or 0)
    week_start_str = data.get("week_start", "")
    
    if not source_project_id:
        return jsonify({"ok": False, "error": "source_project_id eksik"}), 400
    
    if not week_start_str:
        return jsonify({"ok": False, "error": "week_start eksik"}), 400
    
    source_p = Project.query.get(source_project_id)
    if not source_p or source_p.region == "-":
        return jsonify({"ok": False, "error": "Kaynak satır bulunamadı."}), 404
    
    # Yeni proje oluştur (aynı bilgilerle)
    new_p = Project(
        region=source_p.region,
        project_code=source_p.project_code,
        project_name=source_p.project_name,
        responsible=source_p.responsible,
        karsi_firma_sorumlusu=source_p.karsi_firma_sorumlusu,
        is_active=True
    )
    db.session.add(new_p)
    db.session.flush()  # ID'yi almak için
    
    # Haftanın tüm günlerini al
    ws = parse_date(week_start_str)
    if not ws:
        return jsonify({"ok": False, "error": "Geçersiz hafta başlangıcı"}), 400
    
    days = [ws + timedelta(days=i) for i in range(7)]
    
    # Kaynak projenin tüm hücrelerini kopyala
    source_cells = PlanCell.query.filter(
        PlanCell.project_id == source_project_id,
        PlanCell.work_date.in_(days)
    ).all()
    
    for source_cell in source_cells:
        # Yeni hücre oluştur
        new_cell = ensure_cell(new_p.id, source_cell.work_date)
        
        # Hücre bilgilerini kopyala
        new_cell.shift = source_cell.shift
        new_cell.note = source_cell.note
        new_cell.vehicle_info = source_cell.vehicle_info
        new_cell.team_id = source_cell.team_id
        new_cell.subproject_id = source_cell.subproject_id
        if hasattr(source_cell, 'important_note'):
            new_cell.important_note = source_cell.important_note
        if hasattr(source_cell, 'isdp_info'):
            new_cell.isdp_info = source_cell.isdp_info
        if hasattr(source_cell, 'po_info'):
            new_cell.po_info = source_cell.po_info
        if hasattr(source_cell, 'job_mail_body'):
            new_cell.job_mail_body = source_cell.job_mail_body
        if hasattr(source_cell, 'lld_hhd_files'):
            new_cell.lld_hhd_files = source_cell.lld_hhd_files
        if hasattr(source_cell, 'lld_hhd_path'):
            new_cell.lld_hhd_path = source_cell.lld_hhd_path
        if hasattr(source_cell, 'tutanak_files'):
            new_cell.tutanak_files = source_cell.tutanak_files
        if hasattr(source_cell, 'tutanak_path'):
            new_cell.tutanak_path = source_cell.tutanak_path
        
        # Personel atamalarını kopyala
        source_assignments = CellAssignment.query.filter_by(cell_id=source_cell.id).all()
        for source_assignment in source_assignments:
            new_assignment = CellAssignment(
                cell_id=new_cell.id,
                person_id=source_assignment.person_id
            )
            db.session.add(new_assignment)
    
    db.session.commit()
    socketio.emit('update_table')
    
    return jsonify({"ok": True, "new_project_id": new_p.id})


# ---------- DRAG DROP ----------
@planner_bp.post("/api/move_cell")
@login_required
@observer_required
def api_move_cell():
    """
    Hücre içeriğini sürükle-bırak ile taşı / yer değiştir.
    - Aynı proje satırı içinde veya farklı projeler arasında çalışır.
    - mode: "swap" (varsayılan) veya "move"
    """
    data = request.get_json(force=True, silent=True) or {}
    from_project_id = int(data.get("from_project_id") or data.get("project_id") or 0)
    to_project_id = int(data.get("to_project_id") or data.get("project_id") or 0)
    from_d = parse_date(data.get("from_date", ""))
    to_d = parse_date(data.get("to_date", ""))
    mode = (data.get("mode") or "swap").strip().lower()

    if not from_project_id or not to_project_id or not from_d or not to_d:
        return jsonify({"ok": False, "error": "from/to project_id veya from/to_date eksik"}), 400
    if from_project_id == to_project_id and from_d == to_d:
        return jsonify({"ok": True})

    src = PlanCell.query.filter_by(project_id=from_project_id, work_date=from_d).first()
    dst = PlanCell.query.filter_by(project_id=to_project_id, work_date=to_d).first()

    src_snap = snapshot_cell(src)
    dst_snap = snapshot_cell(dst)

    # hücreleri garanti altına al (boş hücreye bırakınca da oluşsun)
    src = src or ensure_cell(from_project_id, from_d)
    dst = dst or ensure_cell(to_project_id, to_d)

    if mode == "move":
        apply_snapshot(dst, src_snap)
        apply_snapshot(src, {"exists": False})
    else:
        apply_snapshot(dst, src_snap)
        apply_snapshot(src, dst_snap)

    db.session.commit()
    socketio.emit('update_table', namespace='/')
    return jsonify({"ok": True})


# ---------- MAP ----------
@planner_bp.route("/map")
@login_required
def map_page():
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", ""))
    start = ws if ws else week_start(d)
    return render_template("map.html", title="Harita", selected_week=iso(start))


@planner_bp.get("/api/map_markers")
def api_map_markers():
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", ""))
    start = ws if ws else week_start(d)
    days = [start + timedelta(days=i) for i in range(7)]

    cells = PlanCell.query.filter(PlanCell.work_date >= days[0], PlanCell.work_date <= days[-1]).all()
    project_ids_in_week = {c.project_id for c in cells}
    active_project_ids = {c.project_id for c in cells if len(c.assignments) > 0}

    by_city = {}
    if not project_ids_in_week:
        return jsonify({"week_start": iso(start), "markers": []})
    
    # Önce şehir bazında iş sayısını hesapla (PlanCell sayısı)
    city_job_counts = {}
    for cell in cells:
        if cell.project_id:
            proj = Project.query.get(cell.project_id)
            if proj:
                city = (proj.region or "").strip()
                if city:
                    key = city.lower()
                    city_job_counts[key] = city_job_counts.get(key, 0) + 1
    
    # Şehir bazında projeleri grupla ve iş sayısını ekle
    for p in Project.query.filter(Project.id.in_(project_ids_in_week)).order_by(Project.region.asc(), Project.project_code.asc()).all():
        city = (p.region or "").strip()
        if not city:
            continue
        key = city.lower()
        if key not in by_city:
            lat, lon = geocode_city(city)
            job_count = city_job_counts.get(key, 0)
            by_city[key] = {
                "city": city, 
                "lat": lat, 
                "lon": lon, 
                "active": False, 
                "projects": [],
                "job_count": job_count  # İş sayısını ekle
            }
        if p.id in active_project_ids:
            by_city[key]["active"] = True
        by_city[key]["projects"].append({"code": p.project_code, "name": p.project_name, "responsible": p.responsible})

    return jsonify({"week_start": iso(start), "markers": list(by_city.values())})


@planner_bp.get("/api/jobs_for_map")
def api_jobs_for_map():
    """
    Haftalık işleri haritaya yüklemek için detaylı veri döndürür.
    - Ekip bazında gruplama
    - Rota optimizasyon önerisi
    - Ekip uygunluk analizi
    """
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    team_id_filter = request.args.get("team_id", type=int)
    start = ws
    end = ws + timedelta(days=6)
    
    # Haftanın tüm hücrelerini al (ekip ve proje bilgileriyle)
    query = (
        db.session.query(PlanCell, Project, Team, SubProject)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(Team, Team.id == PlanCell.team_id)
        .outerjoin(SubProject, SubProject.id == PlanCell.subproject_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
    )
    
    if team_id_filter:
        query = query.filter(PlanCell.team_id == team_id_filter)
    
    rows = query.order_by(PlanCell.work_date.asc(), PlanCell.team_id.asc()).all()
    
    # İşleri ekip bazında grupla
    jobs_by_team = {}
    all_jobs = []
    
    for cell, proj, team, subproj in rows:
        city = (proj.region or "").strip()
        if not city:
            continue
            
        lat, lon = geocode_city(city)
        if lat is None or lon is None:
            continue
        
        # Personel atamaları
        assignments = []
        for a in cell.assignments:
            p = Person.query.get(a.person_id)
            if p:
                assignments.append({
                    "id": p.id,
                    "name": p.full_name,
                    "phone": p.phone or ""
                })
        
        team_id = cell.team_id or 0
        team_name = team.name if team else "Atanmamış"
        
        job_data = {
            "cell_id": cell.id,
            "date": iso(cell.work_date),
            "day_name": ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"][cell.work_date.weekday()],
            "city": city,
            "lat": lat,
            "lon": lon,
            "project_id": proj.id,
            "project_code": proj.project_code,
            "project_name": proj.project_name,
            "responsible": proj.responsible or "",
            "subproject": subproj.name if subproj else "",
            "team_id": team_id,
            "team_name": team_name,
            "shift": cell.shift or "",
            "note": cell.note or "",
            "important_note": getattr(cell, 'important_note', '') or "",
            "status": getattr(cell, 'status', 'active') or "active",
            "assignments": assignments,
            "assignment_count": len(assignments)
        }
        
        all_jobs.append(job_data)
        
        if team_id not in jobs_by_team:
            jobs_by_team[team_id] = {
                "team_id": team_id,
                "team_name": team_name,
                "color": team_color(team_id) if team_id else "#94a3b8",
                "jobs": [],
                "total_km": 0,
                "cities": set(),
                "daily_jobs": {}
            }
        
        jobs_by_team[team_id]["jobs"].append(job_data)
        jobs_by_team[team_id]["cities"].add(city)
        
        date_str = iso(cell.work_date)
        if date_str not in jobs_by_team[team_id]["daily_jobs"]:
            jobs_by_team[team_id]["daily_jobs"][date_str] = []
        jobs_by_team[team_id]["daily_jobs"][date_str].append(job_data)
    
    # Simplified team listing without analysis
    teams_analysis = []
    for tid, team_data in jobs_by_team.items():
        teams_analysis.append({
            "team_id": tid,
            "team_name": team_data["team_name"],
            "color": team_data["color"],
            "job_count": len(team_data["jobs"]),
            "city_count": len(team_data["cities"]),
            "total_km": 0,
            "estimated_hours": 0,
            "feasibility": "uygun", 
            "feasibility_note": "",
            "jobs": team_data["jobs"]
        })
    
    return jsonify({
        "ok": True,
        "week_start": iso(ws),
        "total_jobs": len(all_jobs),
        "total_teams": len(teams_analysis),
        "teams": teams_analysis,
        "all_jobs": all_jobs
    })


@planner_bp.get("/api/routes_all")
def api_routes_all():
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)

    # team_id'leri bul
    team_ids = (
        db.session.query(PlanCell.team_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .filter(PlanCell.team_id.isnot(None))
        .distinct()
        .all()
    )
    team_ids = [t[0] for t in team_ids if t[0]]
    if not team_ids:
        return jsonify({"ok": True, "routes": []})

    # her ekip için sıralı şehir listesi
    routes = []
    for tid in sorted(team_ids):
        # Önceki haftanın tüm rotasını bul
        prev_week_start = start - timedelta(days=7)
        prev_week_end = prev_week_start + timedelta(days=6)
        prev_week_rows = (
            db.session.query(PlanCell.work_date, Project.region, Project.project_code)
            .join(Project, Project.id == PlanCell.project_id)
            .filter(PlanCell.work_date >= prev_week_start, PlanCell.work_date <= prev_week_end)
            .filter(PlanCell.team_id == tid)
            .order_by(PlanCell.work_date.asc(), db.case((PlanCell.shift == "Gündüz", 1),
                                                       (PlanCell.shift == "Gündüz Yol", 2),
                                                       (PlanCell.shift == "Gece", 3),
                                                       else_=9).asc(),
                      Project.project_code.asc())
            .all()
        )
        
        prev_week_points = []
        prev_seen = set()
        for wd, city, pcode in prev_week_rows:
            city = (city or "").strip()
            if not city:
                continue
            key = (wd, city.lower())
            if key in prev_seen:
                continue
            prev_seen.add(key)
            lat, lon = geocode_city(city)
            if lat is not None and lon is not None:
                prev_week_points.append({"date": iso(wd), "city": city, "project_code": pcode, "lat": lat, "lon": lon})
        
        # Önceki hafta yoksa Ankara'dan başla
        if not prev_week_points:
            ankara_lat, ankara_lon = geocode_city("Ankara")
            if ankara_lat is not None and ankara_lon is not None:
                prev_week_points.append({"date": iso(start - timedelta(days=1)), "city": "Ankara", "project_code": "", "lat": ankara_lat, "lon": ankara_lon})
        
        rows = (
            db.session.query(PlanCell.work_date, Project.region, Project.project_code)
            .join(Project, Project.id == PlanCell.project_id)
            .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
            .filter(PlanCell.team_id == tid)
            .order_by(PlanCell.work_date.asc(), db.case((PlanCell.shift == "Gündüz", 1),
                                                       (PlanCell.shift == "Gündüz Yol", 2),
                                                       (PlanCell.shift == "Gece", 3),
                                                       else_=9).asc(),
                      Project.project_code.asc())
            .all()
        )

        seen = set()
        pts = []
        # Önceki haftanın son noktasını ekle (eğer varsa)
        if prev_week_points:
            prev_pt = prev_week_points[-1]  # Son nokta
            pts.append(prev_pt)
        
        for wd, city, pcode in rows:
            city = (city or "").strip()
            if not city:
                continue
            key = (wd, city.lower())
            if key in seen:
                continue
            seen.add(key)
            lat, lon = geocode_city(city)
            if lat is None or lon is None:
                continue
            pts.append({"date": iso(wd), "city": city, "project_code": pcode, "lat": lat, "lon": lon})

            pts.append({"date": iso(wd), "city": city, "project_code": pcode, "lat": lat, "lon": lon})

        t = Team.query.get(tid)
        
        # Ekip sirasina gore 1'den baslayan numara ver (team_ids zaten sorted)
        display_number = team_ids.index(tid) + 1
        display_team_name = f"Ekip {display_number}"
        
        routes.append({
            "team_id": tid,
            "team_name": display_team_name,
            "color": team_color(tid),
            "points": pts,
            "prev_week_points": prev_week_points
        })

    return jsonify({"ok": True, "week_start": iso(ws), "routes": routes})


@planner_bp.get("/api/route")
def api_route_alias():
    # Backward compatible: single team route
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    team_id = int(request.args.get("team_id", 0) or 0)
    if not team_id:
        return jsonify({"ok": False, "error": "team_id eksik"}), 400
    # reuse routes_team logic
    with current_app.test_request_context():
        pass
    # call same code as api_routes_team
    start = ws
    end = ws + timedelta(days=6)
    rows = (
        db.session.query(PlanCell.work_date, Project.region, Project.project_code)
        .join(Project, Project.id == PlanCell.project_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .filter(PlanCell.team_id == team_id)
        .order_by(PlanCell.work_date.asc(),
                  db.case((PlanCell.shift == "Gündüz", 1),
                          (PlanCell.shift == "Gündüz Yol", 2),
                          (PlanCell.shift == "Gece", 3),
                          else_=9).asc(),
                  Project.project_code.asc())
        .all()
    )
    seen = set()
    pts = []
    for wd, city, pcode in rows:
        city = (city or "").strip()
        if not city:
            continue
        key = (wd, city.lower())
        if key in seen:
            continue
        seen.add(key)
        lat, lon = geocode_city(city)
        if lat is None or lon is None:
            continue
        pts.append({"date": iso(wd), "city": city, "project_code": pcode, "lat": lat, "lon": lon})
    t = Team.query.get(team_id)
    return jsonify({
        "ok": True,
        "week_start": iso(ws),
        "team_id": team_id,
        "team_name": (t.name if t else f"Ekip #{team_id}"),
        "color": team_color(team_id),
        "points": pts
    })

@planner_bp.get("/api/routes_team")
def api_routes_team():
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    team_id = int(request.args.get("team_id", 0) or 0)
    start = ws
    end = ws + timedelta(days=6)

    if not team_id:
        return jsonify({"ok": False, "error": "team_id eksik"}), 400

    # Önceki haftanın tüm rotasını bul (Sadece çizgi için temel bilgi yeterli)
    prev_week_start = start - timedelta(days=7)
    prev_week_end = prev_week_start + timedelta(days=6)
    prev_week_rows = (
        db.session.query(PlanCell.work_date, Project.region, Project.project_code)
        .join(Project, Project.id == PlanCell.project_id)
        .filter(PlanCell.work_date >= prev_week_start, PlanCell.work_date <= prev_week_end)
        .filter(PlanCell.team_id == team_id)
        .order_by(PlanCell.work_date.asc(), db.case((PlanCell.shift == "Gündüz", 1),
                                                   (PlanCell.shift == "Gündüz Yol", 2),
                                                   (PlanCell.shift == "Gece", 3),
                                                   else_=9).asc(),
                  Project.project_code.asc())
        .all()
    )
    
    prev_week_points = []
    prev_seen = set()
    for wd, city, pcode in prev_week_rows:
        city = (city or "").strip()
        if not city:
            continue
        key = (wd, city.lower())
        if key in prev_seen:
            continue
        prev_seen.add(key)
        lat, lon = geocode_city(city)
        if lat is not None and lon is not None:
            prev_week_points.append({"date": iso(wd), "city": city, "project_code": pcode, "lat": lat, "lon": lon})
    
    # Önceki hafta yoksa Ankara'dan başla (Depot varsayımı)
    if not prev_week_points:
        ankara_lat, ankara_lon = geocode_city("Ankara")
        if ankara_lat is not None and ankara_lon is not None:
            prev_week_points.append({"date": iso(start - timedelta(days=1)), "city": "Ankara", "project_code": "DEPOT", "lat": ankara_lat, "lon": ankara_lon})

    # Bu haftanın verilerini zenginleştir
    # PlanCell + Project + SubProject (optional) + Job (optional for status)
    rows = (
        db.session.query(PlanCell, Project, SubProject, Job)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(SubProject, SubProject.id == PlanCell.subproject_id)
        .outerjoin(Job, Job.cell_id == PlanCell.id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .filter(PlanCell.team_id == team_id)
        .order_by(PlanCell.work_date.asc(),
                  db.case((PlanCell.shift == "Gündüz", 1),
                          (PlanCell.shift == "Gündüz Yol", 2),
                          (PlanCell.shift == "Gece", 3),
                          else_=9).asc(),
                  Project.project_code.asc())
        .all()
    )

    seen = set()
    pts = []
    # Önceki haftanın son noktası başlangıç noktası olarak eklenebilir mi? 
    # Frontend çizimi için prev_week_points ayrı gönderiliyor.
    # Ancak "Depot" marker'ı için listenin başına eklemek isteyebiliriz.
    # Şimdilik listenin başına prev_week_points'in sonuncusunu "Depot/Başlangıç" olarak ekleyelim.
    if prev_week_points:
        prev_pt = prev_week_points[-1]
        pts.append({
            "lat": prev_pt["lat"],
            "lon": prev_pt["lon"],
            "date": prev_pt["date"],
            "city": prev_pt["city"],
            "project_code": prev_pt["project_code"],
            "project_name": "Başlangıç / Devir",
            "type": "start" # Frontend'de ayırt etmek için
        })

    for cell, proj, subproj, job in rows:
        city = (proj.region or "").strip()
        if not city:
            continue
        
        # Aynı gün/şehir olsa bile detayları farklı olabilir, ama haritada üst üste biner.
        # Frontend, aynı koordinattaki noktaları gruplayabilir veya hafif offset verebilir.
        # Biz hepsini gönderelim.
        
        lat, lon = geocode_city(city)
        if lat is None or lon is None:
            continue
            
        pts.append({
            "lat": lat,
            "lon": lon,
            "date": iso(cell.work_date),
            "city": city,
            "project_code": proj.project_code,
            "project_name": proj.project_name,
            "responsible": proj.responsible or "Belirtilmemiş",
            "karsi_sorumlu": getattr(proj, 'karsi_firma_sorumlusu', '') or "",
            "subproject_name": subproj.name if subproj else "",
            "service_type": subproj.name if subproj else "Genel Çalışma",
            "shift": cell.shift or "",
            "note": cell.note or "",
            "important_note": cell.important_note or "",
            "status": job.kanban_status if job else "PLANLANDI",
            "job_id": job.id if job else None,
            "type": "stop"
        })

    # Son nokta (Bitiş/Depot) eklenebilir, şimdilik rota son müşteride bitiyor.
    
    # Team info
    team = Team.query.get(team_id)
    team_name = team.name if team else f"Ekip {team_id}"

    return jsonify({
        "ok": True,
        "week_start": iso(ws),
        "route": {
            "team_id": team_id,
            "team_name": team_name,
            "color": team_color(team_id),
            "points": pts,
            "prev_week_points": prev_week_points
        }
    })


# ---------- REPORTS ----------
@planner_bp.get("/reports")
@login_required
def reports_page():
    # ensure_schema is called via before_request hook in app.py

    d0 = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", "")) or week_start(d0)
    start = ws
    end = ws + timedelta(days=6)

    tab = (request.args.get("tab") or "open").strip().lower()
    if tab not in ("open", "team", "errors"):
        tab = "open"

    city = (request.args.get("city") or "").strip()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()

    upsert_jobs_for_range(start, end)

    # Filters data
    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()
    cities = sorted({(p.region or "").strip() for p in projects if (p.region or "").strip()})
    subprojects = []
    if project_id:
        owner_pid = _effective_main_project_id_for_subprojects(project_id)
        subprojects = SubProject.query.filter(SubProject.project_id == owner_pid, SubProject.is_active == True).order_by(SubProject.name.asc()).all()

    q0 = Job.query.filter(Job.work_date >= start, Job.work_date <= end).join(Project, Project.id == Job.project_id)
    if city:
        q0 = q0.filter(Project.region == city)
    if project_id:
        q0 = q0.filter(Job.project_id == project_id)
    if subproject_id:
        q0 = q0.filter(Job.subproject_id == subproject_id)
    if team_name:
        q0 = q0.filter(Job.team_name == team_name)

    teams = sorted({(x or "").strip() for (x,) in q0.with_entities(Job.team_name).distinct().all() if (x or "").strip()})

    open_jobs = []
    team_perf = []
    error_rows = []

    if tab == "open":
        open_jobs = (
            q0.filter(Job.kanban_status != "CLOSED")
              .order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc(), Job.id.asc())
              .all()
        )

    if tab == "team":
        jobs = q0.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc()).all()

        # Mail send time map for closure duration (assignment mail -> feedback close)
        send_at = {}
        try:
            import json as _json
            logs = MailLog.query.filter(
                MailLog.created_at >= datetime.combine(start - timedelta(days=7), datetime.min.time()),
                MailLog.created_at <= datetime.combine(end + timedelta(days=7), datetime.max.time()),
            ).all()
            for r in logs:
                if not r.ok or (r.kind or "") != "send":
                    continue
                meta = {}
                try:
                    meta = _json.loads(r.meta_json or "{}")
                except Exception:
                    meta = {}
                if meta.get("type") != "job":
                    continue
                pid = int(meta.get("project_id") or 0)
                wd = str(meta.get("work_date") or "")
                if not pid or not wd:
                    continue
                k = (pid, wd)
                if k not in send_at or (r.created_at and r.created_at < send_at[k]):
                    send_at[k] = r.created_at
        except Exception:
            send_at = {}

        agg = {}
        for j in jobs:
            key = (j.team_name or "").strip() or "-"
            if key not in agg:
                agg[key] = {
                    "team": key,
                    "total": 0,
                    "completed": 0,
                    "problem": 0,
                    "open": 0,
                    "close_deltas": [],
                }
            a = agg[key]
            a["total"] += 1
            if (j.status or "") == "completed":
                a["completed"] += 1
            if (j.status or "") == "problem":
                a["problem"] += 1
            if _normalize_kanban_status(getattr(j, "kanban_status", None)) != "CLOSED":
                a["open"] += 1

            if (j.status or "") == "completed" and j.closed_at:
                sent_time = send_at.get((int(j.project_id), iso(j.work_date)))
                if sent_time:
                    delta = (j.closed_at - sent_time).total_seconds()
                    if delta >= 0:
                        a["close_deltas"].append(delta)

        out = []
        for a in agg.values():
            avg_close_hours = None
            if a["close_deltas"]:
                avg_close_hours = round((sum(a["close_deltas"]) / len(a["close_deltas"])) / 3600.0, 2)
            out.append({
                "team": a["team"],
                "total": a["total"],
                "open": a["open"],
                "completed": a["completed"],
                "problem": a["problem"],
                "avg_close_hours": avg_close_hours,
            })
        team_perf = sorted(out, key=lambda x: (-x["total"], x["team"]))

    if tab == "errors":
        uid = session.get("user_id")
        if uid:
            start_dt = datetime.combine(start, datetime.min.time())
            end_dt = datetime.combine(end, datetime.max.time())
            error_rows = (
                Notification.query
                .filter(Notification.user_id == uid)
                .filter(Notification.event.in_(["mail_fail", "feedback", "sla"]))
                .filter(Notification.created_at >= start_dt, Notification.created_at <= end_dt)
                .order_by(Notification.created_at.desc())
                .limit(200)
                .all()
            )

    excel_url = url_for("planner.reports_excel", tab=tab, week_start=iso(ws), city=city, project_id=(project_id or ""), subproject_id=(subproject_id or ""), team_name=team_name)

    return render_template(
        "reports.html",
        tab=tab,
        week_start=iso(ws),
        prev_week=iso(ws - timedelta(days=7)),
        next_week=iso(ws + timedelta(days=7)),
        city=city,
        project_id=project_id,
        subproject_id=subproject_id,
        subprojects=subprojects,
        team_name=team_name,
        projects=projects,
        cities=cities,
        teams=teams,
        open_jobs=open_jobs,
        team_perf=team_perf,
        error_rows=error_rows,
        excel_url=excel_url,
    )


@planner_bp.get("/reports.xlsx")
@login_required
def reports_excel():
    # ensure_schema is called via before_request hook in app.py

    d0 = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", "")) or week_start(d0)
    start = ws
    end = ws + timedelta(days=6)

    city = (request.args.get("city") or "").strip()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()

    upsert_jobs_for_range(start, end)

    q0 = Job.query.filter(Job.work_date >= start, Job.work_date <= end).join(Project, Project.id == Job.project_id)
    if city:
        q0 = q0.filter(Project.region == city)
    if project_id:
        q0 = q0.filter(Job.project_id == project_id)
    if subproject_id:
        q0 = q0.filter(Job.subproject_id == subproject_id)
    if team_name:
        q0 = q0.filter(Job.team_name == team_name)

    jobs_all = q0.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc(), Job.id.asc()).all()

    # Team performance
    send_at = {}
    try:
        import json as _json
        logs = MailLog.query.filter(
            MailLog.created_at >= datetime.combine(start - timedelta(days=7), datetime.min.time()),
            MailLog.created_at <= datetime.combine(end + timedelta(days=7), datetime.max.time()),
        ).all()
        for r in logs:
            if not r.ok or (r.kind or "") != "send":
                continue
            meta = {}
            try:
                meta = _json.loads(r.meta_json or "{}")
            except Exception:
                meta = {}
            if meta.get("type") != "job":
                continue
            pid = int(meta.get("project_id") or 0)
            wd = str(meta.get("work_date") or "")
            if not pid or not wd:
                continue
            k = (pid, wd)
            if k not in send_at or (r.created_at and r.created_at < send_at[k]):
                send_at[k] = r.created_at
    except Exception:
        send_at = {}

    agg = {}
    for j in jobs_all:
        key = (j.team_name or "").strip() or "-"
        if key not in agg:
            agg[key] = {
                "team": key,
                "total": 0,
                "completed": 0,
                "problem": 0,
                "open": 0,
                "close_deltas": [],
            }
        a = agg[key]
        a["total"] += 1
        if (j.status or "") == "completed":
            a["completed"] += 1
        if (j.status or "") == "problem":
            a["problem"] += 1
        if _normalize_kanban_status(getattr(j, "kanban_status", None)) != "CLOSED":
            a["open"] += 1

        if (j.status or "") == "completed" and j.closed_at:
            sent_time = send_at.get((int(j.project_id), iso(j.work_date)))
            if sent_time:
                delta = (j.closed_at - sent_time).total_seconds()
                if delta >= 0:
                    a["close_deltas"].append(delta)

    team_rows = []
    for a in agg.values():
        avg_close_hours = None
        if a["close_deltas"]:
            avg_close_hours = round((sum(a["close_deltas"]) / len(a["close_deltas"])) / 3600.0, 2)
        team_rows.append([
            a["team"],
            a["total"],
            a["open"],
            a["completed"],
            a["problem"],
            avg_close_hours if avg_close_hours is not None else "",
        ])
    team_rows.sort(key=lambda r: (-int(r[1] or 0), str(r[0])))

    # Error queue
    uid = session.get("user_id")
    error_rows = []
    if uid:
        start_dt = datetime.combine(start, datetime.min.time())
        end_dt = datetime.combine(end, datetime.max.time())
        rows = (
            Notification.query
            .filter(Notification.user_id == uid)
            .filter(Notification.event.in_(["mail_fail", "feedback", "sla"]))
            .filter(Notification.created_at >= start_dt, Notification.created_at <= end_dt)
            .order_by(Notification.created_at.desc())
            .limit(200)
            .all()
        )
        for n in rows:
            error_rows.append([
                n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else "",
                n.event,
                "1" if n.read_at else "0",
                n.title,
                (n.body or "")[:500],
                n.link_url or "",
            ])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Acik Isler"
    ws1.append(["ID", "Date", "City", "Project Code", "Project Name", "Sub-Project", "Team", "Kanban", "Status", "Updated At"])
    for j in jobs_all:
        if _normalize_kanban_status(getattr(j, "kanban_status", None)) == "CLOSED":
            continue
        p = j.project
        ws1.append([
            int(j.id),
            iso(j.work_date) if j.work_date else "",
            (p.region if p else ""),
            (p.project_code if p else ""),
            (p.project_name if p else ""),
            (j.subproject.name if getattr(j, "subproject", None) else ""),
            j.team_name or "",
            _normalize_kanban_status(getattr(j, "kanban_status", None)),
            j.status,
            j.updated_at.strftime("%Y-%m-%d %H:%M") if j.updated_at else "",
        ])

    ws2 = wb.create_sheet("Ekip Performans")
    ws2.append(["Ekip", "Toplam", "Açık", "Tamamlandı", "Sorun", "Ort. Kapanış (Saat)"])
    for r in team_rows:
        ws2.append(r)

    ws3 = wb.create_sheet("Hata Kuyrugu")
    ws3.append(["Created At", "Event", "Read", "Title", "Body", "Link"])
    for r in error_rows:
        ws3.append(r)

    for wsht in [ws1, ws2, ws3]:
        for c in range(1, wsht.max_column + 1):
            wsht.cell(row=1, column=c).font = Font(bold=True)
        for c in range(1, wsht.max_column + 1):
            wsht.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"reports_{iso(ws)}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# ---------- ADVANCED REPORTS ----------
def _parse_date_range_args() -> Tuple[date, date]:
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    if start_s and end_s:
        s = parse_date(start_s) or date.today()
        e = parse_date(end_s) or s
        return s, e
    # If an explicit week/date is provided, keep week-based behavior (backwards compatible).
    if (request.args.get("week_start") or "").strip() or (request.args.get("date") or "").strip():
        d0 = parse_date(request.args.get("date", "")) or date.today()
        ws = parse_date(request.args.get("week_start", "")) or week_start(d0)
        return ws, ws + timedelta(days=6)
    # default: last 7 days (inclusive)
    e = date.today()
    s = e - timedelta(days=6)
    return s, e


def _parse_board_date_range_args() -> Tuple[date, date]:
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    if start_s or end_s:
        s = parse_date(start_s) or date.today()
        e = parse_date(end_s) or s
        return s, e

    today = date.today()
    return today - timedelta(days=30), today + timedelta(days=30)


def _parse_bool_arg(name: str, default: bool = False) -> bool:
    v = (request.args.get(name) or "").strip().lower()
    if v == "":
        return default
    return v in ("1", "true", "t", "yes", "y", "on")


def _build_advanced_jobs_query(
    start: date,
    end: date,
    status: str,
    project_id: int,
    subproject_id: int,
    team_name: str,
    city: str,
    only_overdue: bool,
    only_problem: bool,
):
    q = (
        Job.query
        .join(Project, Project.id == Job.project_id)
        .outerjoin(SubProject, SubProject.id == Job.subproject_id)
        .filter(Job.work_date >= start, Job.work_date <= end)
    )
    if status:
        q = q.filter(Job.status == _job_status_label(status))
    if only_problem:
        q = q.filter(Job.status == "problem")
    if only_overdue:
        today = date.today()
        q = q.filter(Job.status != "completed", Job.work_date < today)
    if project_id:
        q = q.filter(Job.project_id == project_id)
    if subproject_id:
        q = q.filter(Job.subproject_id == subproject_id)
    if team_name:
        q = q.filter(Job.team_name == team_name)
    if city:
        q = q.filter(Project.region == city)
    return q


def _job_status_label(st: str) -> str:
    st = (st or "").lower().strip()
    if st == "completed":
        return "completed"
    if st == "problem":
        return "problem"
    return "pending"


@planner_bp.get("/reports/advanced")
@admin_required
def reports_advanced_page():
    start, end = _parse_date_range_args()
    tab = (request.args.get("tab") or "dashboard").strip().lower()
    status = (request.args.get("status") or "").strip().lower()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()
    city = (request.args.get("city") or "").strip()
    include_leave = _parse_bool_arg("include_leave", default=False)
    only_overdue = _parse_bool_arg("only_overdue", default=False)
    only_problem = _parse_bool_arg("only_problem", default=False)

    if project_id <= 0:
        subproject_id = 0
    elif subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id)):
        subproject_id = 0

    try:
        page = int(request.args.get("page", 1) or 1)
    except Exception:
        page = 1
    page = max(1, page)
    try:
        page_size = int(request.args.get("page_size", 50) or 50)
    except Exception:
        page_size = 50
    page_size = max(10, min(200, page_size))

    upsert_jobs_for_range(start, end)

    today = date.today()
    q = _build_advanced_jobs_query(start, end, status, project_id, subproject_id, team_name, city, only_overdue, only_problem)

    # Metrics
    total_jobs = int(q.count() or 0)
    completed = int(q.filter(Job.status == "completed").count() or 0)
    problem = int(q.filter(Job.status == "problem").count() or 0)
    pending = max(0, total_jobs - completed - problem)
    overdue = int(q.filter(Job.status != "completed", Job.work_date < today).count() or 0)
    completion_rate = round((completed / total_jobs) * 100.0, 1) if total_jobs else 0.0
    cancel_rate = round((problem / total_jobs) * 100.0, 1) if total_jobs else 0.0

    # Kanban distribution + lifecycle KPIs
    kanban_counts = {k: 0 for k in KANBAN_COLUMNS}
    try:
        for st, cnt in (
            q.with_entities(Job.kanban_status, db.func.count(Job.id))
            .group_by(Job.kanban_status)
            .all()
        ):
            kanban_counts[_normalize_kanban_status(st)] = int(cnt or 0)
    except Exception:
        pass

    # publish->report and report->close (hours)
    avg_publish_to_report_hours = None
    avg_report_to_close_hours = None
    try:
        job_ids = [int(x) for (x,) in q.with_entities(Job.id).all()]
        if job_ids:
            # latest feedback per job
            fb_rows = (
                JobFeedback.query
                .filter(JobFeedback.job_id.in_(job_ids), JobFeedback.outcome != None)
                .order_by(JobFeedback.job_id.asc(), JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
                .all()
            )
            latest_fb = {}
            for r in fb_rows:
                if r.job_id not in latest_fb:
                    latest_fb[r.job_id] = r

            jobs_rows = (
                Job.query
                .filter(Job.id.in_(job_ids))
                .with_entities(Job.id, Job.published_at, Job.closed_at)
                .all()
            )
            publish_deltas = []
            close_deltas = []
            for jid, pub_at, closed_at in jobs_rows:
                fb = latest_fb.get(int(jid))
                if fb and pub_at and fb.submitted_at:
                    d = (fb.submitted_at - pub_at).total_seconds()
                    if d >= 0:
                        publish_deltas.append(d)
                if fb and closed_at and fb.submitted_at:
                    d = (closed_at - fb.submitted_at).total_seconds()
                    if d >= 0:
                        close_deltas.append(d)
            if publish_deltas:
                avg_publish_to_report_hours = round((sum(publish_deltas) / len(publish_deltas)) / 3600.0, 2)
            if close_deltas:
                avg_report_to_close_hours = round((sum(close_deltas) / len(close_deltas)) / 3600.0, 2)
    except Exception:
        pass

    # team counts
    team_key = db.func.coalesce(db.func.nullif(db.func.trim(Job.team_name), ""), "-")
    top_teams = (
        q.with_entities(team_key, db.func.count(Job.id))
        .group_by(team_key)
        .order_by(db.func.count(Job.id).desc(), team_key.asc())
        .limit(15)
        .all()
    )

    team_workload = []
    try:
        rows = (
            q.with_entities(
                team_key.label("team_name"),
                db.func.count(Job.id).label("total"),
                db.func.sum(db.case((Job.status == "completed", 1), else_=0)).label("completed"),
                db.func.sum(db.case((Job.status == "problem", 1), else_=0)).label("problem"),
                db.func.sum(db.case((Job.is_published == True, 1), else_=0)).label("on_site"),
            )
            .group_by(team_key)
            .order_by(db.func.count(Job.id).desc(), team_key.asc())
            .limit(50)
            .all()
        )
        for tn, ttotal, tcompleted, tproblem, ton_site in rows:
            total_i = int(ttotal or 0)
            completed_i = int(tcompleted or 0)
            problem_i = int(tproblem or 0)
            on_site_i = int(ton_site or 0)
            team_workload.append(
                {
                    "team": (tn or "").strip() or "-",
                    "total": total_i,
                    "completed": completed_i,
                    "problem": problem_i,
                    "on_site": on_site_i,
                    "completion_rate": round((completed_i / total_i) * 100.0, 1) if total_i else 0.0,
                    "cancel_rate": round((problem_i / total_i) * 100.0, 1) if total_i else 0.0,
                }
            )
    except Exception:
        team_workload = []

    # city counts
    city_key = db.func.coalesce(db.func.nullif(db.func.trim(Project.region), ""), "-")
    top_cities = (
        q.with_entities(city_key, db.func.count(Job.id))
        .group_by(city_key)
        .order_by(db.func.count(Job.id).desc(), city_key.asc())
        .limit(15)
        .all()
    )

    published_total = int(q.filter(Job.is_published == True).count() or 0)
    top_on_site_project = None
    top_on_site_subproject = None
    try:
        row = (
            q.filter(Job.is_published == True)
            .with_entities(Project.project_code, Project.project_name, db.func.count(Job.id))
            .group_by(Project.id)
            .order_by(db.func.count(Job.id).desc(), Project.project_code.asc())
            .first()
        )
        if row:
            p_code, p_name, cnt = row
            label = (p_code or "").strip()
            if (p_name or "").strip():
                label = f"{label} - {p_name}".strip(" -")
            top_on_site_project = {"label": label or "-", "count": int(cnt or 0)}
    except Exception:
        top_on_site_project = None
    try:
        row = (
            q.filter(Job.is_published == True, SubProject.id.isnot(None))
            .with_entities(SubProject.name, SubProject.code, db.func.count(Job.id))
            .group_by(SubProject.id)
            .order_by(db.func.count(Job.id).desc(), SubProject.name.asc())
            .first()
        )
        if row:
            sp_name, sp_code, cnt = row
            label = (sp_name or "").strip() or "-"
            if (sp_code or "").strip():
                label = f"{sp_code} - {label}"
            top_on_site_subproject = {"label": label, "count": int(cnt or 0)}
    except Exception:
        top_on_site_subproject = None

    # Pagination (Jobs tab)
    total_count = total_jobs
    total_pages = max(1, (total_count + page_size - 1) // page_size) if total_count else 1
    page = min(page, total_pages)
    jobs = []
    if tab == "jobs":
        jobs = (
            q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

    # Avg closure (assignment -> feedback.closed_at) using MailLog meta(type=job)
    avg_close_hours = None
    team_avg_close_hours = {}
    try:
        import json as _json
        # mail logs that match range +/- 7 days to capture send time
        logs = MailLog.query.filter(
            MailLog.created_at >= datetime.combine(start - timedelta(days=7), datetime.min.time()),
            MailLog.created_at <= datetime.combine(end + timedelta(days=7), datetime.max.time()),
            MailLog.ok == True,
            (MailLog.kind == "send"),
        ).all()
        send_at = {}  # (project_id, work_date_iso) -> earliest datetime
        for r in logs:
            meta = {}
            try:
                meta = _json.loads(r.meta_json or "{}")
            except Exception:
                meta = {}
            if meta.get("type") != "job":
                continue
            pid = int(meta.get("project_id") or 0)
            wd = str(meta.get("work_date") or "")
            if not pid or not wd:
                continue
            k = (pid, wd)
            if k not in send_at or (r.created_at and r.created_at < send_at[k]):
                send_at[k] = r.created_at

        deltas = []
        closed_rows = (
            q.filter(Job.closed_at.isnot(None))
            .with_entities(Job.project_id, Job.work_date, Job.closed_at, Job.team_name)
            .all()
        )
        by_team = {}
        for pid, wd, closed_at, tname in closed_rows:
            k = (int(pid), iso(wd))
            sent_time = send_at.get(k)
            if not sent_time or not closed_at:
                continue
            delta = (closed_at - sent_time).total_seconds()
            if delta < 0:
                continue
            deltas.append(delta)
            tn = (tname or "").strip() or "-"
            by_team.setdefault(tn, []).append(delta)
        if deltas:
            avg_close_hours = round((sum(deltas) / len(deltas)) / 3600.0, 2)
        for tn, td in by_team.items():
            if td:
                team_avg_close_hours[tn] = round((sum(td) / len(td)) / 3600.0, 2)
    except Exception:
        avg_close_hours = None
        team_avg_close_hours = {}

    # Filters data
    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()
    subprojects = []
    if project_id:
        owner_id = _effective_main_project_id_for_subprojects(project_id)
        subprojects = (
            SubProject.query
            .filter(SubProject.project_id == owner_id, SubProject.is_active == True)
            .order_by(SubProject.name.asc())
            .all()
        )
    teams = sorted({(r[0] or "").strip() for r in q.with_entities(Job.team_name).distinct().all() if (r[0] or "").strip()})
    cities = sorted({(p.region or "").strip() for p in projects if (p.region or "").strip()})
    team_total = len(teams)
    jobs_per_team = round((total_jobs / team_total), 2) if team_total else float(total_jobs)

    # Mail stats in range
    mail_rows = MailLog.query.filter(
        MailLog.created_at >= datetime.combine(start, datetime.min.time()),
        MailLog.created_at <= datetime.combine(end, datetime.max.time())
    ).order_by(MailLog.created_at.desc()).limit(500).all()
    mail_total = len(mail_rows)
    mail_ok = sum(1 for r in mail_rows if r.ok)
    mail_fail = mail_total - mail_ok
    fail_reasons = {}
    for r in mail_rows:
        if r.ok:
            continue
        msg = (r.error or "error")[:120]
        fail_reasons[msg] = fail_reasons.get(msg, 0) + 1
    top_fail_reasons = sorted(fail_reasons.items(), key=lambda x: (-x[1], x[0]))[:10]

    # Build jobs table rows with people count
    job_ids = [j.id for j in jobs]
    people_count = {}
    if job_ids:
        if include_leave:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .filter(JobAssignment.job_id.in_(job_ids))
                .group_by(JobAssignment.job_id)
                .all()
            )
        else:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .join(Job, Job.id == JobAssignment.job_id)
                .outerjoin(
                    PersonDayStatus,
                    (PersonDayStatus.person_id == JobAssignment.person_id) & (PersonDayStatus.work_date == Job.work_date),
                )
                .filter(JobAssignment.job_id.in_(job_ids))
                .filter(or_(PersonDayStatus.id == None, PersonDayStatus.status != "leave"))
                .group_by(JobAssignment.job_id)
                .all()
            )
        for jid, cnt in rows:
            people_count[int(jid)] = int(cnt or 0)

    return render_template(
        "reports_advanced.html",
        tab=tab,
        start=iso(start),
        end=iso(end),
        status=status,
        project_id=project_id,
        subproject_id=subproject_id,
        team_name=team_name,
        city=city,
        include_leave=include_leave,
        only_overdue=only_overdue,
        only_problem=only_problem,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        total_count=total_count,
        projects=projects,
        subprojects=subprojects,
        teams=teams,
        cities=cities,
        total_jobs=total_jobs,
        completed=completed,
        pending=pending,
        problem=problem,
        overdue=overdue,
        completion_rate=completion_rate,
        cancel_rate=cancel_rate,
        published_total=published_total,
        top_on_site_project=top_on_site_project,
        top_on_site_subproject=top_on_site_subproject,
        team_workload=team_workload,
        kanban_counts=kanban_counts,
        avg_publish_to_report_hours=avg_publish_to_report_hours,
        avg_report_to_close_hours=avg_report_to_close_hours,
        team_total=team_total,
        jobs_per_team=jobs_per_team,
        avg_close_hours=avg_close_hours,
        team_avg_close_hours=team_avg_close_hours,
        top_teams=top_teams,
        top_cities=top_cities,
        jobs=jobs,
        people_count=people_count,
        mail_total=mail_total,
        mail_ok=mail_ok,
        mail_fail=mail_fail,
        top_fail_reasons=top_fail_reasons,
        mail_rows=mail_rows,
    )


@planner_bp.get("/api/reports/advanced")
@admin_required
def api_reports_advanced():
    start, end = _parse_date_range_args()
    status = (request.args.get("status") or "").strip().lower()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()
    city = (request.args.get("city") or "").strip()
    include_leave = _parse_bool_arg("include_leave", default=False)
    only_overdue = _parse_bool_arg("only_overdue", default=False)
    only_problem = _parse_bool_arg("only_problem", default=False)

    if project_id <= 0:
        subproject_id = 0
    elif subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id)):
        subproject_id = 0

    try:
        page = int(request.args.get("page", 1) or 1)
    except Exception:
        page = 1
    page = max(1, page)
    try:
        page_size = int(request.args.get("page_size", 50) or 50)
    except Exception:
        page_size = 50
    page_size = max(10, min(200, page_size))

    upsert_jobs_for_range(start, end)
    q = _build_advanced_jobs_query(start, end, status, project_id, subproject_id, team_name, city, only_overdue, only_problem)

    today = date.today()
    total_jobs = int(q.count() or 0)
    completed = int(q.filter(Job.status == "completed").count() or 0)
    problem = int(q.filter(Job.status == "problem").count() or 0)
    pending = max(0, total_jobs - completed - problem)
    overdue = int(q.filter(Job.status != "completed", Job.work_date < today).count() or 0)
    completion_rate = round((completed / total_jobs) * 100.0, 1) if total_jobs else 0.0

    total_pages = max(1, (total_jobs + page_size - 1) // page_size) if total_jobs else 1
    page = min(page, total_pages)

    jobs = (
        q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    job_ids = [j.id for j in jobs]
    people_count = {}
    if job_ids:
        if include_leave:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .filter(JobAssignment.job_id.in_(job_ids))
                .group_by(JobAssignment.job_id)
                .all()
            )
        else:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .join(Job, Job.id == JobAssignment.job_id)
                .outerjoin(
                    PersonDayStatus,
                    (PersonDayStatus.person_id == JobAssignment.person_id) & (PersonDayStatus.work_date == Job.work_date),
                )
                .filter(JobAssignment.job_id.in_(job_ids))
                .filter(or_(PersonDayStatus.id == None, PersonDayStatus.status != "leave"))
                .group_by(JobAssignment.job_id)
                .all()
            )
        for jid, cnt in rows:
            people_count[int(jid)] = int(cnt or 0)

    payload_jobs = []
    for j in jobs:
        p = j.project
        payload_jobs.append(
            {
                "id": j.id,
                "work_date": iso(j.work_date),
                "city": (p.region if p else ""),
                "project_id": int(j.project_id or 0),
                "project_code": (p.project_code if p else ""),
                "project_name": (p.project_name if p else ""),
                "team_name": j.team_name or "",
                "status": j.status,
                "people": people_count.get(j.id, 0),
                "shift": j.shift or "",
                "vehicle": j.vehicle_info or "",
                "closed_at": j.closed_at.isoformat() if getattr(j, "closed_at", None) else None,
            }
        )

    team_key = db.func.coalesce(db.func.nullif(db.func.trim(Job.team_name), ""), "-")
    top_teams = (
        q.with_entities(team_key, db.func.count(Job.id))
        .group_by(team_key)
        .order_by(db.func.count(Job.id).desc(), team_key.asc())
        .limit(15)
        .all()
    )
    city_key = db.func.coalesce(db.func.nullif(db.func.trim(Project.region), ""), "-")
    top_cities = (
        q.with_entities(city_key, db.func.count(Job.id))
        .group_by(city_key)
        .order_by(db.func.count(Job.id).desc(), city_key.asc())
        .limit(15)
        .all()
    )

    return jsonify(
        {
            "ok": True,
            "range": {"start": iso(start), "end": iso(end)},
            "filters": {
                "status": status,
                "project_id": project_id,
                "team_name": team_name,
                "city": city,
                "include_leave": include_leave,
                "only_overdue": only_overdue,
                "only_problem": only_problem,
            },
            "pagination": {"page": page, "page_size": page_size, "total_pages": total_pages, "total_count": total_jobs},
            "kpi": {
                "total_jobs": total_jobs,
                "pending": pending,
                "completed": completed,
                "problem": problem,
                "overdue": overdue,
                "completion_rate": completion_rate,
            },
            "top_teams": [{"team": n, "jobs": int(c)} for n, c in top_teams],
            "top_cities": [{"city": n, "jobs": int(c)} for n, c in top_cities],
            "jobs": payload_jobs,
        }
    )


@planner_bp.get("/reports/advanced.xlsx")
@admin_required
def reports_advanced_xlsx():
    start, end = _parse_date_range_args()
    status = (request.args.get("status") or "").strip().lower()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()
    city = (request.args.get("city") or "").strip()
    include_leave = _parse_bool_arg("include_leave", default=False)
    only_overdue = _parse_bool_arg("only_overdue", default=False)
    only_problem = _parse_bool_arg("only_problem", default=False)

    if project_id <= 0:
        subproject_id = 0
    elif subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id)):
        subproject_id = 0

    upsert_jobs_for_range(start, end)

    q = _build_advanced_jobs_query(start, end, status, project_id, subproject_id, team_name, city, only_overdue, only_problem)
    jobs = q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc()).all()

    today = date.today()
    total_jobs = len(jobs)
    completed = sum(1 for j in jobs if j.status == "completed")
    problem = sum(1 for j in jobs if j.status == "problem")
    pending = max(0, total_jobs - completed - problem)
    overdue = sum(1 for j in jobs if j.status != "completed" and j.work_date < today)
    completion_rate = round((completed / total_jobs) * 100.0, 1) if total_jobs else 0.0

    job_ids_subq = q.with_entities(Job.id).subquery()
    people_count = {}
    try:
        if include_leave:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .filter(JobAssignment.job_id.in_(db.session.query(job_ids_subq.c.id)))
                .group_by(JobAssignment.job_id)
                .all()
            )
        else:
            rows = (
                db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
                .join(Job, Job.id == JobAssignment.job_id)
                .outerjoin(
                    PersonDayStatus,
                    (PersonDayStatus.person_id == JobAssignment.person_id) & (PersonDayStatus.work_date == Job.work_date),
                )
                .filter(JobAssignment.job_id.in_(db.session.query(job_ids_subq.c.id)))
                .filter(or_(PersonDayStatus.id == None, PersonDayStatus.status != "leave"))
                .group_by(JobAssignment.job_id)
                .all()
            )
        for jid, cnt in rows:
            people_count[int(jid)] = int(cnt or 0)
    except Exception:
        people_count = {}

    # Mail summary for range
    mail_rows = MailLog.query.filter(
        MailLog.created_at >= datetime.combine(start, datetime.min.time()),
        MailLog.created_at <= datetime.combine(end, datetime.max.time())
    ).order_by(MailLog.created_at.desc()).all()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.append(["Start", iso(start)])
    ws1.append(["End", iso(end)])
    ws1.append([])
    ws1.append(["Total Jobs", total_jobs])
    ws1.append(["Beklemede", pending])
    ws1.append(["Tamamlandı", completed])
    ws1.append(["Sorun", problem])
    ws1.append(["Overdue", overdue])
    ws1.append(["Completion Rate (%)", completion_rate])
    ws1.append([])
    ws1.append(["Filters", ""])
    ws1.append(["Status", status or "(all)"])
    ws1.append(["City", city or "(all)"])
    ws1.append(["Project ID", project_id or 0])
    ws1.append(["Team", team_name or "(all)"])
    ws1.append(["Only Overdue", "1" if only_overdue else "0"])
    ws1.append(["Sadece Sorun", "1" if only_problem else "0"])
    ws1.append(["Include Leave", "1" if include_leave else "0"])

    ws2 = wb.create_sheet("Jobs")
    ws2.append(["Date", "City", "Project Code", "Project Name", "Team", "Status", "People", "Shift", "Vehicle"])
    for j in jobs:
        p = j.project
        ws2.append([
            iso(j.work_date),
            (p.region if p else ""),
            (p.project_code if p else ""),
            (p.project_name if p else ""),
            j.team_name or "",
            j.status,
            people_count.get(j.id, 0),
            j.shift or "",
            j.vehicle_info or "",
        ])

    ws3 = wb.create_sheet("Teams")
    team_counts = {}
    for j in jobs:
        key = (j.team_name or "").strip() or "-"
        team_counts[key] = team_counts.get(key, 0) + 1
    ws3.append(["Ekip", "İş", "Tamamlandı", "Sorun", "Geciken"])
    for name, cnt in sorted(team_counts.items(), key=lambda x: (-x[1], x[0])):
        t_completed = sum(1 for j in jobs if ((j.team_name or "").strip() or "-") == name and j.status == "completed")
        t_problem = sum(1 for j in jobs if ((j.team_name or "").strip() or "-") == name and j.status == "problem")
        t_overdue = sum(1 for j in jobs if ((j.team_name or "").strip() or "-") == name and j.status != "completed" and j.work_date < today)
        ws3.append([name, cnt, t_completed, t_problem, t_overdue])

    ws4 = wb.create_sheet("MailLogs")
    ws4.append(["Created At", "Kind", "OK", "To", "Subject", "Error"])
    for r in mail_rows[:2000]:
        ws4.append([
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.kind or "",
            "1" if r.ok else "0",
            r.to_addr or "",
            r.subject or "",
            (r.error or "")[:200],
        ])

    for wsht in [ws1, ws2, ws3, ws4]:
        for c in range(1, wsht.max_column + 1):
            wsht.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"advanced_reports_{iso(start)}_{iso(end)}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@planner_bp.get("/api/job_detail")
@login_required
def api_job_detail():
    job_id = int(request.args.get("job_id", 0) or 0)
    if not job_id:
        return jsonify({"ok": False, "error": "job_id eksik"}), 400
    j = Job.query.get(job_id)
    if not j:
        return jsonify({"ok": False, "error": "Job bulunamadi"}), 404

    project = Project.query.get(j.project_id)
    people = (
        db.session.query(Person.full_name, Person.phone, Person.email)
        .join(JobAssignment, JobAssignment.person_id == Person.id)
        .filter(JobAssignment.job_id == j.id)
        .order_by(Person.full_name.asc())
        .all()
    )
    fb = JobFeedback.query.filter_by(job_id=j.id).order_by(JobFeedback.closed_at.desc()).first()

    # related mail logs
    rel_logs = []
    try:
        import json as _json
        logs = MailLog.query.order_by(MailLog.created_at.desc()).limit(400).all()
        for r in logs:
            meta = {}
            try:
                meta = _json.loads(r.meta_json or "{}")
            except Exception:
                meta = {}
            if meta.get("type") != "job":
                continue
            if int(meta.get("project_id") or 0) != int(j.project_id):
                continue
            if str(meta.get("work_date") or "") != iso(j.work_date):
                continue
            rel_logs.append({
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "ok": bool(r.ok),
                "to": r.to_addr,
                "subject": r.subject,
                "error": (r.error or ""),
            })
            if len(rel_logs) >= 30:
                break
    except Exception:
        rel_logs = []

    return jsonify({
        "ok": True,
        "job": {
            "id": j.id,
            "date": iso(j.work_date),
            "status": j.status,
            "closed_at": j.closed_at.isoformat() if j.closed_at else None,
            "city": project.region if project else "",
            "project_code": project.project_code if project else "",
            "project_name": project.project_name if project else "",
            "team_name": j.team_name or "",
            "shift": j.shift or "",
            "vehicle": j.vehicle_info or "",
            "note": j.note or "",
        },
        "people": [{"full_name": n, "phone": ph or "", "email": em or ""} for (n, ph, em) in people],
        "feedback": None if not fb else {
            "status": fb.status,
            "note": fb.note or "",
            "closed_at": fb.closed_at.isoformat() if fb.closed_at else None
        },
        "mail_logs": rel_logs
    })


@planner_bp.post("/api/job_feedback")
@login_required
@observer_required
def api_job_feedback_set():
    data = request.get_json(force=True, silent=True) or {}
    job_id = int(data.get("job_id", 0) or 0)
    status = (data.get("status") or "").strip().lower()
    note = (data.get("note") or "").strip()
    if not job_id:
        return jsonify({"ok": False, "error": "job_id eksik"}), 400
    if status not in ("completed", "problem", "pending"):
        return jsonify({"ok": False, "error": "status gecersiz"}), 400
    j = Job.query.get(job_id)
    if not j:
        return jsonify({"ok": False, "error": "Job bulunamadi"}), 404

    if status == "pending":
        # clear feedback + reset job
        JobFeedback.query.filter_by(job_id=job_id).delete()
        j.status = "pending"
        j.closed_at = None
        db.session.commit()
        return jsonify({"ok": True})

    fb = JobFeedback(job_id=job_id, status=status, note=note or None, closed_at=datetime.now(), created_by_user_id=session.get("user_id"))
    db.session.add(fb)
    j.status = status
    j.closed_at = fb.closed_at
    db.session.commit()

    try:
        project = Project.query.get(j.project_id)
        body = "{} {} | {} | Ekip: {}".format((project.project_code if project else ""), (project.region if project else ""), iso(j.work_date), (j.team_name or "-"))
        _notify_admins(event="feedback", title="Geri bildirim geldi", body=body, link_url=url_for("assignment_page", job_id=j.id), job_id=j.id, meta={"status": status})
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

    return jsonify({"ok": True})


@planner_bp.get("/reports/mail-log")
@login_required
@admin_required
def mail_log_page():
    # Tarih filtreleri
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    
    if start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    else:
        # Varsayılan: son 30 gün
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    
    # Diğer filtreler
    team_name = (request.args.get("team_name", request.args.get("team", "") or "").strip())
    ok_filter = (request.args.get("ok", "").strip())
    mail_type_filter = (request.args.get("mail_type", "").strip())
    q_filter = (request.args.get("q", "").strip())
    
    # Sayfalama
    page = int(request.args.get("page", 1) or 1)
    per_page = 100
    
    q = MailLog.query
    
    # Tarih filtresi
    if start_date and end_date:
        q = q.filter(MailLog.created_at >= datetime.combine(start_date, datetime.min.time()))
        q = q.filter(MailLog.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    if team_name:
        q = q.filter(MailLog.team_name == team_name)
    
    if ok_filter in ("0", "1"):
        q = q.filter(MailLog.ok == (ok_filter == "1"))
    
    if mail_type_filter:
        q = q.filter(MailLog.mail_type == mail_type_filter)
    
    if q_filter:
        # Arama: to_addr veya subject'te ara
        q = q.filter(
            db.or_(
                MailLog.to_addr.ilike(f"%{q_filter}%"),
                MailLog.subject.ilike(f"%{q_filter}%"),
                MailLog.cc_addr.ilike(f"%{q_filter}%")
            )
        )
    
    # Toplam sayı
    total_count = q.count()
    total_pages = (total_count + per_page - 1) // per_page
    
    # Sayfalama uygula
    rows = q.order_by(MailLog.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    
    return render_template(
        "mail_log.html",
        rows=rows,
        start_date=start_date.isoformat() if start_date else "",
        end_date=end_date.isoformat() if end_date else "",
        team_name=team_name,
        ok_filter=ok_filter,
        mail_type_filter=mail_type_filter,
        q_filter=q_filter,
        page=page,
        total_pages=total_pages,
        total_count=total_count
    )


@planner_bp.get("/reports/mail-log.xlsx")
@login_required
@admin_required
def mail_log_excel():
    # Tarih filtreleri
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    
    if start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    else:
        # Varsayılan: son 30 gün
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    
    team_name = (request.args.get("team_name", request.args.get("team", "") or "").strip())
    ok_filter = (request.args.get("ok", "").strip())
    mail_type_filter = (request.args.get("mail_type", "").strip())
    q_filter = (request.args.get("q", "").strip())

    q = MailLog.query
    
    # Tarih filtresi
    if start_date and end_date:
        q = q.filter(MailLog.created_at >= datetime.combine(start_date, datetime.min.time()))
        q = q.filter(MailLog.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    if team_name:
        q = q.filter(MailLog.team_name == team_name)
    
    if ok_filter in ("0", "1"):
        q = q.filter(MailLog.ok == (ok_filter == "1"))
    
    if mail_type_filter:
        q = q.filter(MailLog.mail_type == mail_type_filter)
    
    if q_filter:
        q = q.filter(
            db.or_(
                MailLog.to_addr.ilike(f"%{q_filter}%"),
                MailLog.subject.ilike(f"%{q_filter}%"),
                MailLog.cc_addr.ilike(f"%{q_filter}%")
            )
        )
    
    rows = q.order_by(MailLog.created_at.desc()).all()

    wb = Workbook()
    sh = wb.active
    sh.title = "MailLog"
    sh.append(["Tarih", "Tip", "Durum", "To", "CC", "Konu", "Hafta", "Ekip", "Boyut", "Hata"])
    for r in rows:
        sh.append([
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.mail_type or r.kind or "",
            "OK" if r.ok else "ERR",
            r.to_addr or "",
            r.cc_addr or "",
            r.subject or "",
            r.week_start.strftime("%Y-%m-%d") if r.week_start else "",
            r.team_name or "",
            (r.body_size_bytes or 0),
            (r.error or "")[:500],
        ])
    for c in range(1, 11):
        sh.cell(row=1, column=c).font = Font(bold=True)
    sh.column_dimensions["A"].width = 18
    sh.column_dimensions["B"].width = 12
    sh.column_dimensions["C"].width = 8
    sh.column_dimensions["D"].width = 35
    sh.column_dimensions["E"].width = 35
    sh.column_dimensions["F"].width = 50
    sh.column_dimensions["G"].width = 12
    sh.column_dimensions["H"].width = 18
    sh.column_dimensions["I"].width = 10
    sh.column_dimensions["J"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"mail_log_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@planner_bp.get("/api/mail/log/detail/<int:log_id>")
@login_required
@admin_required
def api_mail_log_detail(log_id: int):
    """Mail log detayını getir"""
    log = MailLog.query.get_or_404(log_id)
    
    # Meta bilgisini parse et
    meta = {}
    if log.meta_json:
        try:
            import json
            meta = json.loads(log.meta_json)
        except:
            pass
    
    return jsonify({
        "ok": True,
        "log": {
            "id": log.id,
            "created_at": log.created_at.isoformat() if log.created_at else None,
            "mail_type": log.mail_type,
            "kind": log.kind,
            "ok": log.ok,
            "error_code": log.error_code,
            "error": log.error,
            "to_addr": log.to_addr,
            "cc_addr": log.cc_addr,
            "bcc_addr": log.bcc_addr,
            "subject": log.subject,
            "body_preview": log.body_preview,
            "body_html": log.body_html,  # Tam HTML içerik - mail önizleme için
            "week_start": log.week_start.isoformat() if log.week_start else None,
            "team_name": log.team_name,
            "project_id": log.project_id,
            "job_id": log.job_id,
            "task_id": log.task_id,
            "user_id": log.user_id,
            "attachments_count": log.attachments_count,
            "body_size_bytes": log.body_size_bytes,
            "sent_at": log.sent_at.isoformat() if log.sent_at else None,
            "meta": meta,
        }
    })


@planner_bp.post("/api/mail/log/resend/<int:log_id>")
@login_required
@admin_required
def api_mail_log_resend(log_id: int):
    """Mail logunu yeniden gönder (düzenlenmiş içerikle)"""
    data = request.get_json(force=True, silent=True) or {}
    log = MailLog.query.get_or_404(log_id)
    
    # Yeni alıcı/konu bilgisi
    new_to = data.get("to_addr", log.to_addr)
    new_subject = data.get("subject", log.subject)
    new_cc = data.get("cc_addr", log.cc_addr)
    
    try:
        from services.mail_service import MailService
        
        # Meta bilgisini al
        meta = {}
        if log.meta_json:
            import json
            try:
                meta = json.loads(log.meta_json)
            except:
                pass
        
        # Basit HTML body oluştur (orijinal body_preview'dan)
        body_html = f"""<!DOCTYPE html>
<html><body style="font-family: Arial, sans-serif; line-height: 1.6; padding: 20px;">
<h2>{new_subject}</h2>
<div style="white-space: pre-wrap;">{log.body_preview or ''}</div>
<hr>
<p style="color: #666; font-size: 12px;">Bu mail yeniden gönderildi. Orijinal gönderim: {log.created_at}</p>
</body></html>"""
        
        ok = MailService.send(
            mail_type=log.mail_type or "resend",
            recipients=new_to,
            subject=new_subject,
            html=body_html,
            cc=new_cc,
            user_id=session.get("user_id"),
            meta={
                **meta,
                "resend_of": log.id,
                "original_subject": log.subject,
            }
        )
        
        if ok:
            return jsonify({"ok": True, "message": "Mail başarıyla yeniden gönderildi"})
        else:
            return jsonify({"ok": False, "error": "Mail gönderimi başarısız"}), 500
            
    except Exception as e:
        log.exception("Mail yeniden gönderme hatası")
        return jsonify({"ok": False, "error": str(e)}), 500


@planner_bp.post("/api/mail/log/delete/<int:log_id>")
@login_required
@admin_required
def api_mail_log_delete(log_id: int):
    """Mail logunu sil"""
    log = MailLog.query.get_or_404(log_id)
    try:
        db.session.delete(log)
        db.session.commit()
        return jsonify({"ok": True, "message": "Log silindi"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@planner_bp.post("/api/mail/log/bulk-resend")
@login_required
@admin_required
def api_mail_log_bulk_resend():
    """Birden fazla başarısız maili yeniden gönder"""
    data = request.get_json(force=True, silent=True) or {}
    log_ids = data.get("log_ids", [])
    
    if not log_ids:
        return jsonify({"ok": False, "error": "Log IDleri belirtilmedi"}), 400
    
    logs = MailLog.query.filter(MailLog.id.in_(log_ids)).all()
    results = []
    
    for log in logs:
        try:
            from services.mail_service import MailService
            
            meta = {}
            if log.meta_json:
                import json
                try:
                    meta = json.loads(log.meta_json)
                except:
                    pass
            
            body_html = f"""<!DOCTYPE html>
<html><body style="font-family: Arial, sans-serif; line-height: 1.6; padding: 20px;">
<h2>{log.subject}</h2>
<div style="white-space: pre-wrap;">{log.body_preview or ''}</div>
<hr>
<p style="color: #666; font-size: 12px;">Toplu yeniden gönderim. Orijinal: {log.created_at}</p>
</body></html>"""
            
            ok = MailService.send(
                mail_type=log.mail_type or "resend",
                recipients=log.to_addr,
                subject=log.subject,
                html=body_html,
                cc=log.cc_addr,
                user_id=session.get("user_id"),
                meta={
                    **meta,
                    "bulk_resend_of": log.id,
                }
            )
            
            results.append({
                "log_id": log.id,
                "to": log.to_addr,
                "ok": ok,
                "subject": log.subject
            })
            
        except Exception as e:
            results.append({
                "log_id": log.id,
                "to": log.to_addr,
                "ok": False,
                "error": str(e),
                "subject": log.subject
            })
    
    success_count = sum(1 for r in results if r.get("ok"))
    return jsonify({
        "ok": True,
        "results": results,
        "success_count": success_count,
        "total_count": len(results)
    })


@planner_bp.get("/mail")
@login_required
def mail_page():
    d = parse_date(request.args.get("date", "")) or date.today()
    ws = parse_date(request.args.get("week_start", "")) or week_start(d)
    start = ws
    end = ws + timedelta(days=6)

    team_names = [
        r[0] for r in db.session.query(Team.name).join(PlanCell, PlanCell.team_id == Team.id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end, Team.name != None).distinct().all()
    ]
    extra = [
        r[0] for r in db.session.query(PlanCell.team_name).filter(
            PlanCell.work_date >= start, PlanCell.work_date <= end,
            PlanCell.team_name != None, PlanCell.team_name != ""
        ).distinct().all()
    ]
    for t in extra:
        if t not in team_names:
            team_names.append(t)

    return render_template(
        "mail.html",
        selected_week=iso(ws),
        prev_week=iso(ws - timedelta(days=7)),
        next_week=iso(ws + timedelta(days=7)),
        team_names=team_names
    )


@planner_bp.get("/mail/compose")
@login_required
@planner_or_admin_required
def mail_compose_page():
    start, end = _parse_date_range_args()
    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()
    cities = sorted({(p.region or "").strip() for p in projects if (p.region or "").strip()})
    teams = Team.query.order_by(Team.name.asc()).all()
    return render_template(
        "mail_compose.html",
        start=iso(start),
        end=iso(end),
        projects=projects,
        cities=cities,
        teams=teams,
    )


@planner_bp.get("/api/teams")
@login_required
@planner_or_admin_required
def api_teams():
    import json as _json

    ws = parse_date(str(request.args.get("week_start") or "").strip())
    if ws:
        start = week_start(ws)
        end = start + timedelta(days=6)
        team_ids = [
            int(x[0]) for x in db.session.query(PlanCell.team_id)
            .filter(PlanCell.team_id.isnot(None), PlanCell.work_date >= start, PlanCell.work_date <= end)
            .distinct()
            .all()
            if x and x[0]
        ]
        if team_ids:
            teams = Team.query.filter(Team.id.in_(team_ids)).order_by(Team.name.asc()).all()
        else:
            teams = []
    else:
        teams = Team.query.order_by(Team.name.asc()).all()
    cfg_rows = TeamMailConfig.query.all()
    cfg_by_team = {int(r.team_id): r for r in cfg_rows if r and r.team_id}

    out = []
    for t in teams:
        cfg = cfg_by_team.get(int(t.id))
        emails = []
        active = True
        if cfg:
            active = bool(cfg.active)
            try:
                emails = _json.loads(cfg.emails_json or "[]") or []
            except Exception:
                emails = []
        emails = [e.strip() for e in emails if isinstance(e, str) and e.strip()]
        member_ids = [int(x) for x in (t.signature or "").split(",") if str(x or "").strip().isdigit()]
        out.append({
            "id": int(t.id),
            "name": t.name,
            "active": active,
            "emails": emails,
            "member_count": len(member_ids),
            "color": team_color(int(t.id)),
        })
    return jsonify({"ok": True, "teams": out})


@planner_bp.get("/api/team/<int:team_id>/members")
@login_required
@planner_or_admin_required
def api_team_members_by_id(team_id: int):
    t = Team.query.get_or_404(team_id)
    try:
        ids = [int(x) for x in (t.signature or "").split(",") if str(x or "").strip().isdigit()]
    except Exception:
        ids = []
    rows = Person.query.filter(Person.id.in_(ids)).all() if ids else []
    by_id = {int(p.id): p for p in rows if p and p.id}
    members = []
    for pid in ids:
        p = by_id.get(int(pid))
        if not p:
            continue
        members.append({
            "id": int(p.id),
            "full_name": p.full_name,
            "phone": p.phone or "",
            "email": p.email or "",
        })
    return jsonify({"ok": True, "team_id": int(t.id), "team_name": t.name, "members": members})


@planner_bp.post("/api/team/<int:team_id>/vehicle")
@login_required
@planner_or_admin_required
def api_team_set_vehicle(team_id: int):
    data = request.get_json(force=True, silent=True) or {}
    vehicle_id = int(data.get("vehicle_id", 0) or 0)
    team = Team.query.get_or_404(team_id)
    vehicle = None
    if vehicle_id:
        vehicle = Vehicle.query.filter_by(id=vehicle_id).first()
        if not vehicle:
            return jsonify({"ok": False, "error": "Araç bulunamadı"}), 404
        conflict = Team.query.filter(Team.vehicle_id == vehicle_id, Team.id != team.id).first()
        if conflict:
            # Aynı aracı isteyen ekip varsa önce ondan kaldır
            conflict.vehicle_id = None
            apply_team_vehicle_to_cells(conflict, None)
    team.vehicle_id = vehicle.id if vehicle else None
    apply_team_vehicle_to_cells(team, vehicle)
    return jsonify({"ok": True, "team_id": int(team.id), "vehicle": _vehicle_payload(vehicle, team.id if vehicle else None)})


@planner_bp.post("/api/teams/save_emails")
@login_required
@planner_or_admin_required
def api_teams_save_emails():
    import json as _json

    data = request.get_json(force=True, silent=True) or {}
    team_id = int(data.get("team_id", 0) or 0)
    if not team_id:
        return jsonify({"ok": False, "error": "team_id eksik"}), 400
    team = Team.query.get(team_id)
    if not team:
        return jsonify({"ok": False, "error": "Ekip bulunamadi"}), 404

    emails_in = data.get("emails")
    active = bool(data.get("active", True))
    emails = []
    if isinstance(emails_in, list):
        emails = [str(x or "").strip() for x in emails_in]
    else:
        emails = _split_email_list(str(emails_in or ""))

    cleaned = []
    seen = set()
    for e in emails:
        e = (e or "").strip()
        if not e or "@" not in e:
            continue
        key = e.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(e)

    row = TeamMailConfig.query.filter_by(team_id=team_id).first()
    if not row:
        row = TeamMailConfig(team_id=team_id, emails_json="[]", active=True)
        db.session.add(row)
    row.active = active
    row.emails_json = _json.dumps(cleaned, ensure_ascii=False)
    db.session.commit()
    return jsonify({"ok": True, "team_id": team_id, "emails": cleaned, "active": active})


@planner_bp.get("/api/jobs")
@login_required
@planner_or_admin_required
def api_jobs_for_mail():
    start, end = _parse_date_range_args()
    status = (request.args.get("status") or "").strip().lower()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    team_name = (request.args.get("team_name") or "").strip()
    city = (request.args.get("city") or "").strip()
    only_overdue = _parse_bool_arg("only_overdue", default=False)
    only_problem = _parse_bool_arg("only_problem", default=False)
    q_text = (request.args.get("q") or "").strip()

    try:
        limit = int(request.args.get("limit", 200) or 200)
    except Exception:
        limit = 200
    limit = max(1, min(500, limit))

    if project_id <= 0:
        subproject_id = 0
    elif subproject_id and (not _subproject_allowed_for_project(subproject_id=subproject_id, project_id=project_id)):
        subproject_id = 0

    upsert_jobs_for_range(start, end)
    q = _build_advanced_jobs_query(start, end, status, project_id, subproject_id, team_name, city, only_overdue, only_problem)
    if q_text:
        like = f"%{q_text}%"
        q = q.filter(or_(Project.project_code.like(like), Project.project_name.like(like)))

    jobs = q.order_by(Job.work_date.desc(), Project.region.asc(), Project.project_code.asc()).limit(limit).all()
    job_ids = [j.id for j in jobs]

    people_count = {}
    if job_ids:
        for jid, cnt in (
            db.session.query(JobAssignment.job_id, db.func.count(db.func.distinct(JobAssignment.person_id)))
            .filter(JobAssignment.job_id.in_(job_ids))
            .group_by(JobAssignment.job_id)
            .all()
        ):
            people_count[int(jid)] = int(cnt or 0)

    out = []
    for j in jobs:
        p = j.project
        out.append({
            "id": int(j.id),
            "work_date": iso(j.work_date),
            "city": (p.region if p else ""),
            "project_id": int(j.project_id or 0),
            "project_code": (p.project_code if p else ""),
            "project_name": (p.project_name if p else ""),
            "team_id": int(j.team_id or 0) if getattr(j, "team_id", None) else 0,
            "team_name": (j.team_name or (j.team.name if j.team else "") or ""),
            "status": j.status,
            "people": people_count.get(int(j.id), 0),
            "shift": j.shift or "",
            "vehicle": j.vehicle_info or "",
        })
    return jsonify({"ok": True, "range": {"start": iso(start), "end": iso(end)}, "jobs": out})


# ---------- FIELD PORTAL (/me) ----------
def _normalize_me_tab(tab: str) -> str:
    t = (tab or "").strip().lower()
    if t in ("current", "completed", "waiting", "reported", "report_pending"):
        return t
    return "current"


def _portal_home_context(kind: str) -> dict:
    user = get_current_user()
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()

    if start_s and end_s:
        start = parse_date(start_s) or date.today()
        end = parse_date(end_s) or start
    else:
        start = date.today()
        end = start + timedelta(days=7)

    project_id = int(request.args.get("project_id", 0) or 0)
    city = (request.args.get("city") or "").strip()
    status = (request.args.get("status") or "").strip().lower()

    tab = _normalize_me_tab(kind)

    # Bu kullanıcıya atanmış işleri listele: yayınlanmış işler + hücrede atanmış (yayınlanmamış) işler
    q = (
        Job.query
        .join(Project, Project.id == Job.project_id)
        .outerjoin(PlanCell, Job.cell_id == PlanCell.id)
        .filter(
            db.or_(
                db.and_(Job.assigned_user_id == user.id, Job.is_published == True),
                PlanCell.assigned_user_id == user.id,
            ),
            Job.work_date >= start,
            Job.work_date <= end,
        )
    )
    if project_id:
        q = q.filter(Job.project_id == project_id)
    if city:
        q = q.filter(Project.region == city)
    if status:
        q = q.filter(Job.status == _job_status_label(status))

    # Tab base filters (can be combined with dropdown filters)
    if tab == "current":
        q = q.filter(Job.status != "completed")
    elif tab == "completed":
        q = q.filter(Job.status == "completed")
    elif tab == "waiting":
        q = q.filter(Job.status == "pending")
    elif tab == "report_pending":
        # Optimization: report pending is always completed + no feedback
        q = q.filter(Job.status == "completed")

    jobs = q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc()).all()

    # Job.assigned_user_id bos ama hücrede atama varsa veritabanını güncelle (bir sonraki yüklemede düzgün görünsün)
    try:
        for j in jobs:
            if getattr(j, "assigned_user_id", None) is None and getattr(j, "cell_id", None):
                cell = PlanCell.query.get(j.cell_id)
                if cell and getattr(cell, "assigned_user_id", None):
                    j.assigned_user_id = cell.assigned_user_id
                    db.session.add(j)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

    # Report flags (for tab filtering + card coloring)
    job_ids = [int(j.id) for j in jobs if j and getattr(j, "id", None)]
    has_report_job_ids = set()
    if job_ids:
        try:
            # "Rapor yazdım" tanımı: bu kullanıcıya ait herhangi bir JobFeedback kaydı varsa.
            for (jid,) in (
                db.session.query(JobFeedback.job_id)
                .filter(JobFeedback.job_id.in_(job_ids), JobFeedback.user_id == user.id)
                .distinct()
                .all()
            ):
                if jid:
                    has_report_job_ids.add(int(jid))
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

    job_meta = {}
    for j in jobs:
        jid = int(j.id)
        has_report = jid in has_report_job_ids
        report_pending = (str(getattr(j, "status", "") or "").strip().lower() == "completed") and (not has_report)
        job_meta[jid] = {"has_report": bool(has_report), "report_pending": bool(report_pending)}

    # Tabs that depend on report flags are applied after fetching jobs
    if tab == "reported":
        jobs = [j for j in jobs if job_meta.get(int(j.id), {}).get("has_report")]
    elif tab == "report_pending":
        jobs = [j for j in jobs if job_meta.get(int(j.id), {}).get("report_pending")]

    today = date.today()
    today_jobs = [j for j in jobs if j.work_date == today]
    upcoming_jobs = [j for j in jobs if j.work_date != today]

    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()
    cities = sorted({(p.region or "").strip() for p in projects if (p.region or "").strip()})

    announcements_preview = []
    announcements_unread = 0
    try:
        announcements_preview = _fetch_announcements(user, limit=5)
        announcements_unread = sum(1 for a in announcements_preview if not a.get("is_read"))
    except Exception:
        announcements_preview = []
        announcements_unread = 0

    return {
        "user": user,
        "start": iso(start),
        "end": iso(end),
        "active_tab": tab,
        "project_id": project_id,
        "city": city,
        "status": status,
        "projects": projects,
        "cities": cities,
        "job_meta": job_meta,
        "announcements_preview": announcements_preview,
        "announcements_unread_count": announcements_unread,
        "today_jobs": today_jobs,
        "upcoming_jobs": upcoming_jobs,
    }


@planner_bp.get("/me")
@login_required
@field_required
def portal_home():
    preset = (request.args.get("preset") or "").strip().lower()
    preset = preset.replace("-", "_")
    if preset:
        tab = _normalize_me_tab(preset or "current")

        args = request.args.to_dict(flat=True) if request.args else {}
        args.pop("preset", None)

        endpoint_map = {
            "current": "planner.portal_home_current",
            "completed": "planner.portal_home_completed",
            "waiting": "planner.portal_home_waiting",
            "reported": "planner.portal_home_reported",
            "report_pending": "planner.portal_home_report_pending",
        }
        return redirect(url_for(endpoint_map.get(tab, "planner.portal_home_current"), **args))

    ctx = _portal_home_context("current")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/current")
@login_required
@field_required
def portal_home_current():
    ctx = _portal_home_context("current")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/completed")
@login_required
@field_required
def portal_home_completed():
    ctx = _portal_home_context("completed")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/waiting")
@login_required
@field_required
def portal_home_waiting():
    ctx = _portal_home_context("waiting")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/reported")
@login_required
@field_required
def portal_home_reported():
    ctx = _portal_home_context("reported")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/report-pending")
@login_required
@field_required
def portal_home_report_pending():
    ctx = _portal_home_context("report_pending")
    return render_template("portal_home.html", **ctx)


@planner_bp.get("/me/where")
@login_required
@field_required
def portal_where():
    """
    Kim Nerede: seçilen güne göre ekiplerin sahadaki planını listeler.
    """
    user = get_current_user()
    d = parse_date((request.args.get("date") or "").strip()) or date.today()

    q = (
        Job.query
        .join(Project, Project.id == Job.project_id)
        .filter(
            Job.is_published == True,
            Job.work_date == d,
        )
    )
    jobs = q.order_by(Project.region.asc(), Job.team_name.asc().nullslast(), Project.project_code.asc()).all()

    job_ids = [int(j.id) for j in jobs if j and getattr(j, "id", None)]
    people_by_job: Dict[int, list] = {}
    if job_ids:
        for jid, full_name in (
            db.session.query(JobAssignment.job_id, Person.full_name)
            .join(Person, Person.id == JobAssignment.person_id)
            .filter(JobAssignment.job_id.in_(job_ids))
            .order_by(Person.full_name.asc())
            .all()
        ):
            if jid:
                people_by_job.setdefault(int(jid), []).append(full_name)

    now = datetime.now()
    cutoff = now - ONLINE_WINDOW
    online_user_ids = set()
    online_team_ids = set()
    try:
        online_user_ids = {
            int(uid) for (uid,) in db.session.query(User.id).filter(User.is_active == True, User.last_seen != None, User.last_seen >= cutoff).all()
            if uid
        }
        online_team_ids = {
            int(tid) for (tid,) in db.session.query(User.team_id).filter(User.is_active == True, User.team_id != None, User.last_seen != None, User.last_seen >= cutoff).distinct().all()
            if tid
        }
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

    return render_template(
        "portal_where.html",
        user=user,
        selected_date=iso(d),
        jobs=jobs,
        people_by_job=people_by_job,
        online_user_ids=online_user_ids,
        online_team_ids=online_team_ids,
    )


@planner_bp.get("/me/job/<int:job_id>")
@login_required
@field_required
def portal_job_detail(job_id: int):
    user = get_current_user()
    job = Job.query.get_or_404(job_id)
    cell = PlanCell.query.get(job.cell_id) if job.cell_id else None
    # Erişim: Job'da veya hücrede bu kullanıcıya atanmışsa detay açılsın (yayınlanmamış olsa da)
    is_assigned = (int(getattr(job, "assigned_user_id", 0) or 0) == int(user.id)) or (
        cell and int(getattr(cell, "assigned_user_id", 0) or 0) == int(user.id)
    )
    if not is_assigned:
        flash("Bu işe erişiminiz yok.", "danger")
        return redirect(url_for('planner.portal_home'))

    project = Project.query.get(job.project_id)

    # Planda (genel plan) yüklenen LLD/HHD ve tutanaklar; eski tek-dosya alanları da dahil
    lld_files = []
    tutanak_files = []
    if cell:
        lld_files = _parse_files(getattr(cell, "lld_hhd_files", None)) or (
            [cell.lld_hhd_path] if getattr(cell, "lld_hhd_path", None) else []
        )
        tutanak_files = _parse_files(getattr(cell, "tutanak_files", None)) or (
            [cell.tutanak_path] if getattr(cell, "tutanak_path", None) else []
        )
    photo_files = _parse_files(getattr(cell, "photo_files", None)) if cell else []

    # Personel: hücredeki atamalardan
    personnel = []
    if cell:
        personnel = (
            db.session.query(Person.full_name, Person.phone)
            .join(CellAssignment, CellAssignment.person_id == Person.id)
            .filter(CellAssignment.cell_id == cell.id)
            .order_by(Person.full_name.asc())
            .all()
        )
    personnel = [{"full_name": (n or "").strip(), "phone": (ph or "").strip()} for (n, ph) in personnel]

    # Status history & mail history
    status_history = (
        JobStatusHistory.query.filter_by(job_id=job.id).order_by(JobStatusHistory.changed_at.desc()).limit(20).all()
    )
    mail_history = (
        MailLog.query.filter(MailLog.job_id == job.id, MailLog.kind == "send")
        .order_by(MailLog.created_at.desc())
        .limit(20)
        .all()
    )
    last_update = (cell.updated_at if cell and getattr(cell, "updated_at", None) else None) or getattr(job, "updated_at", None)
    qc_result = getattr(cell, "qc_result", None) or ""

    return render_template(
        "portal_job_detail.html",
        user=user,
        job=job,
        project=project,
        cell=cell,
        lld_files=lld_files,
        tutanak_files=tutanak_files,
        photo_files=photo_files,
        personnel=personnel,
        status_history=status_history,
        mail_history=mail_history,
        last_update=last_update,
        qc_result=qc_result,
        kanban_columns=KANBAN_COLUMNS,
        kanban_labels=KANBAN_LABEL_TR,
    )


def _portal_job_cell(job_id: int, user):
    """Saha kullanıcısı için job ve cell döner; yetkisi yoksa (None, None)."""
    job = Job.query.get(job_id)
    if not job:
        return None, None
    cell = PlanCell.query.get(job.cell_id) if job.cell_id else None
    is_assigned = (int(getattr(job, "assigned_user_id", 0) or 0) == int(user.id)) or (
        cell and int(getattr(cell, "assigned_user_id", 0) or 0) == int(user.id)
    )
    if not is_assigned:
        return None, None
    return job, cell


@planner_bp.post("/me/job/<int:job_id>/upload/lld")
@login_required
@field_required
def portal_job_upload_lld(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job or not cell:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    if not _csrf_verify(request.form.get("csrf_token", "")):
        return jsonify({"ok": False, "error": "csrf"}), 400
    prefix = f"{job.project_id}-{iso(job.work_date)}"
    lld_files = request.files.getlist("lld_hhd")
    cur = _parse_files(getattr(cell, "lld_hhd_files", None))
    for fs in lld_files:
        if fs and fs.filename:
            cur.append(save_uploaded_file(fs, f"lldhhd-{prefix}"))
    cell.lld_hhd_files = _dump_files(cur) if cur else None
    cell.updated_at = datetime.now()
    db.session.commit()
    return jsonify({"ok": True, "files": cur})


@planner_bp.post("/me/job/<int:job_id>/upload/tutanak")
@login_required
@field_required
def portal_job_upload_tutanak(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job or not cell:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    if not _csrf_verify(request.form.get("csrf_token", "")):
        return jsonify({"ok": False, "error": "csrf"}), 400
    prefix = f"{job.project_id}-{iso(job.work_date)}"
    tutanak_files = request.files.getlist("tutanak")
    cur = _parse_files(getattr(cell, "tutanak_files", None))
    for fs in tutanak_files:
        if fs and fs.filename:
            cur.append(save_uploaded_file(fs, f"tutanak-{prefix}"))
    cell.tutanak_files = _dump_files(cur) if cur else None
    cell.updated_at = datetime.now()
    db.session.commit()
    return jsonify({"ok": True, "files": cur})


@planner_bp.post("/me/job/<int:job_id>/upload/photos")
@login_required
@field_required
def portal_job_upload_photos(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job or not cell:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    if not _csrf_verify(request.form.get("csrf_token", "")):
        return jsonify({"ok": False, "error": "csrf"}), 400
    prefix = f"{job.project_id}-{iso(job.work_date)}"
    photos = request.files.getlist("photos")
    cur = _parse_files(getattr(cell, "photo_files", None))
    for fs in photos:
        if fs and fs.filename:
            cur.append(save_uploaded_file(fs, f"photo-{prefix}"))
    cell.photo_files = _dump_files(cur) if cur else None
    cell.updated_at = datetime.now()
    db.session.commit()
    return jsonify({"ok": True, "files": cur})


@planner_bp.post("/me/job/<int:job_id>/qc")
@login_required
@field_required
def portal_job_save_qc(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job or not cell:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json(force=True, silent=True) or request.form
    if not _csrf_verify((data.get("csrf_token") or "").strip()):
        return jsonify({"ok": False, "error": "csrf"}), 400
    qc = (data.get("qc_result") or "").strip()
    cell.qc_result = qc or None
    cell.updated_at = datetime.now()
    db.session.commit()
    return jsonify({"ok": True})


@planner_bp.post("/me/job/<int:job_id>/status")
@login_required
@field_required
def portal_job_save_status(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json(force=True, silent=True) or request.form
    if not _csrf_verify((data.get("csrf_token") or "").strip()):
        return jsonify({"ok": False, "error": "csrf"}), 400
    new_status = _normalize_kanban_status((data.get("kanban_status") or "").strip())
    if new_status not in KANBAN_COLUMNS:
        return jsonify({"ok": False, "error": "invalid_status"}), 400
    changed = _set_job_kanban_status(job, new_status, changed_by_user_id=user.id, note="portal_set")
    if changed:
        db.session.commit()
    return jsonify({"ok": True, "kanban_status": new_status})


@planner_bp.post("/me/job/<int:job_id>/feedback")
@login_required
@field_required
def portal_job_feedback(job_id: int):
    user = get_current_user()
    job, cell = _portal_job_cell(job_id, user)
    if not job:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    # Form (multipart) veya JSON: outcome + note + opsiyonel media_files
    if request.content_type and "multipart/form-data" in request.content_type:
        data = request.form
        media_files = request.files.getlist("media_files") if request.files else []
    else:
        data = request.get_json(force=True, silent=True) or {}
        media_files = []
    if not _csrf_verify((data.get("csrf_token") or "").strip()):
        return jsonify({"ok": False, "error": "csrf"}), 400
    outcome = (data.get("outcome") or "").strip().lower()
    note = (data.get("note") or "").strip()
    # Eski API: sadece status gönderilmişse (completed/problem/pending)
    if not outcome and data.get("status"):
        status = (data.get("status") or "").strip().lower()
        if status == "pending":
            JobFeedback.query.filter_by(job_id=job_id).delete()
            job.status = "pending"
            job.closed_at = None
            db.session.commit()
            return jsonify({"ok": True})
        outcome = "issue" if status == "problem" else "completed"
    if outcome not in ("completed", "not_completed", "issue"):
        return jsonify({"ok": False, "error": "invalid_status"}), 400
    now = datetime.now()
    fb = JobFeedback(
        job_id=job_id,
        status=("problem" if outcome == "issue" else "completed" if outcome == "completed" else "pending"),
        note=note or None,
        closed_at=now,
        created_by_user_id=user.id,
        user_id=user.id,
        submitted_at=now,
        outcome=outcome,
        notes_text=note or None,
        created_at=now,
    )
    db.session.add(fb)
    db.session.flush()
    try:
        from utils import _save_feedback_uploads
        for m in _save_feedback_uploads(media_files, feedback_id=int(fb.id)):
            db.session.add(m)
    except ValueError as ve:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": "Dosya yükleme hatası."}), 400
    job.status = fb.status
    job.closed_at = fb.closed_at
    try:
        _promote_job_kanban_status(job, "REPORTED", changed_by_user_id=user.id, note="field_feedback")
    except Exception:
        pass
    db.session.commit()
    return jsonify({"ok": True})


@planner_bp.post("/api/jobs/<int:job_id>/reschedule")
@login_required
def api_job_reschedule(job_id: int):
    """
    Copy/move a published job to a future (or any) date by cloning/moving its underlying PlanCell.

    body: { target_date: 'YYYY-MM-DD', mode: 'copy'|'move', mark_old_pending?: bool, csrf_token }
    """
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "auth"}), 401

    data = request.get_json(force=True, silent=True) or {}
    if not _csrf_verify((data.get("csrf_token") or "").strip()):
        return jsonify({"ok": False, "error": "csrf"}), 400

    job = Job.query.get_or_404(job_id)
    if not (_user_is_admin_or_planner(user) or int(getattr(job, "assigned_user_id", 0) or 0) == int(user.id)):
        return jsonify({"ok": False, "error": "forbidden"}), 403

    from_date = getattr(job, "work_date", None)

    mode = (data.get("mode") or "copy").strip().lower()
    if mode not in {"copy", "move"}:
        return jsonify({"ok": False, "error": "mode_invalid"}), 400

    target_date = parse_date((data.get("target_date") or "").strip())
    if not target_date:
        return jsonify({"ok": False, "error": "target_date_invalid"}), 400

    if target_date == getattr(job, "work_date", None):
        return jsonify({"ok": False, "error": "target_date_same"}), 400

    if mode == "move" and (str(getattr(job, "status", "") or "").strip().lower() == "completed"):
        return jsonify({"ok": False, "error": "completed_cannot_move"}), 400

    mark_old_pending = bool(data.get("mark_old_pending", False))

    src_cell = PlanCell.query.get(job.cell_id) if getattr(job, "cell_id", None) else None
    if not src_cell:
        return jsonify({"ok": False, "error": "cell_missing"}), 400

    # Ensure a target cell exists (may commit if created)
    dst_cell = ensure_cell(int(src_cell.project_id), target_date)
    if int(dst_cell.id) == int(src_cell.id):
        return jsonify({"ok": False, "error": "target_date_same"}), 400

    if _cell_has_meaningful_job(dst_cell):
        return jsonify({"ok": False, "error": "target_cell_busy"}), 409

    def _copy_cell_fields(src: PlanCell, dst: PlanCell):
        dst.subproject_id = getattr(src, "subproject_id", None) or None
        dst.shift = src.shift or None
        dst.vehicle_info = src.vehicle_info or None
        dst.note = src.note or None
        dst.job_mail_body = getattr(src, "job_mail_body", None) or None
        dst.isdp_info = getattr(src, "isdp_info", None) or None
        dst.po_info = getattr(src, "po_info", None) or None
        dst.important_note = getattr(src, "important_note", None) or None
        dst.team_name = (src.team_name or None)
        dst.assigned_user_id = getattr(src, "assigned_user_id", None) or None

        # attachments (copy by reference; do not duplicate files)
        dst.lld_hhd_path = getattr(src, "lld_hhd_path", None) or None
        dst.tutanak_path = getattr(src, "tutanak_path", None) or None
        dst.lld_hhd_files = getattr(src, "lld_hhd_files", None) or None
        dst.tutanak_files = getattr(src, "tutanak_files", None) or None
        dst.photo_files = getattr(src, "photo_files", None) or None
        dst.qc_result = getattr(src, "qc_result", None) or None

        dst.updated_at = datetime.now()

    try:
        # clone cell content + people/team
        person_ids = []
        try:
            person_ids = [int(a.person_id) for a in (src_cell.assignments or []) if a and a.person_id]
        except Exception:
            person_ids = []

        _copy_cell_fields(src_cell, dst_cell)
        set_assignments_and_team(dst_cell, person_ids)
        if not (dst_cell.team_name or "").strip():
            dst_cell.team_name = _effective_team_name_for_cell(dst_cell) or None

        now = datetime.now()

        if mode == "copy":
            new_job = _publish_cell(dst_cell, publisher=user, now=now)
            if mark_old_pending:
                job.status = "pending"
                job.closed_at = None
                db.session.add(job)
            db.session.commit()
            try:
                socketio.emit("update_table", namespace="/")
            except Exception:
                pass
            return jsonify({
                "ok": True,
                "mode": "copy",
                "from_date": iso(from_date),
                "to_date": iso(target_date),
                "job_id": int(job.id),
                "new_job_id": int(new_job.id) if new_job else None,
            })

        # move: bind existing job to dst_cell, clear src_cell
        job.cell_id = int(dst_cell.id)
        _sync_job_from_cell(job, dst_cell)
        job.work_date = target_date

        # Clear src cell without deleting uploaded files
        set_assignments_and_team(src_cell, [])
        src_cell.subproject_id = None
        src_cell.shift = None
        src_cell.vehicle_info = None
        src_cell.note = None
        src_cell.job_mail_body = None
        src_cell.isdp_info = None
        src_cell.po_info = None
        src_cell.important_note = None
        src_cell.team_name = None
        src_cell.assigned_user_id = None
        src_cell.lld_hhd_path = None
        src_cell.tutanak_path = None
        src_cell.lld_hhd_files = None
        src_cell.tutanak_files = None
        src_cell.photo_files = None
        src_cell.qc_result = None
        src_cell.updated_at = now

        db.session.add(src_cell)
        db.session.add(dst_cell)
        db.session.add(job)
        db.session.commit()
        try:
            socketio.emit("update_table", namespace="/")
        except Exception:
            pass
        return jsonify({
            "ok": True,
            "mode": "move",
            "from_date": iso(from_date),
            "to_date": iso(target_date),
            "job_id": int(job.id),
            "new_job_id": int(job.id),
        })
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"reschedule_failed: {str(e)}"}), 500


@planner_bp.route("/me/job/<int:job_id>/report", methods=["GET", "POST"])
@login_required
@field_required
def portal_job_report(job_id: int):
    import json as _json

    user = get_current_user()
    job = Job.query.get_or_404(job_id)
    cell = PlanCell.query.get(job.cell_id) if job.cell_id else None
    is_assigned = (int(getattr(job, "assigned_user_id", 0) or 0) == int(user.id)) or (
        cell and int(getattr(cell, "assigned_user_id", 0) or 0) == int(user.id)
    )
    if not is_assigned:
        flash("Bu işe erişiminiz yok.", "danger")
        return redirect(url_for('planner.portal_home'))

    project = Project.query.get(job.project_id)

    latest = (
        JobFeedback.query
        .filter(JobFeedback.job_id == job.id, JobFeedback.user_id == user.id)
        .order_by(JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
        .first()
    )
    latest_answers = {}
    if latest:
        latest_answers = {
            "q1_completed": ("yes" if (latest.outcome or "") == "completed" else "no"),
            "q2_isdp": (latest.isdp_status or ""),
            "extra_work_note": (latest.extra_work_text or ""),
            "issue_note": (latest.notes_text or ""),
        }

    if request.method == "POST":
        try:
            clen = int(request.content_length or 0)
        except Exception:
            clen = 0
        max_bytes = int(current_app.config.get("FEEDBACK_MAX_BYTES") or 0)
        if max_bytes and clen and clen > max_bytes:
            flash(f"Yükleme boyutu çok büyük (maks {int(current_app.config.get('FEEDBACK_MAX_MB') or 50)} MB).", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))

        ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip() or "ip"
        if not _rate_limit(f"upload:feedback:ip:{ip}", limit=30, window_seconds=300):
            flash("Çok fazla yükleme denemesi. Biraz sonra tekrar deneyin.", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))
        if user and user.id:
            if not _rate_limit(f"upload:feedback:user:{user.id}", limit=20, window_seconds=300):
                flash("Çok fazla rapor gönderimi. Biraz sonra tekrar deneyin.", "danger")
                return redirect(url_for("planner.portal_job_report", job_id=job.id))

        if not _csrf_verify(request.form.get("csrf_token", "")):
            flash("Güvenlik doğrulaması başarısız (CSRF).", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))

        prev = (
            JobFeedback.query
            .filter(JobFeedback.job_id == job.id, JobFeedback.user_id == user.id, JobFeedback.outcome != None)
            .order_by(JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
            .first()
        )
        if prev and (getattr(prev, "review_status", "pending") or "pending") != "rejected":
            flash("Rapor incelemede / onaylandı. Yeniden göndermek için reddedilmesi gerekir.", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))

        q1 = (request.form.get("q1_completed") or "").strip().lower()  # yes/no
        q2 = (request.form.get("q2_isdp") or "").strip().lower()  # yes/no/error
        extra_work_note = (request.form.get("extra_work_note") or "").strip()
        issue_note = (request.form.get("issue_note") or "").strip()

        if q1 not in ("yes", "no"):
            flash("Q1 zorunlu.", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))
        if q2 not in ("yes", "no", "error"):
            flash("Q2 zorunlu.", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))

        outcome = "not_completed"
        if q1 == "yes":
            outcome = "completed"
        if q2 == "error":
            outcome = "issue"
        if q1 == "no" and issue_note:
            outcome = "issue"

        now = datetime.now()
        fb = JobFeedback(
            job_id=job.id,
            user_id=user.id,
            submitted_at=now,
            outcome=outcome,
            isdp_status=q2,
            extra_work_text=(extra_work_note if extra_work_note else None),
            notes_text=(issue_note if issue_note else None),
            created_by_user_id=user.id,
            created_at=now,
            closed_at=now,
            note=(issue_note or extra_work_note or None),
            status=("problem" if outcome == "issue" else "completed"),
            review_status="pending",
        )
        db.session.add(fb)
        db.session.flush()

        try:
            media_files = request.files.getlist("media_files") if request.files else []
        except Exception:
            media_files = []
        try:
            media_rows = _save_feedback_uploads(media_files, feedback_id=int(fb.id))
            for m in media_rows:
                db.session.add(m)
        except ValueError as ve:
            db.session.rollback()
            flash(str(ve), "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))
        except Exception:
            db.session.rollback()
            flash("Dosya yükleme hatası.", "danger")
            return redirect(url_for("planner.portal_job_report", job_id=job.id))

        if outcome == "completed":
            job.status = "completed"
            job.closed_at = now
        elif outcome == "issue":
            job.status = "problem"
            job.closed_at = now
        else:
            job.status = "pending"
            job.closed_at = None
        _promote_job_kanban_status(job, "REPORTED", changed_by_user_id=user.id, note="field_report_submit")

        db.session.commit()




        # notify admins
        try:
            title = "Yeni saha raporu"
            body = "{} {} | {} | {} | {}".format(
                (project.project_code if project else ""),
                (project.region if project else ""),
                iso(job.work_date),
                (user.full_name or user.email),
                outcome,
            )
            link_url = url_for("admin_report_detail", report_id=fb.id)
            _notify_admins(event="new_report", title=title, body=body, link_url=link_url, job_id=job.id, meta={
                "report_id": fb.id,
                "job_id": job.id,
                "user_id": user.id,
                "outcome": outcome,
            })
            db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

        # realtime
        try:
            socketio.emit("new_report", {
                "report_id": fb.id,
                "job_id": job.id,
                "user": (user.full_name or user.email),
                "project": (project.project_code if project else ""),
                "city": (project.region if project else ""),
                "submitted_at": fb.submitted_at.isoformat() if fb.submitted_at else now.isoformat(),
                "outcome": outcome,
            }, namespace="/")
        except Exception:
            pass

        flash("Rapor kaydedildi.", "success")
        return redirect(url_for("planner.portal_job_detail", job_id=job.id))

    return render_template(
        "portal_report_form.html",
        user=user,
        job=job,
        project=project,
        latest=latest,
        latest_answers=latest_answers,
    )


@planner_bp.get("/admin/reports")
@login_required
@planner_or_admin_required
def admin_reports():
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    if start_s and end_s:
        start = parse_date(start_s) or (date.today() - timedelta(days=6))
        end = parse_date(end_s) or date.today()
    else:
        end = date.today()
        start = end - timedelta(days=6)

    outcome = (request.args.get("outcome") or "").strip().lower()
    user_id = int(request.args.get("user_id", 0) or 0)
    project_id = int(request.args.get("project_id", 0) or 0)

    q = (
        JobFeedback.query
        .join(Job, Job.id == JobFeedback.job_id)
        .outerjoin(User, User.id == JobFeedback.user_id)
        .join(Project, Project.id == Job.project_id)
        .filter(JobFeedback.outcome != None)
        .filter(JobFeedback.submitted_at >= datetime.combine(start, datetime.min.time()),
                JobFeedback.submitted_at <= datetime.combine(end, datetime.max.time()))
    )
    if outcome:
        q = q.filter(JobFeedback.outcome == outcome)
    if user_id:
        q = q.filter(JobFeedback.user_id == user_id)
    if project_id:
        q = q.filter(Job.project_id == project_id)

    rows = q.order_by(JobFeedback.submitted_at.desc()).limit(500).all()

    users = User.query.filter(User.is_active == True).order_by(User.full_name.asc(), User.email.asc()).all()
    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()

    return render_template(
        "admin_reports.html",
        start=iso(start),
        end=iso(end),
        outcome=outcome,
        user_id=user_id,
        project_id=project_id,
        users=users,
        projects=projects,
        rows=rows,
    )


@planner_bp.get("/admin/reports.xlsx")
@login_required
@planner_or_admin_required
def admin_reports_xlsx():
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    if start_s and end_s:
        start = parse_date(start_s) or (date.today() - timedelta(days=6))
        end = parse_date(end_s) or date.today()
    else:
        end = date.today()
        start = end - timedelta(days=6)

    outcome = (request.args.get("outcome") or "").strip().lower()
    review_status = (request.args.get("review_status") or "").strip().lower()
    user_id = int(request.args.get("user_id", 0) or 0)
    project_id = int(request.args.get("project_id", 0) or 0)

    q = (
        JobFeedback.query
        .join(Job, Job.id == JobFeedback.job_id)
        .outerjoin(User, User.id == JobFeedback.user_id)
        .join(Project, Project.id == Job.project_id)
        .filter(JobFeedback.outcome != None)
        .filter(JobFeedback.submitted_at >= datetime.combine(start, datetime.min.time()),
                JobFeedback.submitted_at <= datetime.combine(end, datetime.max.time()))
    )
    if outcome:
        q = q.filter(JobFeedback.outcome == outcome)
    if review_status in ("pending", "approved", "rejected"):
        q = q.filter(JobFeedback.review_status == review_status)
    if user_id:
        q = q.filter(JobFeedback.user_id == user_id)
    if project_id:
        q = q.filter(Job.project_id == project_id)

    rows = q.order_by(JobFeedback.submitted_at.desc()).limit(5000).all()

    media_counts = {}
    try:
        fb_ids = [r.id for r in rows]
        if fb_ids:
            for fid, cnt in (
                db.session.query(JobFeedbackMedia.feedback_id, db.func.count(JobFeedbackMedia.id))
                .filter(JobFeedbackMedia.feedback_id.in_(fb_ids))
                .group_by(JobFeedbackMedia.feedback_id)
                .all()
            ):
                media_counts[int(fid)] = int(cnt or 0)
    except Exception:
        media_counts = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    ws.append([
        "Report ID",
        "Job ID",
        "Submitted At",
        "Outcome",
        "ISDP",
        "User",
        "City",
        "Project Code",
        "Sub-Project",
        "Team",
        "Review Status",
        "Reviewed At",
        "Review Note",
        "Media Count",
    ])

    for r in rows:
        j = r.job
        p = j.project if j else None
        u = r.user or r.created_by_user
        ws.append([
            int(r.id),
            int(r.job_id or 0),
            r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
            r.outcome or "",
            r.isdp_status or "",
            (u.full_name or u.email) if u else "",
            (p.region if p else ""),
            (p.project_code if p else ""),
            (j.subproject.name if (j and getattr(j, "subproject", None)) else ""),
            (j.team_name if j else ""),
            r.review_status or "pending",
            r.reviewed_at.strftime("%Y-%m-%d %H:%M") if r.reviewed_at else "",
            r.review_note or "",
            media_counts.get(int(r.id), 0),
        ])

    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"admin_reports_{iso(start)}_{iso(end)}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@planner_bp.route("/admin/reports/<int:report_id>", methods=["GET", "POST"])
@login_required
@planner_or_admin_required
def admin_report_detail(report_id: int):
    rep = JobFeedback.query.get_or_404(report_id)
    job = Job.query.get(rep.job_id)
    project = Project.query.get(job.project_id) if job else None
    user = User.query.get(rep.user_id) if getattr(rep, "user_id", None) else None
    if not user and getattr(rep, "created_by_user_id", None):
        user = User.query.get(rep.created_by_user_id)

    answers = {
        "q1_completed": ("yes" if (rep.outcome or "") == "completed" else "no"),
        "q2_isdp": (rep.isdp_status or ""),
        "extra_work_note": (rep.extra_work_text or ""),
        "issue_note": (rep.notes_text or ""),
    }

    if request.method == "POST":
        if not _csrf_verify(request.form.get("csrf_token", "")):
            flash("Güvenlik doğrulaması başarısız (CSRF).", "danger")
            return redirect(url_for("planner.admin_report_detail", report_id=report_id))

        action = (request.form.get("action") or "").strip().lower()
        note = (request.form.get("review_note") or "").strip()
        now = datetime.now()
        reviewer_id = session.get("user_id")

        if action not in ("approve", "reject"):
            flash("İşlem geçersiz.", "danger")
            return redirect(url_for("planner.admin_report_detail", report_id=report_id))

        rep.reviewed_at = now
        rep.reviewed_by_user_id = reviewer_id
        rep.review_note = note or None
        rep.review_status = ("approved" if action == "approve" else "rejected")

        try:
            if job and action == "approve":
                _promote_job_kanban_status(job, "CLOSED", changed_by_user_id=reviewer_id, note="review_approve")
        except Exception:
            pass

        db.session.commit()
        flash("Rapor güncellendi.", "success")
        return redirect(url_for("planner.admin_report_detail", report_id=report_id))

    return render_template(
        "admin_report_detail.html",
        rep=rep,
        job=job,
        project=project,
        user=user,
        answers=answers,
    )


def _render_inline_mail_templates(subject_template: str, body_template: str, ctx: dict) -> Tuple[str, str]:
    subj_t = _mail_subject_env.from_string(subject_template or "")
    body_t = _mail_body_env.from_string(body_template or "")
    subject = (subj_t.render(**ctx) or "").strip()
    body_html = body_t.render(**ctx)
    return subject, body_html


def _jobs_table_html(rows: List[dict]) -> str:
    if not rows:
        return ""
    rows_html = ""
    for it in rows:
        note_html = html.escape(it.get("note") or "").replace(chr(10), "<br>")
        rows_html += f"""
          <tr>
            <td>{html.escape(it.get('date') or '')}</td>
            <td>{html.escape(it.get('city') or '')}</td>
            <td>{html.escape(it.get('project') or '')}</td>
            <td>{html.escape(it.get('team') or '')}</td>
            <td>{html.escape(it.get('shift') or '')}</td>
            <td>{html.escape(it.get('vehicle') or '')}</td>
            <td>{html.escape(it.get('status') or '')}</td>
            <td>{note_html}</td>
          </tr>
        """
    return f"""
      <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">
        <thead>
          <tr>
            <th>Tarih</th><th>Il</th><th>Proje</th><th>Ekip</th><th>Vardiya</th><th>Arac</th><th>Durum</th><th>Not</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    """


@planner_bp.post("/api/mail/preview_bulk")
@login_required
@planner_or_admin_required
def api_mail_preview_bulk():
    import json as _json

    data = request.get_json(force=True, silent=True) or {}
    team_ids = data.get("team_ids") or []
    job_ids = data.get("job_ids") or []

    if not isinstance(team_ids, list) or not team_ids:
        return jsonify({"ok": False, "error": "Ekip secilmedi"}), 400
    if not isinstance(job_ids, list) or not job_ids:
        return jsonify({"ok": False, "error": "Is secilmedi"}), 400

    team_ids = [int(x or 0) for x in team_ids if int(x or 0) > 0]
    job_ids = [int(x or 0) for x in job_ids if int(x or 0) > 0]
    if not team_ids:
        return jsonify({"ok": False, "error": "Ekip secilmedi"}), 400
    if not job_ids:
        return jsonify({"ok": False, "error": "Is secilmedi"}), 400

    jobs = (
        Job.query.join(Project, Project.id == Job.project_id)
        .filter(Job.id.in_(job_ids))
        .order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc())
        .all()
    )
    if not jobs:
        return jsonify({"ok": False, "error": "Isler bulunamadi"}), 404

    projects = {int(j.project_id): j.project for j in jobs if j and j.project_id}
    start = min((j.work_date for j in jobs if j and j.work_date), default=date.today())
    end = max((j.work_date for j in jobs if j and j.work_date), default=date.today())

    recipients_override = data.get("team_emails") or {}
    if not isinstance(recipients_override, dict):
        recipients_override = {}

    cfg_rows = TeamMailConfig.query.filter(TeamMailConfig.team_id.in_(team_ids)).all()
    cfg_by_team = {int(r.team_id): r for r in cfg_rows if r and r.team_id}

    items = []
    for tid in team_ids:
        team = Team.query.get(tid)
        if not team:
            items.append({"team_id": tid, "ok": False, "error": "Ekip bulunamadi"})
            continue

        override_val = recipients_override.get(str(tid)) if str(tid) in recipients_override else recipients_override.get(tid)
        recipients = _split_email_list(str(override_val or ""))
        if not recipients:
            cfg = cfg_by_team.get(int(tid))
            if cfg and cfg.active:
                try:
                    recipients = _json.loads(cfg.emails_json or "[]") or []
                except Exception:
                    recipients = []
        recipients = [r.strip() for r in recipients if isinstance(r, str) and r.strip() and "@" in r]

        rel_jobs = []
        for j in jobs:
            if int(getattr(j, "team_id", 0) or 0) == int(tid):
                rel_jobs.append(j)
                continue
            if (j.team_name or "").strip() and (j.team_name or "").strip() == (team.name or "").strip():
                rel_jobs.append(j)
        job_rows = []
        for j in rel_jobs:
            p = projects.get(int(j.project_id))
            job_rows.append({
                "job_id": int(j.id),
                "date": iso(j.work_date),
                "city": (p.region if p else ""),
                "project": "{} - {}".format((p.project_code if p else ""), (p.project_name if p else "")),
                "team": (j.team_name or team.name or ""),
                "shift": j.shift or "",
                "vehicle": j.vehicle_info or "",
                "status": j.status,
                "note": j.note or "",
            })

        ctx = {
            "team_name": team.name,
            "start": iso(start),
            "end": iso(end),
            "jobs_count": len(job_rows),
            "jobs": job_rows,
            "jobs_table_html": _jobs_table_html(job_rows),
        }

        try:
            subject, body_html = _render_inline_mail_templates(
                DEFAULT_BULK_TEAM_MAIL_SUBJECT_TEMPLATE,
                DEFAULT_BULK_TEAM_MAIL_BODY_TEMPLATE,
                ctx,
            )
        except Exception as e:
            items.append({"team_id": int(team.id), "team_name": team.name, "ok": False, "error": str(e)})
            continue

        email_html = render_template(
            "email_base.html",
            title=subject,
            heading=subject,
            intro="Bu mail /mail/compose ekranindan toplu onizleme olarak uretilmistir.",
            body_html=body_html,
            table_headers=None,
            table_rows=None,
            footer=f"{iso(start)} - {iso(end)}",
        )
        create_mail_log(kind="preview", ok=True, to_addr="(bulk preview)", subject=subject, team_name=team.name, meta={
            "type": "bulk_team",
            "team_id": int(team.id),
            "job_ids": [int(x) for x in job_ids],
        })
        items.append({
            "team_id": int(team.id),
            "team_name": team.name,
            "ok": True,
            "subject": subject,
            "recipients": recipients,
            "jobs_included": len(job_rows),
            "html": email_html,
        })

    overall_ok = all(it.get("ok") for it in items) if items else False
    return jsonify({"ok": overall_ok, "items": items})


@planner_bp.post("/api/mail/send_bulk")
@login_required
@planner_or_admin_required
def api_mail_send_bulk():
    import json as _json
    import time
    import random

    data = request.get_json(force=True, silent=True) or {}
    team_ids = data.get("team_ids") or []
    job_ids = data.get("job_ids") or []

    if not isinstance(team_ids, list) or not team_ids:
        return jsonify({"ok": False, "error": "Ekip secilmedi"}), 400
    if not isinstance(job_ids, list) or not job_ids:
        return jsonify({"ok": False, "error": "Is secilmedi"}), 400

    team_ids = [int(x or 0) for x in team_ids if int(x or 0) > 0]
    job_ids = [int(x or 0) for x in job_ids if int(x or 0) > 0]
    if not team_ids:
        return jsonify({"ok": False, "error": "Ekip secilmedi"}), 400
    if not job_ids:
        return jsonify({"ok": False, "error": "Is secilmedi"}), 400

    jobs = (
        Job.query.join(Project, Project.id == Job.project_id)
        .filter(Job.id.in_(job_ids))
        .order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc())
        .all()
    )
    if not jobs:
        return jsonify({"ok": False, "error": "Isler bulunamadi"}), 404

    projects = {int(j.project_id): j.project for j in jobs if j and j.project_id}
    start = min((j.work_date for j in jobs if j and j.work_date), default=date.today())
    end = max((j.work_date for j in jobs if j and j.work_date), default=date.today())

    # recipients override: {team_id: "a@b.com,c@d.com"}
    recipients_override = data.get("team_emails") or {}
    if not isinstance(recipients_override, dict):
        recipients_override = {}

    cfg_rows = TeamMailConfig.query.filter(TeamMailConfig.team_id.in_(team_ids)).all()
    cfg_by_team = {int(r.team_id): r for r in cfg_rows if r and r.team_id}

    results = []
    for tid in team_ids:
        team = Team.query.get(tid)
        if not team:
            results.append({"team_id": tid, "ok": False, "error": "Ekip bulunamadi", "sent": [], "failed": []})
            continue

        override_val = recipients_override.get(str(tid)) if str(tid) in recipients_override else recipients_override.get(tid)
        recipients = _split_email_list(str(override_val or ""))
        if not recipients:
            cfg = cfg_by_team.get(int(tid))
            if cfg and cfg.active:
                try:
                    recipients = _json.loads(cfg.emails_json or "[]") or []
                except Exception:
                    recipients = []
        recipients = [r.strip() for r in recipients if isinstance(r, str) and r.strip()]
        recipients = [r for r in recipients if "@" in r]

        if not recipients:
            results.append({"team_id": tid, "team_name": team.name, "ok": False, "error": "Ekip mail listesi bos", "sent": [], "failed": []})
            continue

        # jobs relevant to this team (selected jobs are filtered per team)
        rel_jobs = []
        for j in jobs:
            if int(getattr(j, "team_id", 0) or 0) == int(tid):
                rel_jobs.append(j)
                continue
            if (j.team_name or "").strip() and (j.team_name or "").strip() == (team.name or "").strip():
                rel_jobs.append(j)
        job_rows = []
        for j in rel_jobs:
            p = projects.get(int(j.project_id))
            job_rows.append({
                "job_id": int(j.id),
                "date": iso(j.work_date),
                "city": (p.region if p else ""),
                "project": "{} - {}".format((p.project_code if p else ""), (p.project_name if p else "")),
                "team": (j.team_name or team.name or ""),
                "shift": j.shift or "",
                "vehicle": j.vehicle_info or "",
                "status": j.status,
                "note": j.note or "",
            })

        ctx = {
            "team_name": team.name,
            "start": iso(start),
            "end": iso(end),
            "jobs_count": len(job_rows),
            "jobs": job_rows,
            "jobs_table_html": _jobs_table_html(job_rows),
        }

        subject, body_html = _render_inline_mail_templates(
            DEFAULT_BULK_TEAM_MAIL_SUBJECT_TEMPLATE,
            DEFAULT_BULK_TEAM_MAIL_BODY_TEMPLATE,
            ctx,
        )

        email_html = render_template(
            "email_base.html",
            title=subject,
            heading=subject,
            intro="Bu mail /mail/compose ekranindan toplu olarak gonderilmistir.",
            body_html=body_html,
            table_headers=None,
            table_rows=None,
            footer=f"{iso(start)} - {iso(end)}",
        )

        sent = []
        for r in recipients:
            meta = {
                "type": "bulk_team",
                "team_id": int(team.id),
                "job_ids": [int(x) for x in job_ids],
                "recipients": recipients,
                "body_len": len(email_html or ""),
                "body_html_trunc": (email_html or "")[:5000],
            }
            ok = MailService.send(
                mail_type="team",
                recipients=[r],
                subject=subject,
                html=email_html,
                user_id=getattr(current_user, "id", None),
                team_name=team.name,
                meta=meta,
            )
            if ok:
                sent.append(r)
            else:
                err_msg = "Mail gonderilemedi"
                try:
                    _notify_admins(
                        event="mail_fail",
                        title="Mail gonderimi basarisiz",
                        body=f"{team.name} -> {r} | {err_msg}",
                        link_url=url_for("planner.mail_log_page"),
                        meta={"team_id": int(team.id), "to": r, "type": "bulk_team"},
                    )
                    db.session.commit()
                except Exception:
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                failed.append({"to": r, "error": err_msg})

            _time.sleep(random.uniform(1.0, 2.0))

        results.append({
            "team_id": int(team.id),
            "team_name": team.name,
            "ok": len(failed) == 0,
            "subject": subject,
            "recipients": recipients,
            "sent": sent,
            "failed": failed,
            "jobs_included": len(job_rows),
        })

    overall_ok = all(r.get("ok") for r in results) if results else False
    return jsonify({"ok": overall_ok, "results": results})

@planner_bp.post("/api/preview_weekly_email")
@login_required
@admin_required
def api_preview_weekly_email():
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)

    rows = (
        db.session.query(
            Person.id, Person.full_name, Person.email,
            PlanCell.work_date,
            Project.region, Project.project_code, Project.project_name,
            PlanCell.shift, PlanCell.note, PlanCell.vehicle_info,
            PlanCell.job_mail_body,
            Team.name, PlanCell.team_name
        )
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(Team, Team.id == PlanCell.team_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .order_by(Person.full_name.asc(), PlanCell.work_date.asc())
        .all()
    )
    # first person with email
    person = None
    items = []
    for pid, pname, pemail, wd, region, pcode, pname2, shift, note, vehicle, job_mail_body, tname, tname_alt in rows:
        if not pemail:
            continue
        if person is None:
            person = {"id": pid, "full_name": pname, "email": pemail}
        if pid != person["id"]:
            continue
        items.append({
            "date": iso(wd),
            "where": region or "",
            "project": f"{pcode} - {pname2}",
            "shift": shift or "",
            "vehicle": vehicle or "",
            "team": (tname or tname_alt or ""),
            "job_mail_body": job_mail_body or "",
            "note": note or "",
        })

    if not person:
        return jsonify({"ok": False, "error": "Bu hafta icin onizlenecek mail bulunamadi."}), 404

    rows_html = ""
    for it in items:
        job_mail_html = html.escape(it.get("job_mail_body") or "").replace(chr(10), "<br>")
        note_html = html.escape(it.get("note") or "").replace(chr(10), "<br>")
        rows_html += f"""
          <tr>
            <td>{html.escape(it['date'])}</td>
            <td>{html.escape(it['where'])}</td>
            <td>{html.escape(it['project'])}</td>
            <td>{html.escape(it['shift'])}</td>
            <td>{html.escape(it['vehicle'])}</td>
            <td>{html.escape(it['team'])}</td>
            <td>{job_mail_html}</td>
            <td>{note_html}</td>
          </tr>
        """
    table_html = f"""
      <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">
        <thead>
          <tr>
            <th>Tarih</th><th>Il</th><th>Proje</th><th>Vardiya</th><th>Arac</th><th>Ekip</th><th>Is Detay Maili</th><th>Is Detayi</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    """
    subject = f"Haftalik Plan - {iso(ws)}"
    email_html = render_template(
        "email_base.html",
        title=subject,
        heading="Haftalik Is Plani (Onizleme)",
        intro=f"Merhaba {html.escape(person['full_name'] or person['email'])}, bu mail bir onizlemedir.",
        body_html=table_html,
        table_headers=None,
        table_rows=None,
        footer=f"Hafta baslangic: {iso(ws)}",
    )
    create_mail_log(kind="preview", ok=True, to_addr=person["email"], subject=subject, week_start_val=ws, meta={"type": "weekly"})
    return jsonify({"ok": True, "subject": subject, "to": person["email"], "html": email_html})


@planner_bp.post("/api/preview_team_email")
@login_required
@admin_required
def api_preview_team_email():
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", "")) or week_start(date.today())
    team_name = (data.get("team_name") or "").strip()
    if not team_name:
        return jsonify({"ok": False, "error": "team_name eksik"}), 400

    start = ws
    end = ws + timedelta(days=6)
    rows = (
        db.session.query(
            Person.full_name, Person.email,
            PlanCell.work_date,
            Project.region, Project.project_code, Project.project_name,
            PlanCell.shift, PlanCell.note, PlanCell.vehicle_info,
            PlanCell.job_mail_body,
            Team.name, PlanCell.team_name
        )
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(Team, Team.id == PlanCell.team_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .all()
    )

    items = []
    to_email = ""
    for pname, pemail, wd, region, pcode, pname2, shift, note, vehicle, job_mail_body, tname, tname_alt in rows:
        effective_team = (tname or tname_alt or "")
        if effective_team != team_name:
            continue
        if pemail and not to_email:
            to_email = pemail
        items.append({
            "person": pname or "",
            "date": iso(wd),
            "where": region or "",
            "project": f"{pcode} - {pname2}",
            "shift": shift or "",
            "vehicle": vehicle or "",
            "job_mail_body": job_mail_body or "",
            "note": note or "",
        })

    if not items:
        return jsonify({"ok": False, "error": "Bu ekip icin onizleme verisi bulunamadi."}), 404

    rows_html = ""
    for it in sorted(items, key=lambda x: (x["date"], x["person"])):
        job_mail_html = html.escape(it.get("job_mail_body") or "").replace(chr(10), "<br>")
        note_html = html.escape(it.get("note") or "").replace(chr(10), "<br>")
        rows_html += f"""
          <tr>
            <td>{html.escape(it['person'])}</td>
            <td>{html.escape(it['date'])}</td>
            <td>{html.escape(it['where'])}</td>
            <td>{html.escape(it['project'])}</td>
            <td>{html.escape(it['shift'])}</td>
            <td>{html.escape(it['vehicle'])}</td>
            <td>{job_mail_html}</td>
            <td>{note_html}</td>
          </tr>
        """
    table_html = f"""
      <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">
        <thead>
          <tr>
            <th>Personel</th><th>Tarih</th><th>Il</th><th>Proje</th><th>Vardiya</th><th>Arac</th><th>Is Detay Maili</th><th>Is Detayi</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    """
    subject = f"Ekip Plani ({team_name}) - {iso(ws)}"
    email_html = render_template(
        "email_base.html",
        title=subject,
        heading=f"Ekip Plani (Onizleme): {html.escape(team_name)}",
        intro="Bu mail bir onizlemedir. Gonderim oncesi kontrol icindir.",
        body_html=table_html,
        table_headers=None,
        table_rows=None,
        footer=f"Hafta baslangic: {iso(ws)}",
    )
    create_mail_log(kind="preview", ok=True, to_addr=to_email or "-", subject=subject, week_start_val=ws, team_name=team_name, meta={"type": "team"})
    return jsonify({"ok": True, "subject": subject, "to": to_email, "html": email_html})


@planner_bp.post("/api/send_weekly_emails")
@login_required
@observer_required
def api_send_weekly_emails():
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)

    people = Person.query.order_by(Person.full_name.asc()).all()
    tasks: Dict[int, List[dict]] = {p.id: [] for p in people}

    rows = (
        db.session.query(
            Person.id, Person.full_name, Person.email,
            PlanCell.work_date,
            Project.region, Project.project_code, Project.project_name,
            PlanCell.shift, PlanCell.note, PlanCell.vehicle_info,
            PlanCell.job_mail_body, PlanCell.lld_hhd_files, PlanCell.tutanak_files, PlanCell.lld_hhd_path, PlanCell.tutanak_path,
            Team.name, PlanCell.team_name
        )
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(Team, Team.id == PlanCell.team_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .order_by(Person.full_name.asc(), PlanCell.work_date.asc())
        .all()
    )

    for pid, pname, pemail, wd, region, pcode, pname2, shift, note, vehicle, job_mail_body, lld_list_str, tut_list_str, lld_single, tut_single, tname, tname_alt in rows:
        effective_team = (tname or tname_alt or "").strip()
        lld_list = _parse_files(lld_list_str)
        tut_list = _parse_files(tut_list_str)
        if lld_single and not lld_list:
            lld_list = [lld_single]
        if tut_single and not tut_list:
            tut_list = [tut_single]
        tasks[pid].append({
            "date": iso(wd),
            "where": region,
            "project": f"{pcode} - {pname2}",
            "shift": shift or "",
            "note": note or "",
            "vehicle": vehicle or "",
            "team": effective_team,
            "job_mail_body": job_mail_body or "",
            "lld_hhd_files": lld_list,
            "tutanak_files": tut_list,
        })

    sent = 0
    skipped = 0
    errors = []

    for p in people:
        if not p.email:
            skipped += 1
            continue
        items = tasks.get(p.id, [])
        if not items:
            skipped += 1
            continue

        rows_html = ""
        attachments_paths = set()
        for it in items:
            for f in it.get("lld_hhd_files") or []:
                attachments_paths.add(f)
            for f in it.get("tutanak_files") or []:
                attachments_paths.add(f)

            job_mail_html = html.escape(it.get("job_mail_body") or "").replace("\n", "<br>")
            note_html = html.escape(it.get("note") or "").replace("\n", "<br>")
            where_html = html.escape(it.get("where") or "")
            project_html = html.escape(it.get("project") or "")
            shift_html = html.escape(it.get("shift") or "")
            vehicle_html = html.escape(it.get("vehicle") or "")
            team_html = html.escape(it.get("team") or "")
            lld_names = ", ".join([os.path.basename(x) for x in (it.get("lld_hhd_files") or [])])
            tutanak_names = ", ".join([os.path.basename(x) for x in (it.get("tutanak_files") or [])])
            rows_html += f"""
              <tr>
                <td>{it['date']}</td>
                <td>{where_html}</td>
                <td>{project_html}</td>
                <td>{shift_html}</td>
                <td>{vehicle_html}</td>
                <td>{team_html}</td>
                <td>{job_mail_html}</td>
                <td>{note_html}</td>
                <td>{lld_names}</td>
                <td>{tutanak_names}</td>
              </tr>
            """

        table_html = f"""
          <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">
            <thead>
              <tr>
                <th>Tarih</th><th>Il</th><th>Proje</th><th>Vardiya</th><th>Arac</th><th>Ekip</th><th>Is Detay Maili</th><th>Is Detayi</th><th>LLD/HHD</th><th>Tutanak</th>
              </tr>
            </thead>
            <tbody>{rows_html}</tbody>
          </table>
        """

        subject = f"Haftalik Plan - {iso(ws)}"
        email_html = render_template(
            "email_base.html",
            title=subject,
            heading="Haftalik Is Plani",
            intro=f"Merhaba {html.escape(p.full_name or '')}, asagida bu haftaki planin yer aliyor.",
            body_html=table_html,
            table_headers=None,
            table_rows=None,
            footer=f"Hafta baslangic: {iso(ws)}",
        )

        attachments_payload = []
        for fname in attachments_paths:
            full_path = os.path.join(current_app.config["UPLOAD_FOLDER"], fname)
            if not fname or not os.path.exists(full_path):
                continue
            with open(full_path, "rb") as f:
                data_blob = f.read()
            ctype = mimetypes.guess_type(full_path)[0] or "application/octet-stream"
            attachments_payload.append({
                "filename": os.path.basename(full_path),
                "data": data_blob,
                "content_type": ctype
            })

        ok = MailService.send(
            mail_type="weekly",
            recipients=[p.email],
            subject=subject,
            html=email_html,
            attachments=attachments_payload,
            week_start=ws,
            user_id=getattr(current_user, "id", None),
            meta={"type": "weekly", "attachments": sorted(list(attachments_paths))},
        )
        if ok:
            sent += 1
        else:
            errors.append(f"{p.full_name}: Mail gonderimi basarisiz")

    return jsonify({"ok": True, "sent": sent, "skipped": skipped, "errors": errors})


@planner_bp.post("/api/send_team_emails")
@login_required
@observer_required
def api_send_team_emails():
    data = request.get_json(force=True, silent=True) or {}
    ws = parse_date(data.get("week_start", "")) or week_start(date.today())
    start = ws
    end = ws + timedelta(days=6)
    team_name = (data.get("team_name") or "").strip()
    if not team_name:
        return jsonify({"ok": False, "error": "Ekip adi gerekli"}), 400

    people = Person.query.order_by(Person.full_name.asc()).all()
    tasks: Dict[int, List[dict]] = {p.id: [] for p in people}

    rows = (
        db.session.query(
            Person.id, Person.full_name, Person.email,
            PlanCell.work_date,
            Project.region, Project.project_code, Project.project_name,
            PlanCell.shift, PlanCell.note, PlanCell.vehicle_info,
            PlanCell.job_mail_body, PlanCell.lld_hhd_files, PlanCell.tutanak_files, PlanCell.lld_hhd_path, PlanCell.tutanak_path,
            Team.name, PlanCell.team_name
        )
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(Team, Team.id == PlanCell.team_id)
        .filter(PlanCell.work_date >= start, PlanCell.work_date <= end)
        .order_by(Person.full_name.asc(), PlanCell.work_date.asc())
        .all()
    )

    for pid, pname, pemail, wd, region, pcode, pname2, shift, note, vehicle, job_mail_body, lld_list_str, tut_list_str, lld_single, tut_single, tname, tname_alt in rows:
        effective_team = (tname or tname_alt or "").strip()
        if effective_team != team_name:
            continue

        lld_list = _parse_files(lld_list_str)
        tut_list = _parse_files(tut_list_str)
        if lld_single and not lld_list:
            lld_list = [lld_single]
        if tut_single and not tut_list:
            tut_list = [tut_single]

        tasks[pid].append({
            "date": iso(wd),
            "where": region,
            "project": f"{pcode} - {pname2}",
            "shift": shift or "",
            "note": note or "",
            "vehicle": vehicle or "",
            "team": effective_team,
            "job_mail_body": job_mail_body or "",
            "lld_hhd_files": lld_list,
            "tutanak_files": tut_list,
        })

    sent = 0
    skipped = 0
    errors = []

    for p in people:
        items = tasks.get(p.id, [])
        if not items:
            continue
        if not p.email:
            skipped += 1
            continue

        rows_html = ""
        attachments_paths = set()
        for it in items:
            for f in it.get("lld_hhd_files") or []:
                attachments_paths.add(f)
            for f in it.get("tutanak_files") or []:
                attachments_paths.add(f)

            job_mail_html = html.escape(it.get("job_mail_body") or "").replace("\n", "<br>")
            note_html = html.escape(it.get("note") or "").replace("\n", "<br>")
            where_html = html.escape(it.get("where") or "")
            project_html = html.escape(it.get("project") or "")
            shift_html = html.escape(it.get("shift") or "")
            vehicle_html = html.escape(it.get("vehicle") or "")
            team_html = html.escape(it.get("team") or "")
            lld_names = ", ".join([os.path.basename(x) for x in (it.get("lld_hhd_files") or [])])
            tutanak_names = ", ".join([os.path.basename(x) for x in (it.get("tutanak_files") or [])])
            rows_html += f"""
              <tr>
                <td>{it['date']}</td>
                <td>{where_html}</td>
                <td>{project_html}</td>
                <td>{shift_html}</td>
                <td>{vehicle_html}</td>
                <td>{team_html}</td>
                <td>{job_mail_html}</td>
                <td>{note_html}</td>
                <td>{lld_names}</td>
                <td>{tutanak_names}</td>
              </tr>
            """

        table_html = f"""
          <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">
            <thead>
              <tr>
                <th>Tarih</th><th>Il</th><th>Proje</th><th>Vardiya</th><th>Arac</th><th>Ekip</th><th>Is Detay Maili</th><th>Is Detayi</th><th>LLD/HHD</th><th>Tutanak</th>
              </tr>
            </thead>
            <tbody>{rows_html}</tbody>
          </table>
        """

        subject = f"Ekip Plani ({team_name}) - {iso(ws)}"
        email_html = render_template(
            "email_base.html",
            title=subject,
            heading=f"Ekip Plani: {html.escape(team_name)}",
            intro=f"Merhaba {html.escape(p.full_name or '')}, bu ekip icin planin:",
            body_html=table_html,
            table_headers=None,
            table_rows=None,
            footer=f"Hafta baslangic: {iso(ws)}",
        )

        attachments_payload = []
        for fname in attachments_paths:
            full_path = os.path.join(current_app.config["UPLOAD_FOLDER"], fname)
            if not fname or not os.path.exists(full_path):
                continue
            with open(full_path, "rb") as f:
                data_blob = f.read()
            ctype = mimetypes.guess_type(full_path)[0] or "application/octet-stream"
            attachments_payload.append({
                "filename": os.path.basename(full_path),
                "data": data_blob,
                "content_type": ctype
            })

        ok = MailService.send(
            mail_type="team",
            recipients=[p.email],
            subject=subject,
            html=email_html,
            attachments=attachments_payload,
            week_start=ws,
            team_name=team_name,
            user_id=getattr(current_user, "id", None),
            meta={"type": "team", "attachments": sorted(list(attachments_paths))},
        )
        if ok:
            sent += 1
        else:
            errors.append(f"{p.full_name}: Mail gonderimi basarisiz")

    return jsonify({"ok": True, "sent": sent, "skipped": skipped, "errors": errors})


# ---------- TIMESHEET EXCEL ----------
@planner_bp.get("/timesheet.xlsx")
@login_required
def timesheet_excel():
    """
    Personel bazlı haftalık timesheet Excel'i.
    
    Özellikler:
    - Hafta hafta raporlama
    - Personel bazlı listeleme
    - Günlük olarak hangi işe gittiği
    - Hangi bölgede/şehirde çalıştığı
    - Mesai (overtime) bilgisi
    - Proje kodları ve proje adları
    """
    ws = parse_date(request.args.get("week_start", "")) or week_start(date.today())
    start = ws
    days = [start + timedelta(days=i) for i in range(7)]
    
    # Tüm aktif personeli al
    people = Person.query.filter(Person.durum == "Aktif").order_by(Person.full_name.asc()).all()
    
    # Tüm atamaları çek - daha detaylı bilgi
    rows = (
        db.session.query(
            Person.id, Person.full_name,
            PlanCell.work_date,
            Project.region, Project.project_code, Project.project_name,
            PlanCell.shift, PlanCell.note, PlanCell.important_note,
            SubProject.name.label("subproject_name")
        )
        .join(CellAssignment, CellAssignment.person_id == Person.id)
        .join(PlanCell, PlanCell.id == CellAssignment.cell_id)
        .join(Project, Project.id == PlanCell.project_id)
        .outerjoin(SubProject, SubProject.id == PlanCell.subproject_id)
        .filter(PlanCell.work_date >= days[0], PlanCell.work_date <= days[-1])
        .all()
    )
    
    # Mesai kayıtlarını çek (TeamOvertime tablosundan)
    overtime_rows = (
        db.session.query(
            TeamOvertime.person_id, TeamOvertime.work_date, TeamOvertime.duration_hours
        )
        .filter(
            TeamOvertime.work_date >= days[0],
            TeamOvertime.work_date <= days[-1]
        )
        .all()
    )
    
    # (person_id, date_iso) -> mesai saati
    overtime_map: Dict[Tuple[int, str], float] = {}
    for row in overtime_rows:
        key = (int(row.person_id), iso(row.work_date))
        overtime_map[key] = float(row.duration_hours or 0)
    
    # (person_id, date_iso) -> detaylı görev listesi
    mp: Dict[Tuple[int, str], List[Dict]] = {}
    for row in rows:
        key = (row.id, iso(row.work_date))
        task_info = {
            "city": row.region or "",
            "project_code": row.project_code or "",
            "project_name": row.project_name or "",
            "subproject": row.subproject_name or "",
            "shift": row.shift or "",
            "note": row.note or "",
            "important_note": row.important_note or ""
        }
        mp.setdefault(key, []).append(task_info)
    
    # Personel durumlarını çek (izin, üretimde vb.)
    st = get_person_status_map(days)
    
    # Excel dosyası oluştur
    wb = Workbook()
    wsht = wb.active
    wsht.title = f"Timesheet {start.strftime('%d.%m.%Y')}"
    
    # Başlık satırı
    header = ["Personel", "Firma", "Seviye"]
    for i, dday in enumerate(days):
        header.append(f"{dday.strftime('%d.%m')}\\n{TR_DAYS[i]}")
    header.append("Toplam\\nMesai")
    
    wsht.append(header)
    
    # Başlık stili
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, len(header)+1):
        cell = wsht.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Veri satırları
    r = 2
    for p in people:
        row = [
            p.full_name or "",
            p.firma.name if p.firma else "",
            p.seviye.name if p.seviye else ""
        ]
        
        for dday in days:
            k = iso(dday)
            tasks = mp.get((p.id, k), [])
            
            if tasks:
                cell_parts = []
                for task in tasks:
                    parts = []
                    # Şehir
                    if task["city"]:
                        parts.append(task["city"])
                    # Proje kodu ve adı
                    if task["project_code"]:
                        parts.append(f"{task['project_code']} {task['project_name']}")
                    # Alt proje
                    if task["subproject"]:
                        parts.append(task["subproject"])
                    
                    # Normal Çalışma Saati
                    work_hours = calculate_hours_from_shift(task["shift"])
                    if work_hours and work_hours > 0:
                        parts.append(f"Çalışma: {work_hours} saat")

                    # Mesai saati
                    overtime = overtime_map.get((p.id, k), 0)
                    if overtime and overtime > 0:
                        parts.append(f"Mesai: +{overtime} saat")
                    
                    # Önemli not
                    if task["important_note"]:
                        parts.append(f"Not: {task['important_note']}")
                    
                    cell_parts.append(" | ".join(parts))
                
                row.append("\\n".join(cell_parts))
            else:
                stv = st.get((p.id, k), "available")
                status_map = {
                    "leave": "IZINLI",
                    "production": "URETIMDE",
                    "available": "",
                    "sick": "HASTA",
                    "permission": "IZIN",
                }
                row.append(status_map.get(stv, ""))
        
        # Toplam mesai
        total_overtime = sum(overtime_map.get((p.id, iso(d)), 0) for d in days)
        row.append(f"{total_overtime} saat" if total_overtime > 0 else "")
        
        wsht.append(row)
        r += 1
    
    # Sütun genişliklerini ayarla
    wsht.column_dimensions["A"].width = 25
    wsht.column_dimensions["B"].width = 18
    wsht.column_dimensions["C"].width = 12
    for i in range(4, 11):
        wsht.column_dimensions[get_column_letter(i)].width = 40
    wsht.column_dimensions["K"].width = 12
    
    # Satır yüksekliği ve hücre stili
    for rr in range(2, wsht.max_row + 1):
        wsht.row_dimensions[rr].height = 45
        for cc in range(1, wsht.max_column + 1):
            cell = wsht.cell(row=rr, column=cc)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Hafta sonu günlerini renklendir
            if cc >= 4 and cc <= 10:
                day_idx = cc - 4
                if day_idx >= 5:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Excel dosyasını döndür
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"timesheet_{start.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )






# ---------- ASSIGNMENT DETAIL ----------
@planner_bp.get("/assignment/<int:job_id>")
@login_required
def assignment_page(job_id: int):
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    project = Project.query.get(j.project_id)
    cell = PlanCell.query.get(j.cell_id) if j.cell_id else None

    people = (
        db.session.query(Person)
        .join(JobAssignment, JobAssignment.person_id == Person.id)
        .filter(JobAssignment.job_id == j.id)
        .order_by(Person.full_name.asc())
        .all()
    )

    lld_list = _parse_files(getattr(cell, "lld_hhd_files", None)) if cell else []
    tut_list = _parse_files(getattr(cell, "tutanak_files", None)) if cell else []
    if cell and getattr(cell, "lld_hhd_path", None) and not lld_list:
        lld_list = [cell.lld_hhd_path]
    if cell and getattr(cell, "tutanak_path", None) and not tut_list:
        tut_list = [cell.tutanak_path]
    photo_list = _parse_files(getattr(cell, "photo_files", None)) if cell else []
    qc_result = (getattr(cell, "qc_result", "") or "") if cell else ""

    feedback_rows = JobFeedback.query.filter_by(job_id=j.id).order_by(JobFeedback.closed_at.desc()).all()
    status_rows = []
    try:
        status_rows = JobStatusHistory.query.filter_by(job_id=j.id).order_by(JobStatusHistory.changed_at.desc()).all()
    except Exception:
        status_rows = []

    mail_rows = []
    try:
        import json as _json
        logs = MailLog.query.order_by(MailLog.created_at.desc()).limit(600).all()
        for r in logs:
            meta = {}
            try:
                meta = _json.loads(r.meta_json or "{}")
            except Exception:
                meta = {}
            if meta.get("type") != "job":
                continue
            if int(meta.get("project_id") or 0) != int(j.project_id):
                continue
            if str(meta.get("work_date") or "") != iso(j.work_date):
                continue
            mail_rows.append({
                "created_at": r.created_at,
                "kind": r.kind or "",
                "ok": bool(r.ok),
                "to": r.to_addr or "",
                "subject": r.subject or "",
                "error": (r.error or ""),
                "meta": meta,
            })
            if len(mail_rows) >= 60:
                break
    except Exception:
        mail_rows = []

    return render_template(
        "assignment.html",
        j=j,
        project=project,
        cell=cell,
        people=people,
        lld_files=lld_list,
        tutanak_files=tut_list,
        photo_files=photo_list,
        qc_result=qc_result,
        feedback_rows=feedback_rows,
        status_rows=status_rows,
        mail_rows=mail_rows,
        kanban_columns=KANBAN_COLUMNS,
    )

@planner_bp.post("/assignment/<int:job_id>/kanban")
@login_required
@observer_required
def assignment_set_kanban(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    new_status = _normalize_kanban_status((request.form.get("kanban_status") or "").strip())
    if new_status not in KANBAN_COLUMNS:
        flash("Gecersiz durum.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    changed = _set_job_kanban_status(j, new_status, changed_by_user_id=session.get("user_id"), note="manual_set")
    if changed:
        db.session.commit()
        flash("Durum guncellendi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))

@planner_bp.post("/assignment/<int:job_id>/feedback")
@login_required
@observer_required
def assignment_set_feedback(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    status = (request.form.get("status") or "").strip().lower()
    note = (request.form.get("note") or "").strip()
    if status not in ("completed", "problem", "pending"):
        flash("Durum gecersiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    if status == "pending":
        JobFeedback.query.filter_by(job_id=job_id).delete()
        j.status = "pending"
        j.closed_at = None
        db.session.commit()
        flash("Geri bildirim temizlendi.", "success")
        return redirect(url_for("assignment_page", job_id=job_id))
    fb = JobFeedback(job_id=job_id, status=status, note=note or None, closed_at=datetime.now(), created_by_user_id=session.get("user_id"))
    db.session.add(fb)
    j.status = status
    j.closed_at = fb.closed_at
    try:
        _promote_job_kanban_status(j, "REPORTED", changed_by_user_id=session.get("user_id"), note="staff_feedback")
    except Exception:
        pass
    db.session.commit()
    flash("Geri bildirim kaydedildi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))

@planner_bp.post("/assignment/<int:job_id>/attachments/upload")
@login_required
@observer_required
def assignment_upload_attachments(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    cell = PlanCell.query.get(j.cell_id) if j.cell_id else None
    if not cell:
        flash("Hucre bulunamadi.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    prefix = f"{j.project_id}-{iso(j.work_date)}"
    lld_files = request.files.getlist("lld_hhd")
    if lld_files:
        cur = _parse_files(getattr(cell, "lld_hhd_files", None))
        for fs in lld_files:
            if fs and fs.filename:
                cur.append(save_uploaded_file(fs, f"lldhhd-{prefix}"))
        cell.lld_hhd_files = _dump_files(cur) if cur else None
    tutanak_files = request.files.getlist("tutanak")
    if tutanak_files:
        cur = _parse_files(getattr(cell, "tutanak_files", None))
        for fs in tutanak_files:
            if fs and fs.filename:
                cur.append(save_uploaded_file(fs, f"tutanak-{prefix}"))
        cell.tutanak_files = _dump_files(cur) if cur else None
    cell.updated_at = datetime.now()
    db.session.commit()
    flash("Ekler yüklendi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))

@planner_bp.post("/assignment/<int:job_id>/photos/upload")
@login_required
@observer_required
def assignment_upload_photos(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    cell = PlanCell.query.get(j.cell_id) if j.cell_id else None
    if not cell:
        flash("Hucre bulunamadi.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    prefix = f"{j.project_id}-{iso(j.work_date)}"
    photos = request.files.getlist("photos")
    cur = _parse_files(getattr(cell, "photo_files", None))
    for fs in photos:
        if fs and fs.filename:
            cur.append(save_uploaded_file(fs, f"photo-{prefix}"))
    cell.photo_files = _dump_files(cur) if cur else None
    cell.updated_at = datetime.now()
    db.session.commit()
    flash("Fotograflar yüklendi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))

@planner_bp.post("/assignment/<int:job_id>/file/delete")
@login_required
@observer_required
def assignment_delete_file(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    kind = (request.form.get("kind") or "").strip()
    filename = (request.form.get("filename") or "").strip()
    if not filename or any(x in filename for x in ("..", "/", "\\")):
        flash("Dosya adi gecersiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    cell = PlanCell.query.get(j.cell_id) if j.cell_id else None
    if not cell:
        flash("Hucre bulunamadi.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    if kind == "lld_hhd":
        cur = _parse_files(getattr(cell, "lld_hhd_files", None))
        cur = [f for f in cur if f != filename]
        cell.lld_hhd_files = _dump_files(cur) if cur else None
        if getattr(cell, "lld_hhd_path", None) == filename:
            cell.lld_hhd_path = None
    elif kind == "tutanak":
        cur = _parse_files(getattr(cell, "tutanak_files", None))
        cur = [f for f in cur if f != filename]
        cell.tutanak_files = _dump_files(cur) if cur else None
        if getattr(cell, "tutanak_path", None) == filename:
            cell.tutanak_path = None
    elif kind == "photo":
        cur = _parse_files(getattr(cell, "photo_files", None))
        cur = [f for f in cur if f != filename]
        cell.photo_files = _dump_files(cur) if cur else None
    else:
        flash("Tur gecersiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    delete_upload(filename)
    cell.updated_at = datetime.now()
    db.session.commit()
    flash("Dosya silindi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))

@planner_bp.post("/assignment/<int:job_id>/qc")
@login_required
@observer_required
def assignment_set_qc(job_id: int):
    if not _csrf_verify(request.form.get("csrf_token", "")):
        flash("CSRF dogrulamasi basarisiz.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    qc = (request.form.get("qc_result") or "").strip()
    j = Job.query.get(job_id)
    if not j:
        flash("Is bulunamadi.", "danger")
        return redirect(url_for('planner.plan_week'))
    cell = PlanCell.query.get(j.cell_id) if j.cell_id else None
    if not cell:
        flash("Hucre bulunamadi.", "danger")
        return redirect(url_for("assignment_page", job_id=job_id))
    cell.qc_result = qc or None
    cell.updated_at = datetime.now()
    db.session.commit()
    flash("QC kaydedildi.", "success")
    return redirect(url_for("assignment_page", job_id=job_id))
# ---------- BOARD (OPERATIONAL LIST) ----------
def _build_board_jobs_query(
    *,
    start: date,
    end: date,
    q_text: str,
    assignee_q: str,
    status: str,
    project_id: int,
    subproject_id: int,
    city: str,
    assigned_user_id: int,
    overdue_only: bool,
    unassigned_only: bool,
    published_only: bool,
):
    q = (
        db.session.query(Job)
        .join(Project, Project.id == Job.project_id)
        .outerjoin(SubProject, SubProject.id == Job.subproject_id)
        .outerjoin(User, User.id == Job.assigned_user_id)
        .filter(Job.work_date >= start, Job.work_date <= end)
    )

    st = _normalize_kanban_status((status or "").strip())
    if st:
        q = q.filter(Job.kanban_status == st)

    if project_id:
        q = q.filter(Job.project_id == int(project_id))
    if subproject_id:
        q = q.filter(Job.subproject_id == int(subproject_id))
    if city:
        q = q.filter(Project.region == city)
    if assigned_user_id:
        q = q.filter(Job.assigned_user_id == int(assigned_user_id))

    if published_only:
        q = q.filter(Job.is_published == True)
    if unassigned_only:
        q = q.filter(Job.assigned_user_id.is_(None))
    if overdue_only:
        today = date.today()
        q = q.filter(Job.work_date < today, Job.kanban_status != "CLOSED")

    if q_text:
        like = f"%{q_text}%"
        q = q.filter(or_(
            Project.project_code.ilike(like),
            Project.project_name.ilike(like),
            Project.region.ilike(like),
            SubProject.name.ilike(like),
            SubProject.code.ilike(like),
            Job.team_name.ilike(like),
            Job.note.ilike(like),
            User.full_name.ilike(like),
            User.email.ilike(like),
        ))

    if assignee_q:
        like = f"%{assignee_q}%"
        q = q.filter(or_(
            User.full_name.ilike(like),
            User.email.ilike(like),
            db.session.query(JobAssignment.id)
            .join(Person, Person.id == JobAssignment.person_id)
            .filter(JobAssignment.job_id == Job.id, Person.full_name.ilike(like))
            .exists(),
        ))

    return q


@planner_bp.get("/board")
@login_required
@planner_or_admin_required
def board_page():
    start, end = _parse_board_date_range_args()

    projects = Project.query.filter(Project.region != "-").order_by(Project.region.asc(), Project.project_code.asc()).all()
    cities = sorted({(p.region or "").strip() for p in projects if (p.region or "").strip()})
    users = User.query.filter(User.is_active == True).order_by(User.full_name.asc().nullslast(), User.email.asc()).all()

    return render_template(
        "board.html",
        start=iso(start),
        end=iso(end),
        projects=projects,
        cities=cities,
        users=users,
        statuses=KANBAN_COLUMNS,
        kanban_labels=KANBAN_LABEL_TR,
    )


@planner_bp.get("/kanban")
@login_required
@planner_or_admin_required
def kanban_page():
    start, end = _parse_date_range_args()
    upsert_jobs_for_range(start, end)

    q = Job.query.filter(Job.work_date >= start, Job.work_date <= end)
    q = q.join(Project, Project.id == Job.project_id)
    jobs = q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc()).all()

    jobs_by_status = {c: [] for c in KANBAN_COLUMNS}
    for j in jobs:
        st = _normalize_kanban_status(getattr(j, "kanban_status", None))
        jobs_by_status[st].append(j)

    return render_template(
        "kanban.html",
        columns=KANBAN_COLUMNS,
        jobs_by_status=jobs_by_status,
        start=iso(start),
        end=iso(end),
    )


@planner_bp.get("/api/board/jobs")
@login_required
@planner_or_admin_required
def api_board_jobs():
    start, end = _parse_board_date_range_args()
    try:
        page = int(request.args.get("page", 1) or 1)
    except Exception:
        page = 1
    page = max(1, page)
    try:
        page_size = int(request.args.get("page_size", 50) or 50)
    except Exception:
        page_size = 50
    page_size = max(10, min(200, page_size))

    q_text = (request.args.get("q") or "").strip()
    assignee_q = (request.args.get("assignee_q") or "").strip()
    status = (request.args.get("status") or "").strip()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    city = (request.args.get("city") or "").strip()
    assigned_user_id = int(request.args.get("assigned_user_id", 0) or 0)
    overdue_only = _parse_bool_arg("overdue_only", default=False)
    unassigned_only = _parse_bool_arg("unassigned_only", default=False)
    published_only = _parse_bool_arg("published_only", default=False)

    upsert_jobs_for_range(start, end)

    base_q = _build_board_jobs_query(
        start=start,
        end=end,
        q_text=q_text,
        assignee_q=assignee_q,
        status=status,
        project_id=project_id,
        subproject_id=subproject_id,
        city=city,
        assigned_user_id=assigned_user_id,
        overdue_only=overdue_only,
        unassigned_only=unassigned_only,
        published_only=published_only,
    )

    total_count = int((base_q.with_entities(db.func.count(db.func.distinct(Job.id))).scalar() or 0))

    # latest report per job (submitted) + review status + id
    last_report_at = (
        db.session.query(db.func.max(JobFeedback.submitted_at))
        .filter(JobFeedback.job_id == Job.id, JobFeedback.outcome != None)
        .correlate(Job)
        .scalar_subquery()
    )
    last_report_id = (
        db.session.query(JobFeedback.id)
        .filter(JobFeedback.job_id == Job.id, JobFeedback.outcome != None, JobFeedback.submitted_at != None)
        .order_by(JobFeedback.submitted_at.desc(), JobFeedback.id.desc())
        .limit(1)
        .correlate(Job)
        .scalar_subquery()
    )
    last_review_status = (
        db.session.query(JobFeedback.review_status)
        .filter(JobFeedback.job_id == Job.id, JobFeedback.outcome != None, JobFeedback.submitted_at != None)
        .order_by(JobFeedback.submitted_at.desc(), JobFeedback.id.desc())
        .limit(1)
        .correlate(Job)
        .scalar_subquery()
    )

    # Önce tüm işleri al (sayfalama öncesi birleştirme için)
    rows = (
        base_q.with_entities(Job, Project, SubProject, User, last_report_at, last_review_status, last_report_id)
        .order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc(), Job.id.asc())
        .all()
    )

    # Birleştirilmiş işleri grupla
    today = date.today()
    job_groups = {}  # (project_id, subproject_id, team_name) -> list of jobs
    job_to_group = {}  # job_id -> group_key
    
    for j, p, sp, u, rep_at, review_st, rep_id in rows:
        subproject_id = int(getattr(j, "subproject_id", 0) or 0)
        team_name = (j.team_name or "").strip()
        group_key = (j.project_id, subproject_id, team_name)
        
        if group_key not in job_groups:
            job_groups[group_key] = []
        job_groups[group_key].append((j, p, sp, u, rep_at, review_st, rep_id))
        job_to_group[j.id] = group_key
    
    # Ardışık tarihlerde olan işleri birleştir
    merged_groups = []
    for group_key, group_jobs in job_groups.items():
        # Tarihe göre sırala
        group_jobs.sort(key=lambda x: x[0].work_date if x[0].work_date else date.min)
        
        # Ardışık tarihlerde olan işleri grupla
        current_group = [group_jobs[0]]
        for i in range(1, len(group_jobs)):
            prev_job = current_group[-1][0]
            curr_job = group_jobs[i][0]
            
            # Ardışık tarihler mi kontrol et
            if prev_job.work_date and curr_job.work_date:
                days_diff = (curr_job.work_date - prev_job.work_date).days
                if days_diff == 1:
                    # Ardışık, gruba ekle
                    current_group.append(group_jobs[i])
                else:
                    # Ardışık değil, mevcut grubu kaydet ve yeni grup başlat
                    if len(current_group) > 1:
                        merged_groups.append(current_group)
                    current_group = [group_jobs[i]]
            else:
                # Tarih yok, ayrı grup
                if len(current_group) > 1:
                    merged_groups.append(current_group)
                current_group = [group_jobs[i]]
        
        # Son grubu kaydet
        if len(current_group) > 1:
            merged_groups.append(current_group)
    
    # Birleştirilmiş grupları işaretle
    merged_job_ids = set()
    for group in merged_groups:
        for j, _, _, _, _, _, _ in group:
            merged_job_ids.add(j.id)
    
    # Sayfalama için tüm işleri al (birleştirilmiş olanlar tek sayılacak)
    all_jobs = []
    processed_job_ids = set()
    
    for j, p, sp, u, rep_at, review_st, rep_id in rows:
        if j.id in processed_job_ids:
            continue
        
        # Bu iş birleştirilmiş bir grubun parçası mı?
        if j.id in merged_job_ids:
            # Grubu bul
            group = None
            for g in merged_groups:
                if any(gj[0].id == j.id for gj in g):
                    group = g
                    break
            
            if group:
                # Grubun tüm işlerini işaretle
                for gj, _, _, _, _, _, _ in group:
                    processed_job_ids.add(gj.id)
                
                # İlk işin bilgilerini kullan (tarih aralığını ekle)
                first_job = group[0][0]
                last_job = group[-1][0]
                assigned_name = ""
                if group[0][3]:  # User
                    assigned_name = (group[0][3].full_name or group[0][3].email or group[0][3].username or "").strip()
                
                # En son raporu bul (tüm gruptan)
                latest_rep_at = None
                latest_rep_id = 0
                latest_review_st = ""
                for gj, _, _, _, rep_at, review_st, rep_id in group:
                    if rep_at and (not latest_rep_at or rep_at > latest_rep_at):
                        latest_rep_at = rep_at
                        latest_rep_id = int(rep_id or 0) if rep_id else 0
                        latest_review_st = (str(review_st or "") if review_st else "")
                
                # En son kapanış tarihini bul
                latest_closed_at = None
                for gj, _, _, _, _, _, _ in group:
                    if getattr(gj, "closed_at", None):
                        if not latest_closed_at or gj.closed_at > latest_closed_at:
                            latest_closed_at = gj.closed_at
                
                # Yayınlanma durumunu kontrol et (tüm grup yayınlanmış mı?)
                all_published = all(bool(getattr(gj, "is_published", False)) for gj, _, _, _, _, _, _ in group)
                latest_published_at = None
                for gj, _, _, _, _, _, _ in group:
                    if getattr(gj, "published_at", None):
                        if not latest_published_at or gj.published_at > latest_published_at:
                            latest_published_at = gj.published_at
                
                # Kanban durumunu kontrol et (en kötü durumu al)
                statuses = [_normalize_kanban_status(getattr(gj, "kanban_status", None)) for gj, _, _, _, _, _, _ in group]
                worst_status = "CLOSED"
                if "PLANNED" in statuses:
                    worst_status = "PLANNED"
                elif "ASSIGNED" in statuses:
                    worst_status = "ASSIGNED"
                elif "PUBLISHED" in statuses:
                    worst_status = "PUBLISHED"
                elif "REPORTED" in statuses:
                    worst_status = "REPORTED"
                
                # Gecikme durumunu kontrol et
                is_overdue = any(
                    gj.work_date and gj.work_date < today and _normalize_kanban_status(getattr(gj, "kanban_status", None)) != "CLOSED"
                    for gj, _, _, _, _, _, _ in group
                )
                
                all_jobs.append({
                    "id": int(first_job.id),  # İlk işin ID'sini kullan
                    "merged_job_ids": [int(gj.id) for gj, _, _, _, _, _, _ in group],  # Tüm iş ID'leri
                    "work_date": iso(first_job.work_date) if first_job.work_date else "",
                    "work_date_end": iso(last_job.work_date) if last_job.work_date else "",
                    "city": (group[0][1].region if group[0][1] else ""),
                    "project_id": int(first_job.project_id or 0),
                    "project_code": (group[0][1].project_code if group[0][1] else ""),
                    "project_name": (group[0][1].project_name if group[0][1] else ""),
                    "subproject_id": int(getattr(first_job, "subproject_id", 0) or 0),
                    "subproject_name": (group[0][2].name if group[0][2] else ""),
                    "team_name": (first_job.team_name or ""),
                    "assigned_user_id": int(getattr(first_job, "assigned_user_id", 0) or 0) if getattr(first_job, "assigned_user_id", None) else 0,
                    "assigned_user_name": assigned_name,
                    "kanban_status": worst_status,
                    "is_published": all_published,
                    "published_at": latest_published_at.isoformat() if latest_published_at else None,
                    "closed_at": latest_closed_at.isoformat() if latest_closed_at else None,
                    "last_report_at": latest_rep_at.isoformat() if latest_rep_at else None,
                    "last_report_id": latest_rep_id,
                    "review_status": latest_review_st,
                    "is_overdue": is_overdue,
                    "is_merged": True,
                })
                continue
        
        # Birleştirilmemiş iş
        processed_job_ids.add(j.id)
        assigned_name = ""
        if u:
            assigned_name = (u.full_name or u.email or u.username or "").strip()
        all_jobs.append({
            "id": int(j.id),
            "merged_job_ids": [int(j.id)],
            "work_date": iso(j.work_date) if j.work_date else "",
            "work_date_end": iso(j.work_date) if j.work_date else "",
            "city": (p.region if p else ""),
            "project_id": int(j.project_id or 0),
            "project_code": (p.project_code if p else ""),
            "project_name": (p.project_name if p else ""),
            "subproject_id": int(getattr(j, "subproject_id", 0) or 0),
            "subproject_name": (sp.name if sp else ""),
            "team_name": (j.team_name or ""),
            "assigned_user_id": int(getattr(j, "assigned_user_id", 0) or 0) if getattr(j, "assigned_user_id", None) else 0,
            "assigned_user_name": assigned_name,
            "kanban_status": _normalize_kanban_status(getattr(j, "kanban_status", None)),
            "is_published": bool(getattr(j, "is_published", False)),
            "published_at": j.published_at.isoformat() if getattr(j, "published_at", None) else None,
            "closed_at": j.closed_at.isoformat() if getattr(j, "closed_at", None) else None,
            "last_report_at": rep_at.isoformat() if rep_at else None,
            "last_report_id": int(rep_id or 0) if rep_id else 0,
            "review_status": (str(review_st or "") if review_st else ""),
            "is_overdue": bool(j.work_date and j.work_date < today and _normalize_kanban_status(getattr(j, "kanban_status", None)) != "CLOSED"),
            "is_merged": False,
        })
    
    # Sayfalama uygula
    total_count = len(all_jobs)
    out = all_jobs[(page - 1) * page_size:page * page_size]

    return jsonify({
        "ok": True,
        "range": {"start": iso(start), "end": iso(end)},
        "page": page,
        "page_size": page_size,
        "total_count": total_count,
        "rows": out,
    })


@planner_bp.get("/api/board/job/<int:job_id>/detail")
@login_required
@planner_or_admin_required
def api_board_job_detail(job_id: int):
    j = Job.query.get(job_id)
    if not j:
        return jsonify({"ok": False, "error": "Is bulunamadi."}), 404

    p = Project.query.get(j.project_id) if getattr(j, "project_id", None) else None
    sp = SubProject.query.get(j.subproject_id) if getattr(j, "subproject_id", None) else None
    u = User.query.get(j.assigned_user_id) if getattr(j, "assigned_user_id", None) else None

    people = (
        db.session.query(Person.full_name, Person.phone, Person.email)
        .join(JobAssignment, JobAssignment.person_id == Person.id)
        .filter(JobAssignment.job_id == j.id)
        .order_by(Person.full_name.asc())
        .all()
    )

    cell = PlanCell.query.get(j.cell_id) if getattr(j, "cell_id", None) else None
    attachments = {"lld_hhd": [], "tutanak": [], "photo": []}
    try:
        if cell:
            attachments["lld_hhd"] = _parse_files(getattr(cell, "lld_hhd_files", None) or getattr(cell, "lld_hhd_path", None))
            attachments["tutanak"] = _parse_files(getattr(cell, "tutanak_files", None) or getattr(cell, "tutanak_path", None))
            attachments["photo"] = _parse_files(getattr(cell, "photo_files", None))
    except Exception:
        attachments = {"lld_hhd": [], "tutanak": [], "photo": []}

    latest = (
        JobFeedback.query
        .filter(JobFeedback.job_id == j.id, JobFeedback.outcome != None)
        .order_by(JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
        .first()
    )
    latest_media = []
    if latest:
        try:
            latest_media = JobFeedbackMedia.query.filter_by(feedback_id=latest.id).order_by(JobFeedbackMedia.id.asc()).all()
        except Exception:
            latest_media = []

    hist = JobStatusHistory.query.filter_by(job_id=j.id).order_by(JobStatusHistory.changed_at.desc(), JobStatusHistory.id.desc()).limit(50).all()

    def _u_name(user: Optional[User]) -> str:
        if not user:
            return ""
        return (user.full_name or user.email or user.username or "").strip()

    return jsonify({
        "ok": True,
        "job": {
            "id": int(j.id),
            "work_date": iso(j.work_date) if j.work_date else "",
            "city": (p.region if p else ""),
            "project_code": (p.project_code if p else ""),
            "project_name": (p.project_name if p else ""),
            "subproject_name": (sp.name if sp else ""),
            "team_name": (j.team_name or ""),
            "shift": (j.shift or ""),
            "vehicle_info": (j.vehicle_info or ""),
            "note": (j.note or ""),
            "kanban_status": _normalize_kanban_status(getattr(j, "kanban_status", None)),
            "is_published": bool(getattr(j, "is_published", False)),
            "published_at": j.published_at.isoformat() if getattr(j, "published_at", None) else None,
            "published_by": _u_name(j.published_by_user) if getattr(j, "published_by_user", None) else "",
            "assigned_user": _u_name(u),
            "closed_at": j.closed_at.isoformat() if getattr(j, "closed_at", None) else None,
        },
        "people": [{"full_name": n, "phone": ph or "", "email": em or ""} for (n, ph, em) in people],
        "attachments": attachments,
        "latest_report": None if not latest else {
            "id": int(latest.id),
            "submitted_at": latest.submitted_at.isoformat() if getattr(latest, "submitted_at", None) else None,
            "outcome": (latest.outcome or ""),
            "isdp_status": (latest.isdp_status or ""),
            "extra_work_text": (latest.extra_work_text or ""),
            "notes_text": (latest.notes_text or ""),
            "review_status": (getattr(latest, "review_status", "") or ""),
            "reviewed_at": latest.reviewed_at.isoformat() if getattr(latest, "reviewed_at", None) else None,
            "review_note": (getattr(latest, "review_note", "") or ""),
            "media": [{
                "id": int(m.id),
                "file_path": (m.file_path or ""),
                "file_type": (m.file_type or ""),
                "original_name": (m.original_name or ""),
                "uploaded_at": m.uploaded_at.isoformat() if getattr(m, "uploaded_at", None) else None,
            } for m in (latest_media or [])],
        },
        "status_history": [{
            "changed_at": h.changed_at.isoformat() if getattr(h, "changed_at", None) else None,
            "from_status": (h.from_status or ""),
            "to_status": (h.to_status or ""),
            "note": (h.note or ""),
            "changed_by": _u_name(h.changed_by) if getattr(h, "changed_by", None) else "",
        } for h in (hist or [])],
    })

@planner_bp.post("/api/board/move")
@login_required
@planner_or_admin_required
def api_board_move():
    data = request.get_json(force=True, silent=True) or {}
    token = str(data.get("csrf_token") or "")
    if not _csrf_verify(token):
        return jsonify({"ok": False, "error": "CSRF dogrulamasi basarisiz."}), 400

    current_user = get_current_user()
    if current_user and current_user.role == "g”zlemci":
        return jsonify({"ok": False, "error": "Gozlemci rolu degisiklik yapamaz."}), 403

    job_id = int(data.get("job_id", 0) or 0)
    new_status = _normalize_kanban_status((data.get("new_status") or "").strip())
    if not job_id:
        return jsonify({"ok": False, "error": "job_id eksik"}), 400
    if new_status not in KANBAN_COLUMNS:
        return jsonify({"ok": False, "error": "Gecersiz durum"}), 400

    j = Job.query.get(job_id)
    if not j:
        return jsonify({"ok": False, "error": "Is bulunamadi"}), 404

    old_status = _normalize_kanban_status(getattr(j, "kanban_status", None))
    if old_status == new_status:
        return jsonify({
            "ok": True,
            "changed": False,
            "status": old_status,
            "updated_at": j.updated_at.isoformat() if getattr(j, "updated_at", None) else None,
        })

    _set_job_kanban_status(j, new_status, changed_by_user_id=(current_user.id if current_user else None), note="board_move")
    db.session.commit()

    return jsonify({
        "ok": True,
        "changed": True,
        "status": new_status,
        "updated_at": j.updated_at.isoformat() if getattr(j, "updated_at", None) else None,
    })


@planner_bp.post("/api/board/jobs/publish")
@login_required
@planner_or_admin_required
def api_board_bulk_publish():
    data = request.get_json(force=True, silent=True) or {}
    job_ids_in = data.get("job_ids") or []
    if not isinstance(job_ids_in, list) or not job_ids_in:
        return jsonify({"ok": False, "error": "job_ids eksik"}), 400

    current_user = get_current_user()
    if not current_user:
        return jsonify({"ok": False, "error": "Yetkisiz"}), 401

    ids = sorted({int(x) for x in job_ids_in if int(x or 0) > 0})
    if not ids:
        return jsonify({"ok": False, "error": "job_ids eksik"}), 400

    now = datetime.now()
    updated = 0
    for jid in ids:
        j = Job.query.get(jid)
        if not j:
            continue
        cell = PlanCell.query.get(j.cell_id) if getattr(j, "cell_id", None) else None
        if not cell:
            continue
        _publish_cell(cell, publisher=current_user, now=now)
        updated += 1

    db.session.commit()
    return jsonify({"ok": True, "updated": updated})


@planner_bp.post("/api/board/jobs/close")
@login_required
@planner_or_admin_required
def api_board_bulk_close():
    data = request.get_json(force=True, silent=True) or {}
    job_ids_in = data.get("job_ids") or []
    if not isinstance(job_ids_in, list) or not job_ids_in:
        return jsonify({"ok": False, "error": "job_ids eksik"}), 400

    current_user = get_current_user()
    now = datetime.now()
    updated = 0
    for jid in sorted({int(x) for x in job_ids_in if int(x or 0) > 0}):
        j = Job.query.get(jid)
        if not j:
            continue
        if _normalize_kanban_status(getattr(j, "kanban_status", None)) == "CLOSED":
            continue
        try:
            j.closed_at = now
        except Exception:
            pass
        _promote_job_kanban_status(j, "CLOSED", changed_by_user_id=(current_user.id if current_user else None), note="board_close")
        updated += 1

    db.session.commit()
    return jsonify({"ok": True, "updated": updated})


@planner_bp.post("/api/board/jobs/reassign")
@login_required
@planner_or_admin_required
def api_board_bulk_reassign():
    data = request.get_json(force=True, silent=True) or {}
    job_ids_in = data.get("job_ids") or []
    assigned_user_id = int(data.get("assigned_user_id", 0) or 0)
    if not isinstance(job_ids_in, list) or not job_ids_in:
        return jsonify({"ok": False, "error": "job_ids eksik"}), 400
    if assigned_user_id <= 0:
        return jsonify({"ok": False, "error": "assigned_user_id eksik"}), 400
    user = User.query.get(assigned_user_id)
    if not user or not bool(getattr(user, "is_active", True)):
        return jsonify({"ok": False, "error": "Kullanici bulunamadi"}), 404

    current_user = get_current_user()
    updated = 0
    for jid in sorted({int(x) for x in job_ids_in if int(x or 0) > 0}):
        j = Job.query.get(jid)
        if not j:
            continue
        j.assigned_user_id = assigned_user_id
        try:
            cell = PlanCell.query.get(j.cell_id) if getattr(j, "cell_id", None) else None
            if cell:
                cell.assigned_user_id = assigned_user_id
                cell.updated_at = datetime.now()
        except Exception:
            pass
        _promote_job_kanban_status(j, "ASSIGNED", changed_by_user_id=(current_user.id if current_user else None), note="board_reassign")
        updated += 1

    db.session.commit()
    return jsonify({"ok": True, "updated": updated})


@planner_bp.get("/board/export.xlsx")
@login_required
@planner_or_admin_required
def board_export_xlsx():
    start, end = _parse_board_date_range_args()

    q_text = (request.args.get("q") or "").strip()
    assignee_q = (request.args.get("assignee_q") or "").strip()
    status = (request.args.get("status") or "").strip()
    project_id = int(request.args.get("project_id", 0) or 0)
    subproject_id = int(request.args.get("subproject_id", 0) or 0)
    city = (request.args.get("city") or "").strip()
    assigned_user_id = int(request.args.get("assigned_user_id", 0) or 0)
    overdue_only = _parse_bool_arg("overdue_only", default=False)
    unassigned_only = _parse_bool_arg("unassigned_only", default=False)
    published_only = _parse_bool_arg("published_only", default=False)

    ids_raw = (request.args.get("ids") or "").strip()
    ids = []
    if ids_raw:
        try:
            ids = [int(x) for x in re.split(r"[ ,;]+", ids_raw) if x.strip()]
        except Exception:
            ids = []
        ids = sorted({int(x) for x in ids if int(x or 0) > 0})

    upsert_jobs_for_range(start, end)

    base_q = _build_board_jobs_query(
        start=start,
        end=end,
        q_text=q_text,
        assignee_q=assignee_q,
        status=status,
        project_id=project_id,
        subproject_id=subproject_id,
        city=city,
        assigned_user_id=assigned_user_id,
        overdue_only=overdue_only,
        unassigned_only=unassigned_only,
        published_only=published_only,
    )
    if ids:
        base_q = base_q.filter(Job.id.in_(ids))

    rows = (
        base_q.with_entities(Job, Project, SubProject, User)
        .order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc(), Job.id.asc())
        .limit(20000)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append([
        "Job ID",
        "Date",
        "City",
        "Project Code",
        "Project Name",
        "Sub-Project",
        "Assigned User",
        "Team",
        "Kanban Status",
        "Published",
        "Published At",
        "Closed At",
    ])

    for j, p, sp, u in rows:
        assigned = (u.full_name or u.email) if u else ""
        ws.append([
            int(j.id),
            iso(j.work_date) if j.work_date else "",
            (p.region if p else ""),
            (p.project_code if p else ""),
            (p.project_name if p else ""),
            (sp.name if sp else ""),
            assigned,
            (j.team_name or ""),
            _normalize_kanban_status(getattr(j, "kanban_status", None)),
            "1" if bool(getattr(j, "is_published", False)) else "0",
            j.published_at.strftime("%Y-%m-%d %H:%M") if getattr(j, "published_at", None) else "",
            j.closed_at.strftime("%Y-%m-%d %H:%M") if getattr(j, "closed_at", None) else "",
        ])

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Jobs", len(rows)])
    by_status = {}
    for j, _, _, _ in rows:
        st = _normalize_kanban_status(getattr(j, "kanban_status", None))
        by_status[st] = by_status.get(st, 0) + 1
    for st, cnt in sorted(by_status.items(), key=lambda x: (KANBAN_STATUS_ORDER.get(x[0], 99), x[0])):
        ws2.append([f"Status: {st}", cnt])

    for wsht in [ws, ws2]:
        for c in range(1, wsht.max_column + 1):
            wsht.cell(row=1, column=c).font = Font(bold=True)
            wsht.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"board_export_{iso(start)}_{iso(end)}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@planner_bp.get("/board.xlsx")
@login_required
@planner_or_admin_required
def board_xlsx():
    start, end = _parse_date_range_args()
    upsert_jobs_for_range(start, end)

    q = Job.query.join(Project, Project.id == Job.project_id).filter(Job.work_date >= start, Job.work_date <= end)
    jobs = q.order_by(Job.work_date.asc(), Project.region.asc(), Project.project_code.asc()).all()
    job_ids = [j.id for j in jobs]

    latest_fb = {}
    if job_ids:
        rows = (
            JobFeedback.query
            .filter(JobFeedback.job_id.in_(job_ids), JobFeedback.outcome != None)
            .order_by(JobFeedback.job_id.asc(), JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
            .all()
        )
        for r in rows:
            if r.job_id not in latest_fb:
                latest_fb[r.job_id] = r

    wb = Workbook()
    ws = wb.active
    ws.title = "Board"
    ws.append([
        "Job ID",
        "Date",
        "City",
        "Project Code",
        "Project Name",
        "Sub-Project",
        "Team",
        "Assigned User",
        "Job Status",
        "Kanban",
        "Published At",
        "Last Report At",
        "Review Status",
    ])

    for j in jobs:
        p = j.project
        fb = latest_fb.get(j.id)
        assigned = ""
        try:
            assigned = (j.assigned_user.full_name or j.assigned_user.email) if j.assigned_user else ""
        except Exception:
            assigned = ""
        ws.append([
            int(j.id),
            iso(j.work_date) if j.work_date else "",
            (p.region if p else ""),
            (p.project_code if p else ""),
            (p.project_name if p else ""),
            (j.subproject.name if getattr(j, "subproject", None) else ""),
            j.team_name or "",
            assigned,
            j.status or "",
            _normalize_kanban_status(getattr(j, "kanban_status", None)),
            j.published_at.strftime("%Y-%m-%d %H:%M") if getattr(j, "published_at", None) else "",
            fb.submitted_at.strftime("%Y-%m-%d %H:%M") if (fb and fb.submitted_at) else "",
            (fb.review_status if fb else ""),
        ])

    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"board_{iso(start)}_{iso(end)}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@planner_bp.get("/api/job/<int:job_id>/latest_report")
@login_required
@planner_or_admin_required
def api_job_latest_report(job_id: int):
    j = Job.query.get_or_404(job_id)
    rep = (
        JobFeedback.query
        .filter(JobFeedback.job_id == j.id, JobFeedback.outcome != None)
        .order_by(JobFeedback.submitted_at.desc().nullslast(), JobFeedback.id.desc())
        .first()
    )
    if not rep:
        return jsonify({"ok": True, "found": False, "job_id": int(j.id)})

    media_items = []
    try:
        rows = list(getattr(rep, "media", None) or [])
    except Exception:
        rows = []

    for m in rows:
        fp = getattr(m, "file_path", None) or ""
        media_items.append({
            "id": int(m.id),
            "file_path": fp,
            "file_type": (m.file_type or "file"),
            "original_name": (m.original_name or ""),
            "uploaded_at": m.uploaded_at.isoformat() if getattr(m, "uploaded_at", None) else None,
            "view_url": url_for("view_file_inline", filename=fp) if fp else "",
            "download_url": url_for("download_file", filename=fp) if fp else "",
        })

    u = None
    try:
        if getattr(rep, "user_id", None):
            u = User.query.get(rep.user_id)
    except Exception:
        u = None

    payload = {
        "id": int(rep.id),
        "job_id": int(rep.job_id),
        "user": (u.full_name or u.email) if u else "",
        "submitted_at": rep.submitted_at.isoformat() if getattr(rep, "submitted_at", None) else None,
        "outcome": rep.outcome or "",
        "isdp_status": rep.isdp_status or "",
        "extra_work_text": rep.extra_work_text or "",
        "notes_text": rep.notes_text or "",
        "reviewed_at": rep.reviewed_at.isoformat() if getattr(rep, "reviewed_at", None) else None,
        "review_status": rep.review_status or "pending",
        "review_note": rep.review_note or "",
        "media": media_items,
    }
    return jsonify({"ok": True, "found": True, "report": payload})


# -------------------------------------------------------------------------
# ARVENTO SETTINGS PAGE
# -------------------------------------------------------------------------

@planner_bp.get("/arvento-ayarlari")
@login_required
@planner_or_admin_required
def arvento_settings_page():
    """Arvento araç takip ayarları sayfası"""
    return render_template("arvento_settings.html")


# -------------------------------------------------------------------------
# ANALYTICS EXPORT
# -------------------------------------------------------------------------


# ----------------------------------------------------------------
# NEW ANALYTICS MODULE (v4.0 - Real Data Only)
# ----------------------------------------------------------------


# -------------------------------------------------------------------------
# MAP ICON UPLOAD API
# -------------------------------------------------------------------------

@planner_bp.post("/api/map/upload-icon")
@login_required
def api_map_upload_icon():
    """Harita için özel simge yükle"""
    if 'icon' not in request.files:
        return jsonify({"ok": False, "error": "Dosya yüklenmedi."}), 400
    
    file = request.files['icon']
    icon_type = request.form.get('icon_type', 'custom')
    
    if file.filename == '':
        return jsonify({"ok": False, "error": "Dosya seçilmedi."}), 400
    
    # Dosya uzantısı kontrolü
    allowed_extensions = {'.png', '.svg', '.webp'}
    ext = os.path.splitext(file.filename.lower())[1]
    if ext not in allowed_extensions:
        return jsonify({"ok": False, "error": "Sadece PNG, SVG ve WEBP dosyaları yüklenebilir."}), 400
    
    # Boyut kontrolü (500KB max)
    file.seek(0, 2)  # End of file
    file_size = file.tell()
    file.seek(0)  # Back to start
    
    if file_size > 500 * 1024:
        return jsonify({"ok": False, "error": "Dosya boyutu 500KB'dan küçük olmalıdır."}), 400
    
    try:
        # Yükleme klasörünü oluştur
        upload_folder = os.path.join(current_app.static_folder, 'uploads', 'map_icons')
        os.makedirs(upload_folder, exist_ok=True)
        
        # Benzersiz dosya adı oluştur
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        safe_icon_type = re.sub(r'[^a-zA-Z0-9_]', '', icon_type)
        filename = f"{safe_icon_type}_{timestamp}{ext}"
        filepath = os.path.join(upload_folder, filename)
        
        # Dosyayı kaydet
        file.save(filepath)
        
        # URL oluştur
        icon_url = url_for('static', filename=f'uploads/map_icons/{filename}', _external=False)
        
        return jsonify({
            "ok": True,
            "icon_url": icon_url,
            "filename": filename
        })
        
    except Exception as e:
        current_app.logger.error(f"Icon upload error: {str(e)}")
        return jsonify({"ok": False, "error": "Dosya yüklenirken hata oluştu."}), 500

